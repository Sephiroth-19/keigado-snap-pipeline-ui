from __future__ import annotations

# ── Make onnxruntime-gpu find CUDA via PyTorch's bundled libraries ──
# Kept because some lazily-imported helpers (error endpoints) still pull torch.
import errno
import os
try:
    import torch as _torch
    _torch_lib = os.path.join(os.path.dirname(_torch.__file__), "lib")
    if os.path.isdir(_torch_lib):
        os.environ["PATH"] = _torch_lib + os.pathsep + os.environ.get("PATH", "")
    del _torch, _torch_lib
except Exception:
    pass
# ───────────────────────────────────────────────────────────────────

import json
import logging
import re
import shutil
import zipfile
from pathlib import Path
from typing import Any
from urllib.parse import quote

logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s │ %(name)s │ %(message)s",
    force=True,
)

from fastapi import Depends, FastAPI, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

from backend import auth, job_repo, s3_storage, storage
from backend.auth_routes import router as auth_router
from backend.db import init_db
from backend.excel_labels import CLUB_SHEET_LABELS, excel_label, translate_display_value
from backend.project_routes import router as project_router
from backend.teacher_jobs import router as teacher_router
from backend.preview_images import (
    IMAGE_EXTENSIONS,
    collect_snap_preview_images,
    image_media_type,
    list_preview_images,
    safe_resolve_preview_path,
)

ROOT = Path(__file__).resolve().parent.parent
FRONTEND_DIR = ROOT / "frontend"

DEFAULT_BEST_SHOT_COUNT = 25  # local copy to avoid importing the heavy snap module here

load_dotenv()
app = FastAPI(title="Snap Pipeline Local App")

# Origins allowed to call the API. The frontend is served from the same origin as the API,
# so the default needs no cross-origin grant. Set CORS_ALLOW_ORIGINS (comma-separated) for a
# split deployment (e.g. CloudFront). "*" stays available for local dev but should not ship.
_cors_env = os.getenv("CORS_ALLOW_ORIGINS", "").strip()
_cors_origins = [o.strip() for o in _cors_env.split(",") if o.strip()] if _cors_env else []
app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins or ["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def _on_startup() -> None:
    init_db()
    auth.seed_owner()


pipeline: SnapPipeline | Any | None = None
latest_summary: dict[str, Any] | None = None
latest_zip_path: Path | None = None
logger = logging.getLogger(__name__)


@app.exception_handler(Exception)
async def _unhandled_exception_handler(request: Request, exc: Exception):
    """Guarantee a JSON body for unhandled errors instead of Starlette's plain-text
    'Internal Server Error' — the frontend always does res.json(), so a non-JSON body
    surfaces as a cryptic 'Unexpected token' parse error. Disk-full (ENOSPC) is mapped to
    507 with an operator-facing message so it can be told apart from a generic 500."""
    if isinstance(exc, OSError) and exc.errno == errno.ENOSPC:
        logger.error("No space left on device handling %s %s", request.method, request.url.path)
        return JSONResponse(
            status_code=507,
            content={"detail": "サーバーの空き容量が不足しています。管理者にお問い合わせください。"},
        )
    logger.exception("Unhandled error on %s %s", request.method, request.url.path)
    return JSONResponse(
        status_code=500,
        content={"detail": "サーバー内部エラーが発生しました。"},
    )


# ──────────────────────────── authentication gate ───────────────────────────────────
# Every /api/* request must carry a valid bearer token, except the public login endpoint
# (and CORS preflight). The static frontend (index.html, login.html, assets) is not under
# /api/ and is served freely — it is not secret; the data behind the API is what's guarded.
# On success the resolved user is stashed on request.state.user for route dependencies.

PUBLIC_API_PATHS = {"/api/auth/login"}


@app.middleware("http")
async def auth_gate(request: Request, call_next):
    path = request.url.path
    if request.method == "OPTIONS" or not path.startswith("/api/") or path in PUBLIC_API_PATHS:
        return await call_next(request)

    # Prefer the Authorization header; fall back to an ?access_token= query param so that
    # top-level GETs that can't set headers — file downloads (window.open) and <img> preview
    # src — can still authenticate with the same bearer token.
    header = request.headers.get("Authorization", "")
    if header.startswith("Bearer "):
        token = header[len("Bearer "):].strip()
    else:
        token = request.query_params.get("access_token", "")
    if not token:
        return JSONResponse({"detail": "認証が必要です。"}, status_code=401)
    try:
        request.state.user = auth.authenticate_token(token)
    except HTTPException as exc:
        return JSONResponse({"detail": exc.detail}, status_code=exc.status_code)
    return await call_next(request)


# ──────────────────────────── shared job helpers ────────────────────────────────────

def _get_job_checked(
    job_id: str, pipeline: str, user: dict[str, Any] | None = None
) -> dict[str, Any]:
    job = job_repo.get_job(job_id)
    if not job or job["pipeline"] != pipeline:
        raise HTTPException(status_code=404, detail="Job not found")
    # Hide other users' jobs behind the same 404 so existence isn't leaked.
    if not job_repo.may_access(job, user):
        raise HTTPException(status_code=404, detail="Job not found")
    return job


def _status_payload(job: dict[str, Any]) -> dict[str, Any]:
    return {"job_id": job["id"], "status": job["status"], "error": job.get("error")}


def _result_or_202(job: dict[str, Any]):
    if job["status"] == "completed":
        return job["summary"] or {}
    message = job.get("error") or "Job not completed"
    return JSONResponse(
        status_code=202,
        content={"job_id": job["id"], "status": job["status"], "message": message},
    )


def _download_result_zip(job: dict[str, Any], filename: str):
    if not job.get("result_zip"):
        raise HTTPException(status_code=404, detail="No output zip available yet.")
    rel = job["result_zip"]
    # If S3 archival is on and the object is there, hand the client a short-lived
    # presigned URL; otherwise fall back to streaming from the local EBS volume.
    if s3_storage.enabled() and s3_storage.object_exists(rel):
        url = s3_storage.presigned_url(rel, filename=filename)
        if url:
            return RedirectResponse(url, status_code=307)
    path = storage.resolve(rel)
    if not path.exists():
        raise HTTPException(status_code=404, detail="Output zip not found.")
    return FileResponse(path, media_type="application/zip", filename=filename)


def _serve_job_file(prefix: str, *, media_type: str, filename: str | None = None):
    """Serve a job artifact by its storage prefix, preferring local EBS and falling back to
    S3. Once a job's local copy is evicted (it lives only in S3), previews/downloads keep
    working by redirecting to a short-lived presigned URL. 404 when present in neither."""
    path = storage.resolve(prefix)
    if path.is_file():
        return FileResponse(path, media_type=media_type, filename=filename)
    if s3_storage.enabled() and s3_storage.object_exists(prefix):
        url = s3_storage.presigned_url(prefix, filename=filename)
        if url:
            return RedirectResponse(url, status_code=307)
    raise HTTPException(status_code=404, detail="File not found")


def _output_dir(job: dict[str, Any]) -> Path:
    return storage.resolve(job["output_prefix"])


def _find_individual_output_file(output_dir: Path, filename: str) -> Path:
    """Locate a single output image by bare filename anywhere under the job's output dir."""
    if not filename or Path(filename).name != filename or any(sep in filename for sep in ("/", "\\")):
        raise HTTPException(status_code=400, detail="Invalid preview filename")
    if not output_dir.exists():
        raise HTTPException(status_code=404, detail="Individual output directory not found")
    output_root = output_dir.resolve()
    for candidate in output_root.rglob("*"):
        if not candidate.is_file() or candidate.name != filename:
            continue
        resolved = candidate.resolve()
        try:
            resolved.relative_to(output_root)
        except ValueError:
            continue
        return resolved
    raise HTTPException(status_code=404, detail="Preview file not found")


def _club_output_dir(output_dir: Path) -> Path:
    return output_dir / "Club_Output"


def _club_result_json(output_dir: Path) -> Path:
    return _club_output_dir(output_dir) / "club_result.json"


def _club_rank_label(rank: int | None) -> str:
    return f"本{rank:02d}" if rank else "除外"


def _parse_club_final_rank(value: Any, excluded: bool = False) -> int | None:
    if excluded:
        return None
    if value is None:
        return None
    raw = str(value).strip()
    if raw == "" or raw.lower() in {"exclude", "excluded", "not_use", "not use", "none", "ng"} or raw in {"除外", "未使用"}:
        return None
    raw = raw.replace("本", "")
    if not re.fullmatch(r"[1-9]\d{0,2}", raw):
        raise HTTPException(status_code=400, detail="final_rank must be 本1, 本2, 本3, or exclude.")
    return int(raw)


def _club_output_name(item: dict[str, Any], final_rank: int) -> str:
    original = Path(str(item.get("original_file") or "photo.jpg"))
    club_name = str(item.get("club_name") or "Club")
    shooting_date = str(item.get("shooting_date") or "00000000")
    return f"{club_name}_{shooting_date}_{original.stem}_本{final_rank:02d}{original.suffix.lower() or '.jpg'}"


def _load_club_result(output_dir: Path) -> dict[str, Any]:
    path = _club_result_json(output_dir)
    if not path.exists():
        raise HTTPException(status_code=404, detail="Club result not found.")
    return json.loads(path.read_text(encoding="utf-8"))


def _save_club_result(output_dir: Path, data: dict[str, Any]) -> None:
    _club_result_json(output_dir).write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _club_preview_path_for_url(path: Any) -> str:
    rel = str(path or "").strip()
    if not rel:
        return ""
    if rel.startswith("Club_Output/"):
        return rel
    return f"Club_Output/{rel}"


def _public_club_result(job_id: str, data: dict[str, Any]) -> dict[str, Any]:
    public = json.loads(json.dumps(data, ensure_ascii=False))
    for idx, item in enumerate(public.get("items") or [], 1):
        clean_rel = item.get("ranked_relative_path") or item.get("clean_relative_path")
        marked_rel = item.get("ranked_marked_relative_path") or item.get("marked_relative_path")
        item.setdefault("item_id", f"{item.get('club_name', '')}::{item.get('original_file', '')}::{idx}")
        if clean_rel:
            item["preview_url"] = f"/api/club/{job_id}/preview-file?path={quote(_club_preview_path_for_url(clean_rel), safe='')}"
            item["thumbnail_url"] = item["preview_url"]
        if marked_rel:
            item["marked_preview_url"] = f"/api/club/{job_id}/preview-file?path={quote(_club_preview_path_for_url(marked_rel), safe='')}"
    public["job_id"] = job_id
    public["excel_url"] = f"/api/club/{job_id}/excel"
    public["output_zip_url"] = f"/api/club/{job_id}/download"
    return public


def _write_club_excel_from_manifest(club_output_dir: Path, data: dict[str, Any]) -> None:
    items = list(data.get("items") or [])
    included = [x for x in items if not x.get("excluded") and x.get("final_rank")]
    club_names = {str(x.get("club_name") or "") for x in items}

    wb = Workbook()
    ws = wb.active
    ws.title = CLUB_SHEET_LABELS["Summary"]
    ws.append(
        [
            excel_label("club_count"),
            excel_label("photo_count"),
            excel_label("closed_eye_photo_count"),
            excel_label("closed_eye_face_count"),
            excel_label("ranked_output_count"),
        ]
    )
    ws.append(
        [
            len(club_names),
            len(items),
            sum(1 for x in items if x.get("eyes_closed_photo")),
            sum(int(x.get("closed_eye_face_count") or 0) for x in items),
            len(included),
        ]
    )

    ews = wb.create_sheet(CLUB_SHEET_LABELS["Eye Closure Summary"])
    ews.append([excel_label("club"), excel_label("file_name"), excel_label("person_count"), excel_label("closed_eye_faces"), excel_label("eyes_closed_photo")])
    for item in items:
        ews.append(
            [
                item.get("club_name"),
                item.get("original_file"),
                item.get("person_count"),
                item.get("closed_eye_face_count"),
                translate_display_value(item.get("eyes_closed_photo")),
            ]
        )

    fws = wb.create_sheet(CLUB_SHEET_LABELS["Face Detail"])
    fws.append([excel_label("club"), excel_label("file_name"), excel_label("face_index"), excel_label("bbox"), excel_label("left_eye_ratio"), excel_label("right_eye_ratio"), excel_label("eye_closed")])

    rws = wb.create_sheet(CLUB_SHEET_LABELS["Best Shot Ranking"])
    rws.append(
        [
            excel_label("club"),
            "AI順位",
            "最終順位",
            excel_label("file_name"),
            excel_label("formality"),
            excel_label("beauty_score"),
            excel_label("expression_score"),
            excel_label("emotion_score"),
            excel_label("people_count_score"),
            excel_label("gesture_expression_penalty"),
            excel_label("is_ng"),
            excel_label("ng_reason"),
            excel_label("short_comment"),
            excel_label("total_score"),
        ]
    )
    for item in sorted(items, key=lambda x: (str(x.get("club_name") or ""), int(x.get("final_rank") or 9999), str(x.get("original_file") or ""))):
        ev = item.get("evaluation") or {}
        rws.append(
            [
                item.get("club_name"),
                item.get("ai_rank") or item.get("rank"),
                item.get("final_rank_label") or _club_rank_label(item.get("final_rank")),
                item.get("original_file"),
                ev.get("formality_score"),
                ev.get("beauty_score"),
                ev.get("expression_score"),
                ev.get("emotion_score"),
                ev.get("people_count_score"),
                ev.get("gesture_expression_penalty"),
                translate_display_value(item.get("ng_flag")),
                translate_display_value(item.get("ng_reason")),
                translate_display_value(ev.get("short_comment")),
                item.get("total_score"),
            ]
        )

    nws = wb.create_sheet(CLUB_SHEET_LABELS["Rename Output"])
    nws.append([excel_label("club"), excel_label("original_file"), excel_label("renamed_file"), excel_label("shooting_date"), "AI順位", "最終順位", "出力対象"])
    for item in sorted(items, key=lambda x: (str(x.get("club_name") or ""), int(x.get("final_rank") or 9999), str(x.get("original_file") or ""))):
        nws.append(
            [
                item.get("club_name"),
                item.get("original_file"),
                item.get("final_renamed_file") or "",
                item.get("shooting_date"),
                item.get("ai_rank") or item.get("rank"),
                item.get("final_rank_label") or _club_rank_label(item.get("final_rank")),
                "いいえ" if item.get("excluded") else "はい",
            ]
        )

    ngs = wb.create_sheet("NG写真・要確認")
    ngs.append([excel_label("club"), excel_label("file_name"), excel_label("ng_flag"), excel_label("reason")])
    for item in items:
        if item.get("ng_flag") or item.get("excluded"):
            reason = item.get("ng_reason") or ("manual exclude" if item.get("excluded") else "")
            ngs.append([item.get("club_name"), item.get("original_file"), translate_display_value(item.get("ng_flag")), translate_display_value(reason)])
    wb.save(club_output_dir / "club_result.xlsx")


def _zip_club_output(output_dir: Path, output_zip: Path) -> Path:
    club_output_dir = _club_output_dir(output_dir)
    if output_zip.exists():
        output_zip.unlink()
    output_zip.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(output_zip, "w", zipfile.ZIP_DEFLATED) as zf:
        allowed = {"ranked_photos", "ranked_photos_marked", "club_result.xlsx"}
        for p in club_output_dir.rglob("*"):
            if not p.is_file():
                continue
            rel = p.relative_to(club_output_dir)
            if rel.parts and rel.parts[0] in allowed:
                zf.write(p, p.relative_to(club_output_dir.parent))
    return output_zip


def _rebuild_club_outputs(output_dir: Path, output_zip: Path, data: dict[str, Any]) -> None:
    club_output_dir = _club_output_dir(output_dir)
    ranked = club_output_dir / "ranked_photos"
    ranked_marked = club_output_dir / "ranked_photos_marked"
    for folder in (ranked, ranked_marked):
        if folder.exists():
            shutil.rmtree(folder)
        folder.mkdir(parents=True, exist_ok=True)

    for item in data.get("items") or []:
        final_rank = item.get("final_rank")
        if item.get("excluded") or not final_rank:
            item["excluded"] = True
            item["final_rank"] = None
            item["final_rank_label"] = _club_rank_label(None)
            item["status"] = "Excluded"
            item["final_renamed_file"] = ""
            item["ranked_relative_path"] = ""
            item["ranked_marked_relative_path"] = ""
            continue

        final_rank = int(final_rank)
        club_name = str(item.get("club_name") or "Club")
        renamed = _club_output_name(item, final_rank)
        clean_rel = item.get("clean_relative_path") or (Path("clean_images") / club_name / str(item.get("original_file") or "")).as_posix()
        marked_rel = item.get("marked_relative_path") or (Path("marked_images") / club_name / str(item.get("original_file") or "")).as_posix()
        clean_src = safe_resolve_preview_path(club_output_dir, str(clean_rel))
        marked_src = safe_resolve_preview_path(club_output_dir, str(marked_rel))
        clean_dst = ranked / club_name / renamed
        marked_dst = ranked_marked / club_name / renamed
        clean_dst.parent.mkdir(parents=True, exist_ok=True)
        marked_dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(clean_src, clean_dst)
        shutil.copy2(marked_src, marked_dst)

        item["excluded"] = False
        item["final_rank"] = final_rank
        item["final_rank_label"] = _club_rank_label(final_rank)
        item["status"] = "Best Shot" if final_rank == 1 else "Passing"
        item["final_renamed_file"] = renamed
        item["renamed_file"] = renamed
        item["ranked_relative_path"] = clean_dst.relative_to(club_output_dir).as_posix()
        item["ranked_marked_relative_path"] = marked_dst.relative_to(club_output_dir).as_posix()

    _write_club_excel_from_manifest(club_output_dir, data)
    _save_club_result(output_dir, data)
    _zip_club_output(output_dir, output_zip)


def _add_individual_preview_urls(manifest: dict[str, Any], job_id: str) -> dict[str, Any]:
    """Inject per-file ``preview_url`` entries into a manifest so the UI can show thumbnails."""
    classes = manifest.get("classes")
    if not isinstance(classes, dict):
        return manifest
    for class_data in classes.values():
        entries = class_data.get("entries") if isinstance(class_data, dict) else None
        if not isinstance(entries, list):
            continue
        for entry in entries:
            files = entry.get("files") if isinstance(entry, dict) else None
            if not isinstance(files, dict):
                continue
            for tag, value in list(files.items()):
                if isinstance(value, str):
                    filename = value
                    files[tag] = {
                        "filename": filename,
                        "preview_url": f"/api/individual/{job_id}/preview/{quote(filename, safe='')}",
                    }
                elif isinstance(value, dict):
                    filename = value.get("filename") or value.get("name")
                    if filename:
                        value.setdefault("filename", filename)
                        value.setdefault(
                            "preview_url",
                            f"/api/individual/{job_id}/preview/{quote(str(filename), safe='')}",
                        )
    return manifest


# ════════════════════════════════════ SNAP ══════════════════════════════════════════

def _parse_snap_best_shot_count(value: str | None) -> int | None:
    if value is None or value == "":
        return None
    raw = value.strip()
    if not re.fullmatch(r"[1-9]\d*", raw):
        raise HTTPException(status_code=400, detail="Best Shot Count must be an integer between 1 and 200.")
    parsed = int(raw)
    if parsed > 200:
        raise HTTPException(status_code=400, detail="Best Shot Count must be an integer between 1 and 200.")
    return parsed


@app.post("/api/snap/run")
async def run_snap_pipeline(
    images: list[UploadFile] | None = File(default=None),
    folder_zip: UploadFile | None = File(default=None),
    best_shot_count: str | None = Form(default=None),
    project_id: str | None = Form(default=None),
    current: dict[str, Any] = Depends(auth.get_current_user),
) -> dict[str, Any]:
    if not images and not folder_zip:
        raise HTTPException(status_code=400, detail="Upload image files or one zip folder.")
    parsed_best_shot_count = _parse_snap_best_shot_count(best_shot_count)

    job_id = job_repo.new_job_id("snap")
    in_dir, _ = storage.ensure_job_dirs(job_id)

    if images:
        for image in images:
            await storage.save_upload_async(image, in_dir / image.filename)
    if folder_zip:
        if not (folder_zip.filename or "").lower().endswith(".zip"):
            raise HTTPException(status_code=400, detail="folder_zip must be a .zip file.")
        await storage.save_upload_async(folder_zip, in_dir / "folder.zip")

    job_repo.create_job(
        "snap",
        input_prefix=storage.input_prefix(job_id),
        output_prefix=storage.output_prefix(job_id),
        options={"best_shot_count": parsed_best_shot_count},
        project_id=project_id or None,
        user_id=current.get("id"),
        job_id=job_id,
    )
    return {
        "job_id": job_id,
        "status": "queued",
        "status_url": f"/api/snap/status/{job_id}",
        "result_url": f"/api/snap/result/{job_id}",
        "download_url": f"/api/snap/download/{job_id}",
    }


@app.get("/api/snap/status/{job_id}")
def snap_status(
    job_id: str, current: dict[str, Any] = Depends(auth.get_current_user)
) -> dict[str, Any]:
    return _status_payload(_get_job_checked(job_id, "snap", current))


@app.get("/api/snap/result/{job_id}")
def snap_result(job_id: str, current: dict[str, Any] = Depends(auth.get_current_user)):
    return _result_or_202(_get_job_checked(job_id, "snap", current))


@app.get("/api/snap/download/{job_id}")
def snap_download(
    job_id: str, current: dict[str, Any] = Depends(auth.get_current_user)
) -> FileResponse:
    job = _get_job_checked(job_id, "snap", current)
    return _download_result_zip(job, "snap_outputs.zip")


@app.get("/api/snap/{job_id}/preview-images")
def snap_preview_images(
    job_id: str,
    mode: str = Query(default="final", pattern="^(final|similarity|candidates|all)$"),
    event_id: str | None = Query(default=None),
    event_name: str | None = Query(default=None),
    bucket: str | None = Query(
        default=None,
        pattern="^(final_selected|best|bestshot|other_passing|passing|ng_photos|ng|similarity_clusters|similarity)$",
    ),
    current: dict[str, Any] = Depends(auth.get_current_user),
) -> dict[str, Any]:
    job = _get_job_checked(job_id, "snap", current)
    images = collect_snap_preview_images(
        _output_dir(job),
        mode=mode,
        event_id=event_id,
        event_name=event_name,
        bucket=bucket,
        preview_base=f"/api/snap/{job_id}/preview-file",
    )
    return {
        "status": "ok",
        "mode": mode,
        "event_id": event_id,
        "event_name": event_name,
        "bucket": bucket,
        "count": len(images),
        "images": images,
    }


@app.get("/api/snap/{job_id}/preview-file")
def snap_preview_file(
    job_id: str, path: str, current: dict[str, Any] = Depends(auth.get_current_user)
) -> FileResponse:
    job = _get_job_checked(job_id, "snap", current)
    output_root = _output_dir(job).resolve()
    target = (output_root / path).resolve()
    try:
        target.relative_to(output_root)
    except ValueError:
        raise HTTPException(status_code=404, detail="Preview file not found.") from None
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="Preview file not found.")
    return FileResponse(target)


def _excel_sort_key(path: Path) -> tuple[int, str]:
    name = path.name.lower()
    if name == "snap_pipeline_report.xlsx":
        return (0, path.as_posix())
    if name == "all_events_summary.xlsx":
        return (1, path.as_posix())
    return (2, path.as_posix())


def _cell_to_json(value: Any) -> Any:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _preview_excel_file(path: Path, output_dir: Path, row_limit: int = 100) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets: list[dict[str, Any]] = []
        for sheet in workbook.worksheets:
            rows_iter = sheet.iter_rows(values_only=True)
            first_row = next(rows_iter, None)
            headers = [_cell_to_json(v) for v in first_row] if first_row else []
            rows: list[list[Any]] = []
            for idx, row in enumerate(rows_iter):
                if idx >= row_limit:
                    break
                rows.append([_cell_to_json(v) for v in row])
            sheets.append({
                "sheet_name": sheet.title,
                "headers": headers,
                "rows": rows,
                "row_count": sheet.max_row or 0,
                "column_count": sheet.max_column or 0,
            })
        return {
            "file_name": path.name,
            "relative_path": path.relative_to(output_dir).as_posix(),
            "sheets": sheets,
        }
    finally:
        workbook.close()


@app.get("/api/snap/{job_id}/preview-excel")
def snap_preview_excel(
    job_id: str, current: dict[str, Any] = Depends(auth.get_current_user)
) -> dict[str, Any]:
    job = _get_job_checked(job_id, "snap", current)
    output_dir = _output_dir(job)
    no_files = {"status": "ok", "files": [], "message": "プレビュー可能なExcelファイルが見つかりませんでした。"}
    if not output_dir.exists():
        return no_files
    excel_files = sorted((p for p in output_dir.rglob("*.xlsx") if p.is_file()), key=_excel_sort_key)
    if not excel_files:
        return no_files
    return {"status": "ok", "files": [_preview_excel_file(p, output_dir) for p in excel_files]}


# ═════════════════════════════════ INDIVIDUAL ═══════════════════════════════════════

@app.post("/api/individual/run")
async def run_individual(
    photos_zip: UploadFile = File(...),
    roster_file: UploadFile | None = File(default=None),
    frame_config_file: UploadFile | None = File(default=None),
    school_name: str = Form(default=""),
    year: str = Form(default=""),
    scoring: str = Form(default="local"),
    project_id: str | None = Form(default=None),
    current: dict[str, Any] = Depends(auth.get_current_user),
) -> dict[str, Any]:
    if not photos_zip.filename or not photos_zip.filename.lower().endswith(".zip"):
        raise HTTPException(status_code=400, detail="photos_zip must be .zip")

    missing_fields: list[str] = []
    if not school_name or not school_name.strip():
        missing_fields.append("school_name")
    if not year or not year.strip():
        missing_fields.append("year")
    if missing_fields:
        return {
            "status": "error",
            "reason": "missing_required_form_field",
            "missing_fields": missing_fields,
            "received_school_name": school_name,
            "received_year": year,
            "received_scoring": scoring,
        }

    job_id = job_repo.new_job_id("individual")
    in_dir, _ = storage.ensure_job_dirs(job_id)

    await storage.save_upload_async(photos_zip, in_dir / "photos.zip")

    roster_filename = None
    if roster_file and roster_file.filename:
        roster_filename = roster_file.filename
        await storage.save_upload_async(roster_file, in_dir / roster_filename)

    has_frame_config = False
    if frame_config_file and frame_config_file.filename:
        if not frame_config_file.filename.lower().endswith(".json"):
            raise HTTPException(status_code=400, detail="frame_config_file must be .json")
        await storage.save_upload_async(frame_config_file, in_dir / "frame_config.json")
        has_frame_config = True

    job_repo.create_job(
        "individual",
        input_prefix=storage.input_prefix(job_id),
        output_prefix=storage.output_prefix(job_id),
        options={
            "school_name": school_name,
            "year": year,
            "scoring": scoring,
            "roster_filename": roster_filename,
            "has_frame_config": has_frame_config,
        },
        project_id=project_id or None,
        user_id=current.get("id"),
        job_id=job_id,
    )
    return {
        "job_id": job_id,
        "status": "queued",
        "status_url": f"/api/individual/{job_id}/status",
        "result_url": f"/api/individual/{job_id}/result",
        "download_url": f"/api/individual/{job_id}/download",
        "received_school_name": school_name,
        "received_year": year,
        "received_scoring": scoring,
    }


@app.get("/api/individual/{job_id}/status")
def individual_status(
    job_id: str, current: dict[str, Any] = Depends(auth.get_current_user)
) -> dict[str, Any]:
    return _status_payload(_get_job_checked(job_id, "individual", current))


@app.get("/api/individual/{job_id}/result")
def get_individual_result(
    job_id: str, current: dict[str, Any] = Depends(auth.get_current_user)
):
    job = _get_job_checked(job_id, "individual", current)
    if job["status"] != "completed":
        return _result_or_202(job)
    # On completion return the on-disk manifest (carries per-class entries/files) enriched
    # with preview URLs so the review UI can render real thumbnails. Fall back to the stored
    # summary if a manifest was never written.
    manifest_path = _output_dir(job) / "manifest.json"
    if not manifest_path.exists():
        return job["summary"] or {}
    data = json.loads(manifest_path.read_text(encoding="utf-8"))
    # The manifest carries per-class entries; the review-summary counts live under "summary"
    # (embedded by the pipeline) or fall back to the stored job summary. Expose them at the top
    # level so the review UI shows real counts instead of 0/-, without clobbering manifest keys.
    summary = data.get("summary") if isinstance(data.get("summary"), dict) else (job.get("summary") or {})
    merged = {**(summary or {}), **data}
    return _add_individual_preview_urls(merged, job_id)


@app.get("/api/individual/{job_id}/preview/{filename}")
def preview_individual_output(
    job_id: str, filename: str, current: dict[str, Any] = Depends(auth.get_current_user)
) -> FileResponse:
    job = _get_job_checked(job_id, "individual", current)
    image_path = _find_individual_output_file(_output_dir(job), filename)
    return FileResponse(image_path, filename=image_path.name)


@app.get("/api/individual/{job_id}/preview-images")
def get_individual_preview_images(
    job_id: str, current: dict[str, Any] = Depends(auth.get_current_user)
) -> dict[str, Any]:
    job = _get_job_checked(job_id, "individual", current)
    images = list_preview_images(
        _output_dir(job), f"/api/individual/{job_id}/preview-image", "Individual Photo"
    )
    return {"job_id": job_id, "count": len(images), "images": images}


@app.get("/api/individual/{job_id}/preview-image")
def get_individual_preview_image(
    job_id: str, path: str, current: dict[str, Any] = Depends(auth.get_current_user)
) -> FileResponse:
    job = _get_job_checked(job_id, "individual", current)
    image_path = safe_resolve_preview_path(_output_dir(job), path)
    return FileResponse(image_path, media_type=image_media_type(image_path), filename=image_path.name)


@app.get("/api/individual/{job_id}/download")
def download_individual_output(
    job_id: str, current: dict[str, Any] = Depends(auth.get_current_user)
) -> FileResponse:
    job = _get_job_checked(job_id, "individual", current)
    return _download_result_zip(job, "individual_output.zip")


@app.get("/api/individual/{job_id}/errors")
def get_individual_errors(
    job_id: str, current: dict[str, Any] = Depends(auth.get_current_user)
) -> dict[str, Any]:
    # Lazy import: pulls the heavy pipeline module only when an operator opens the queue.
    from backend.individual_pipeline import filter_unresolved_error_queue, load_error_resolutions

    job = _get_job_checked(job_id, "individual", current)
    out = _output_dir(job)
    queue_path = out / "error_queue.json"
    queue = json.loads(queue_path.read_text(encoding="utf-8")) if queue_path.exists() else []
    resolutions = load_error_resolutions(out / "error_resolutions.json")
    return {"errors": filter_unresolved_error_queue(queue, resolutions)}


@app.post("/api/individual/{job_id}/resolve")
def resolve_individual_errors(
    job_id: str,
    payload: dict[str, Any],
    current: dict[str, Any] = Depends(auth.get_current_user),
) -> dict[str, Any]:
    from backend.individual_pipeline import load_error_resolutions, save_error_resolutions

    job = _get_job_checked(job_id, "individual", current)
    out = _output_dir(job)
    out.mkdir(parents=True, exist_ok=True)
    path = out / "error_resolutions.json"
    existing = load_error_resolutions(path)
    existing.update(payload.get("resolutions", {}))
    save_error_resolutions(path, existing)
    return {"status": "ok", "count": len(existing)}


@app.post("/api/individual/{job_id}/reexport")
def reexport_individual(
    job_id: str, current: dict[str, Any] = Depends(auth.get_current_user)
) -> dict[str, Any]:
    """Re-run the export against resolved errors — enqueued, reusing the same job dirs."""
    job = _get_job_checked(job_id, "individual", current)
    new_id = job_repo.new_job_id("individual")
    job_repo.create_job(
        "individual",
        input_prefix=job["input_prefix"],   # reuse already-extracted photos
        output_prefix=job["output_prefix"],  # reuse output dir (holds error_resolutions.json)
        options={**(job.get("options") or {}), "reexport": True},
        project_id=job.get("project_id"),
        user_id=job.get("user_id") or current.get("id"),
        job_id=new_id,
    )
    return {
        "job_id": new_id,
        "status": "queued",
        "status_url": f"/api/individual/{new_id}/status",
        "result_url": f"/api/individual/{new_id}/result",
    }


_OVERRIDES_FILENAME = "review_overrides.json"


@app.get("/api/individual/{job_id}/overrides")
def get_individual_overrides(
    job_id: str, current: dict[str, Any] = Depends(auth.get_current_user)
) -> dict[str, Any]:
    """Return the operator's saved review edits so a page reload restores them."""
    job = _get_job_checked(job_id, "individual", current)
    path = _output_dir(job) / _OVERRIDES_FILENAME
    if not path.exists():
        return {"files": {}}
    return json.loads(path.read_text(encoding="utf-8"))


@app.post("/api/individual/{job_id}/overrides")
def save_individual_overrides(
    job_id: str,
    payload: dict[str, Any],
    current: dict[str, Any] = Depends(auth.get_current_user),
) -> dict[str, Any]:
    """Persist review-screen rename/tag/status edits. Survives reloads; applied on re-export."""
    job = _get_job_checked(job_id, "individual", current)
    out = _output_dir(job)
    out.mkdir(parents=True, exist_ok=True)
    files = payload.get("files")
    if not isinstance(files, dict):
        raise HTTPException(status_code=400, detail="Body must contain a 'files' object.")
    path = out / _OVERRIDES_FILENAME
    existing = json.loads(path.read_text(encoding="utf-8")) if path.exists() else {"files": {}}
    existing.setdefault("files", {}).update(files)
    path.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"status": "ok", "count": len(existing["files"])}


@app.post("/api/individual/{job_id}/apply-overrides")
def apply_individual_overrides(
    job_id: str, current: dict[str, Any] = Depends(auth.get_current_user)
) -> dict[str, Any]:
    """Enqueue a fast job that applies saved review edits to the output and re-zips it."""
    job = _get_job_checked(job_id, "individual", current)
    new_id = job_repo.new_job_id("individual")
    job_repo.create_job(
        "individual",
        input_prefix=job["input_prefix"],
        output_prefix=job["output_prefix"],
        options={**(job.get("options") or {}), "apply_overrides_only": True},
        project_id=job.get("project_id"),
        user_id=job.get("user_id") or current.get("id"),
        job_id=new_id,
    )
    return {
        "job_id": new_id,
        "status": "queued",
        "status_url": f"/api/individual/{new_id}/status",
        "result_url": f"/api/individual/{new_id}/result",
    }


# ════════════════════════════════════ CLUB ══════════════════════════════════════════

@app.post("/api/club/run")
async def run_club(
    folder_zip: UploadFile = File(...),
    project_id: str | None = Form(default=None),
    current: dict[str, Any] = Depends(auth.get_current_user),
) -> dict[str, Any]:
    if not folder_zip.filename or not folder_zip.filename.lower().endswith(".zip"):
        raise HTTPException(status_code=400, detail="Please upload one .zip file.")

    job_id = job_repo.new_job_id("club")
    in_dir, _ = storage.ensure_job_dirs(job_id)
    await storage.save_upload_async(folder_zip, in_dir / "folder.zip")

    job_repo.create_job(
        "club",
        input_prefix=storage.input_prefix(job_id),
        output_prefix=storage.output_prefix(job_id),
        options={},
        project_id=project_id or None,
        user_id=current.get("id"),
        job_id=job_id,
    )
    return {
        "job_id": job_id,
        "status": "queued",
        "status_url": f"/api/club/{job_id}/status",
        "result_url": f"/api/club/{job_id}/result",
        "download_url": f"/api/club/{job_id}/download",
        "excel_url": f"/api/club/{job_id}/excel",
    }


@app.get("/api/club/{job_id}/status")
def club_status(
    job_id: str, current: dict[str, Any] = Depends(auth.get_current_user)
) -> dict[str, Any]:
    return _status_payload(_get_job_checked(job_id, "club", current))


@app.get("/api/club/{job_id}/result")
def club_result(job_id: str, current: dict[str, Any] = Depends(auth.get_current_user)):
    job = _get_job_checked(job_id, "club", current)
    if job["status"] != "completed":
        return _result_or_202(job)
    result_path = _club_result_json(_output_dir(job))
    if result_path.exists():
        data = _load_club_result(_output_dir(job))
        data["status"] = "completed"
        data.setdefault("summary", job["summary"] or {})
        return _public_club_result(job_id, data)
    return _result_or_202(job)


@app.post("/api/club/{job_id}/adjust-ranks")
def adjust_club_ranks(
    job_id: str,
    payload: dict[str, Any],
    current: dict[str, Any] = Depends(auth.get_current_user),
) -> dict[str, Any]:
    job = _get_job_checked(job_id, "club", current)
    if job["status"] != "completed":
        raise HTTPException(status_code=409, detail="Club job is not completed yet.")
    output_dir = _output_dir(job)
    data = _load_club_result(output_dir)
    adjustments = payload.get("adjustments")
    if not isinstance(adjustments, list):
        raise HTTPException(status_code=400, detail="adjustments must be a list.")

    by_key: dict[tuple[str, str], dict[str, Any]] = {}
    for item in data.get("items") or []:
        by_key[(str(item.get("club_name") or ""), str(item.get("original_file") or ""))] = item

    for adjustment in adjustments:
        if not isinstance(adjustment, dict):
            continue
        key = (str(adjustment.get("club_name") or ""), str(adjustment.get("original_file") or ""))
        item = by_key.get(key)
        if item is None:
            raise HTTPException(status_code=400, detail=f"Unknown club photo: {key[0]} / {key[1]}")
        final_rank = _parse_club_final_rank(
            adjustment.get("final_rank"),
            bool(adjustment.get("excluded")),
        )
        item["final_rank"] = final_rank
        item["excluded"] = final_rank is None

    zip_path = storage.resolve(job.get("result_zip") or f"jobs/{job_id}/club_output.zip")
    _rebuild_club_outputs(output_dir, zip_path, data)
    refreshed = _load_club_result(output_dir)
    refreshed["status"] = "completed"
    refreshed.setdefault("summary", job["summary"] or {})
    return _public_club_result(job_id, refreshed)


@app.get("/api/club/{job_id}/excel")
def download_club_excel(
    job_id: str, current: dict[str, Any] = Depends(auth.get_current_user)
) -> FileResponse:
    job = _get_job_checked(job_id, "club", current)
    excel_path = _output_dir(job) / "Club_Output" / "club_result.xlsx"
    if not excel_path.exists():
        raise HTTPException(status_code=404, detail="Club Excel not found.")
    return FileResponse(
        excel_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="club_result.xlsx",
    )


@app.get("/api/club/{job_id}/download")
def download_club_output(
    job_id: str, current: dict[str, Any] = Depends(auth.get_current_user)
) -> FileResponse:
    job = _get_job_checked(job_id, "club", current)
    return _download_result_zip(job, "club_output.zip")


@app.get("/api/club/{job_id}/preview-file")
def club_preview_file(
    job_id: str, path: str, current: dict[str, Any] = Depends(auth.get_current_user)
) -> FileResponse:
    """Serve a single club output image (e.g. ranked_photos_marked/<club>/<file>)."""
    job = _get_job_checked(job_id, "club", current)
    image_path = safe_resolve_preview_path(_output_dir(job), path)
    return FileResponse(image_path, media_type=image_media_type(image_path))


# ════════════════════════════════════ AI DETECTION ══════════════════════════════════
# Standalone "AI 検出チェック": people-count / eye-closure on ad-hoc uploads. Runs as a
# queued "ai" job (the worker owns the heavy ML deps), so the flow mirrors the pipelines:
# run → poll status → fetch result → show per-image rows.

@app.post("/api/ai/run")
async def run_ai_detect(
    files: list[UploadFile] = File(...),
    eye: bool = Form(default=True),
    count: bool = Form(default=True),
    project_id: str | None = Form(default=None),
    current: dict[str, Any] = Depends(auth.get_current_user),
) -> dict[str, Any]:
    if not files:
        raise HTTPException(status_code=400, detail="画像を1枚以上アップロードしてください。")
    if not eye and not count:
        raise HTTPException(status_code=400, detail="目閉じ検出または人数確認を1つ以上選択してください。")

    job_id = job_repo.new_job_id("ai")
    in_dir, _ = storage.ensure_job_dirs(job_id)
    saved = 0
    for upload in files:
        name = Path(upload.filename or "").name
        if not name or Path(name).suffix.lower() not in {
            ".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff", ".heic", ".heif", ".zip",
        }:
            continue
        await storage.save_upload_async(upload, in_dir / name)
        saved += 1
    if saved == 0:
        raise HTTPException(status_code=400, detail="対応する画像ファイルが見つかりませんでした。")

    job_repo.create_job(
        "ai",
        input_prefix=storage.input_prefix(job_id),
        output_prefix=storage.output_prefix(job_id),
        options={"eye": bool(eye), "count": bool(count)},
        project_id=project_id or None,
        user_id=current.get("id"),
        job_id=job_id,
    )
    return {
        "job_id": job_id,
        "status": "queued",
        "status_url": f"/api/ai/{job_id}/status",
        "result_url": f"/api/ai/{job_id}/result",
    }


@app.get("/api/ai/{job_id}/status")
def ai_status(
    job_id: str, current: dict[str, Any] = Depends(auth.get_current_user)
) -> dict[str, Any]:
    return _status_payload(_get_job_checked(job_id, "ai", current))


@app.get("/api/ai/{job_id}/result")
def ai_result(job_id: str, current: dict[str, Any] = Depends(auth.get_current_user)):
    return _result_or_202(_get_job_checked(job_id, "ai", current))


@app.get("/api/ai/{job_id}/preview-file")
def ai_preview_file(
    job_id: str, path: str, current: dict[str, Any] = Depends(auth.get_current_user)
) -> FileResponse:
    job = _get_job_checked(job_id, "ai", current)
    image_path = safe_resolve_preview_path(_output_dir(job), path)
    return FileResponse(image_path, media_type=image_media_type(image_path))


# ════════════════════════════════════ PERSON SEARCH ═════════════════════════════════

@app.post("/api/person/search")
def person_search(
    body: dict[str, Any], current: dict[str, Any] = Depends(auth.get_current_user)
) -> dict[str, Any]:
    from backend import person_search as person_search_mod

    project_id = (body or {}).get("project_id")
    if not project_id:
        raise HTTPException(status_code=400, detail="プロジェクトを選択してください。")
    if (body or {}).get("mode") == "photo":
        raise HTTPException(
            status_code=400,
            detail="参照写真・PDF検索は /api/person/search/query に添付して実行してください。",
        )
    return person_search_mod.search(
        project_id=project_id,
        keyword=(body or {}).get("keyword") or "",
        targets=(body or {}).get("targets") or [],
        scopes=(body or {}).get("scopes") or [],
    )


@app.post("/api/person/search/query")
async def person_search_query(
    project_id: str = Form(...),
    scopes: str = Form(""),
    class_name: str = Form(""),
    class_number: str = Form(""),
    teacher_name: str = Form(""),
    top_k: int = Form(200),
    photo: UploadFile | None = File(default=None),
    ref_pdf: UploadFile | None = File(default=None),
    album: UploadFile | None = File(default=None),
    current: dict[str, Any] = Depends(auth.get_current_user),
) -> dict[str, Any]:
    """Combined reference-photo / PDF / album / text query.

    Saves the uploads under a new ``face_query`` job and queues it; the worker builds the
    query embeddings (local InsightFace + fitz) and runs the combined search. Poll the
    result via ``GET /api/jobs/{id}``. Uploads are named with deterministic prefixes
    (``photo_`` / ``ref_`` / ``album_``) so the runner can tell them apart — the album is
    detected by its ``album`` prefix.
    """
    scope_list = [s.strip() for s in scopes.split(",") if s.strip()]
    has_any = any([photo, ref_pdf, album, class_name.strip(), class_number.strip(), teacher_name.strip()])
    if not has_any:
        raise HTTPException(
            status_code=400,
            detail="検索条件（写真・PDF・クラス名・番号・教師名）を1つ以上指定してください。",
        )
    job_id = job_repo.new_job_id("face_query")
    in_dir, _ = storage.ensure_job_dirs(job_id)
    if photo is not None:
        await storage.save_upload_async(photo, in_dir / f"photo_{photo.filename or 'ref.png'}")
    if ref_pdf is not None:
        await storage.save_upload_async(ref_pdf, in_dir / f"ref_{ref_pdf.filename or 'ref.pdf'}")
    if album is not None:
        await storage.save_upload_async(album, in_dir / f"album_{album.filename or 'album.pdf'}")
    job_repo.create_job(
        "face_query",
        input_prefix=storage.input_prefix(job_id),
        output_prefix=storage.output_prefix(job_id),
        options={"project_id": project_id, "scopes": scope_list,
                 "class_name": class_name.strip(), "class_number": class_number.strip(),
                 "teacher_name": teacher_name.strip(), "top_k": int(top_k)},
        project_id=project_id, user_id=current.get("id"), job_id=job_id,
    )
    return {"job_id": job_id}


def _resolve_person_image(src: str, roots: list[Path]) -> Path | None:
    """Locate a result image by its stored ``source_file`` under a job's input/output roots.

    Tries the exact relative path first, then falls back to a basename search (the indexer's
    ``source_file`` is relative to whichever root held the file). Rejects traversal and any
    path that escapes the job roots. Web-tier safe: just filesystem lookups, no cv2/fitz."""
    rel = Path(src.split("#")[0])
    if rel.is_absolute() or ".." in rel.parts:
        return None
    base = rel.name
    for root in roots:
        if root is None or not root.exists():
            continue
        cand = (root / rel).resolve()
        if cand.is_relative_to(root) and cand.is_file() and cand.suffix.lower() in IMAGE_EXTENSIONS:
            return cand
    for root in roots:
        if root is None or not root.exists():
            continue
        for cand in root.rglob(base):
            rc = cand.resolve()
            if cand.is_file() and cand.suffix.lower() in IMAGE_EXTENSIONS and rc.is_relative_to(root):
                return rc
    return None


@app.get("/api/person/preview")
def person_preview(
    job_id: str = Query(...),
    src: str = Query(...),
    page: int | None = Query(default=None),
    current: dict[str, Any] = Depends(auth.get_current_user),
):
    """Serve a thumbnail/preview for a 人物検索 result.

    The match row carries ``job_id`` + ``source_file``; access is checked against that job.
    PDF-sourced rows resolve to the page render the worker cached under
    ``<output>/face_pages/`` (so the web tier never imports fitz); plain photos are streamed
    directly from the owning job's input/output dir. 404 hides both missing files and jobs
    the caller may not access."""
    from backend import pdf_utils  # import-safe: no fitz/cv2 at module top

    job = job_repo.get_job(job_id)
    if not job or not job_repo.may_access(job, current):
        raise HTTPException(status_code=404, detail="Preview not found")
    input_root = storage.resolve(job["input_prefix"]).resolve() if job.get("input_prefix") else None
    output_root = storage.resolve(job["output_prefix"]).resolve() if job.get("output_prefix") else None

    is_pdf_page = "#p" in src or src.split("#")[0].lower().endswith(".pdf") or page is not None
    if is_pdf_page:
        if output_root is None:
            raise HTTPException(status_code=404, detail="Preview not found")
        cache = (output_root / "face_pages" / pdf_utils.page_image_name(src)).resolve()
        if not cache.is_relative_to(output_root):
            raise HTTPException(status_code=404, detail="Preview not found")
        if cache.is_file():
            return FileResponse(cache, media_type="image/png")
        # Evicted locally — fall back to the cached page in S3 (deterministic key).
        served = _person_preview_from_s3(
            f"{job['output_prefix']}/face_pages/{pdf_utils.page_image_name(src)}", "image/png")
        if served is not None:
            return served
        raise HTTPException(status_code=404, detail="Preview not found")

    path = _resolve_person_image(src, [input_root, output_root])
    if path is not None:
        return FileResponse(path, media_type=image_media_type(path))
    # Evicted locally — fall back to S3 by the indexer's deterministic key (input then output).
    served = _person_preview_from_s3_for_src(job, src)
    if served is not None:
        return served
    raise HTTPException(status_code=404, detail="Preview not found")


def _person_preview_from_s3(prefix: str, media_type: str):
    """Presigned redirect for an evicted preview file, or None if S3 lacks it."""
    if s3_storage.enabled() and s3_storage.object_exists(prefix):
        url = s3_storage.presigned_url(prefix)
        if url:
            return RedirectResponse(url, status_code=307)
    return None


def _person_preview_from_s3_for_src(job: dict[str, Any], src: str):
    """Resolve a plain-photo preview from S3 after local eviction. The indexer stores
    ``source_file`` relative to the job's input (or output) root, so the S3 key is
    ``<prefix>/<src>``; try both roots. Rejects path traversal."""
    rel = Path(src.split("#")[0])
    if rel.is_absolute() or ".." in rel.parts:
        return None
    for root_prefix in (job.get("input_prefix"), job.get("output_prefix")):
        if not root_prefix:
            continue
        prefix = f"{root_prefix}/{rel.as_posix()}"
        served = _person_preview_from_s3(prefix, image_media_type(rel))
        if served is not None:
            return served
    return None


# ════════════════════════════════════ JOBS ══════════════════════════════════════════

@app.get("/api/jobs")
def list_jobs(
    current: dict[str, Any] = Depends(auth.get_current_user),
    pipeline: str | None = Query(default=None),
    project_id: str | None = Query(default=None),
    status: str | None = Query(default=None),
    mine: bool = Query(default=True),
    limit: int = Query(default=100, ge=1, le=500),
) -> dict[str, Any]:
    """List jobs so the UI can rebuild its state (and resume polling) after a reload.

    By default an operator sees only their own jobs. The owner role may pass ``mine=false``
    to list across all users for support/debugging.
    """
    scope_user = current.get("id")
    if current.get("role") == "owner" and not mine:
        scope_user = None
    jobs = job_repo.list_jobs(
        user_id=scope_user,
        pipeline=pipeline,
        project_id=project_id,
        status=status,
        limit=limit,
    )
    return {"jobs": jobs, "count": len(jobs)}


@app.get("/api/jobs/{job_id}")
def get_job(
    job_id: str, current: dict[str, Any] = Depends(auth.get_current_user)
) -> dict[str, Any]:
    """Return a single job (status/summary/error) so the UI can poll for completion.

    Backs the 人物検索 reference-photo poller, which queues a ``face_query`` job and waits on
    ``status``/``summary.matches`` here. 404 hides both missing jobs and jobs the caller may
    not access (consistent with the other per-job routes)."""
    job = job_repo.get_job(job_id)
    if not job or not job_repo.may_access(job, current):
        raise HTTPException(status_code=404, detail="Job not found")
    return job


app.include_router(auth_router)
app.include_router(project_router)
app.include_router(teacher_router)

if (FRONTEND_DIR / "index.html").exists():
    app.mount("/", StaticFiles(directory=FRONTEND_DIR, html=True), name="frontend")
