from __future__ import annotations

import math
import re
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

import numpy as np
import torch
from openpyxl import Workbook
from openpyxl.styles import Font
from PIL import ExifTags, Image, ImageOps
from torchvision import models, transforms

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff", ".heic", ".heif"}

# Client-aligned relaxed similarity rules (from Colab-tested version)
STRICT_COMBINED_THRESHOLD = 0.88
RELAXED_EMBED_THRESHOLD = 0.84
RELAXED_EMBED_THRESHOLD_WITH_TIME = 0.82
MAX_TIME_GAP_SAME_SEQUENCE = 240.0
MAX_TIME_GAP_STRICT_RELAXED = 120.0
MAX_SEQUENCE_GAP = 10

WEIGHT_TECHNICAL = 0.30
WEIGHT_EXPRESSION = 0.25
WEIGHT_COMPOSITION = 0.25
WEIGHT_RARITY = 0.20
DEFAULT_BEST_SHOT_COUNT = None
DEFAULT_BEST_SHOT_RATIO = 0.30
NG_TAIL_RATIO = 0.05


@dataclass
class PipelineResult:
    total_input_images: int
    total_clusters: int
    total_representative_candidates: int
    dedup_reduction_rate: float
    best_shot_count: int
    final_selected_count: int
    ng_count_after_menna: int
    other_passing_count: int


@dataclass
class ImageRecord:
    path: Path
    capture_time: Optional[datetime]
    sequence_tail: Optional[int]
    embedding: np.ndarray
    focus_score: float
    brightness_score: float


class SnapPipeline:
    def __init__(self) -> None:
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        weights = models.ResNet50_Weights.IMAGENET1K_V2
        model = models.resnet50(weights=weights)
        model.fc = torch.nn.Identity()
        self.model = model.to(self.device).eval()
        self.preprocess = weights.transforms()

    @staticmethod
    def _collect_images(input_dir: Path) -> list[Path]:
        return sorted(p for p in input_dir.rglob("*") if p.suffix.lower() in IMAGE_EXTS)

    @staticmethod
    def _safe_open_image(path: Path) -> Image.Image:
        with Image.open(path) as img:
            return ImageOps.exif_transpose(img).convert("RGB")

    @staticmethod
    def _get_capture_time(path: Path) -> Optional[datetime]:
        try:
            with Image.open(path) as img:
                exif = img.getexif()
                if exif:
                    tag_map = {ExifTags.TAGS.get(k, k): v for k, v in exif.items()}
                    for key in ["DateTimeOriginal", "DateTimeDigitized", "DateTime"]:
                        if key in tag_map:
                            try:
                                return datetime.strptime(str(tag_map[key]), "%Y:%m:%d %H:%M:%S")
                            except Exception:
                                continue
        except Exception:
            pass
        try:
            return datetime.fromtimestamp(path.stat().st_mtime)
        except Exception:
            return None

    @staticmethod
    def _parse_filename_numeric_tail(name: str) -> Optional[int]:
        nums = re.findall(r"(\d+)", Path(name).stem)
        if not nums:
            return None
        try:
            return int(nums[-1])
        except Exception:
            return None

    @staticmethod
    def _time_gap_seconds(a: Optional[datetime], b: Optional[datetime]) -> Optional[float]:
        if a is None or b is None:
            return None
        return abs((a - b).total_seconds())

    @staticmethod
    def _seq_gap(a: Optional[int], b: Optional[int]) -> Optional[int]:
        if a is None or b is None:
            return None
        return abs(a - b)

    @staticmethod
    def _cosine(a: np.ndarray, b: np.ndarray) -> float:
        return float(np.dot(a, b) / ((np.linalg.norm(a) * np.linalg.norm(b)) + 1e-12))

    @staticmethod
    def _focus_score(img_rgb: Image.Image) -> float:
        arr = np.array(img_rgb.convert("L"), dtype=np.float32)
        gy, gx = np.gradient(arr)
        mag = np.sqrt(gx**2 + gy**2)
        return float(np.var(mag))

    @staticmethod
    def _brightness_score(img_rgb: Image.Image) -> float:
        arr = np.array(img_rgb.convert("L"), dtype=np.float32)
        return float(arr.mean())

    @staticmethod
    def _contrast_score(img_rgb: Image.Image) -> float:
        arr = np.array(img_rgb.convert("L"), dtype=np.float32)
        return float(arr.std())

    @staticmethod
    def _thirds_composition_score(img_rgb: Image.Image) -> float:
        gray = np.array(img_rgb.convert("L"), dtype=np.float32)
        h, w = gray.shape
        gy, gx = np.gradient(gray)
        edge = np.sqrt(gx**2 + gy**2)

        points = [
            (h // 3, w // 3),
            (h // 3, 2 * w // 3),
            (2 * h // 3, w // 3),
            (2 * h // 3, 2 * w // 3),
        ]
        r = max(4, min(h, w) // 12)
        energy = 0.0
        for py, px in points:
            y1, y2 = max(0, py - r), min(h, py + r)
            x1, x2 = max(0, px - r), min(w, px + r)
            energy += float(edge[y1:y2, x1:x2].mean())
        return energy / max(1, len(points))

    def _embed(self, img: Image.Image) -> np.ndarray:
        tensor = self.preprocess(img).unsqueeze(0).to(self.device)
        with torch.no_grad():
            vec = self.model(tensor).detach().cpu().numpy().reshape(-1).astype(np.float32)
        vec /= np.linalg.norm(vec) + 1e-12
        return vec

    def _load_records(self, image_paths: list[Path]) -> list[ImageRecord]:
        records: list[ImageRecord] = []
        for path in image_paths:
            img = self._safe_open_image(path)
            records.append(
                ImageRecord(
                    path=path,
                    capture_time=self._get_capture_time(path),
                    sequence_tail=self._parse_filename_numeric_tail(path.name),
                    embedding=self._embed(img),
                    focus_score=self._focus_score(img),
                    brightness_score=self._brightness_score(img),
                )
            )

        records.sort(key=lambda r: (r.capture_time or datetime.min, r.path.name))
        return records

    def _is_similar(self, a: ImageRecord, b: ImageRecord) -> bool:
        embed_sim = self._cosine(a.embedding, b.embedding)
        tgap = self._time_gap_seconds(a.capture_time, b.capture_time)
        sgap = self._seq_gap(a.sequence_tail, b.sequence_tail)

        strict_context = (
            (tgap is not None and tgap <= MAX_TIME_GAP_SAME_SEQUENCE)
            or (sgap is not None and sgap <= MAX_SEQUENCE_GAP)
        )
        if embed_sim >= STRICT_COMBINED_THRESHOLD and strict_context:
            return True

        relaxed_context = (
            (tgap is not None and tgap <= MAX_TIME_GAP_STRICT_RELAXED)
            or (sgap is not None and sgap <= MAX_SEQUENCE_GAP)
        )
        if embed_sim >= RELAXED_EMBED_THRESHOLD and relaxed_context:
            return True

        if embed_sim >= RELAXED_EMBED_THRESHOLD_WITH_TIME and tgap is not None and tgap <= MAX_TIME_GAP_STRICT_RELAXED:
            return True

        return False

    def _cluster_records(self, records: list[ImageRecord]) -> list[list[ImageRecord]]:
        clusters: list[list[ImageRecord]] = []
        for rec in records:
            placed = False
            for cluster in clusters:
                # Compare against cluster representative and recent members for looser grouping continuity
                rep = max(cluster, key=lambda x: x.focus_score)
                if self._is_similar(rec, rep) or any(self._is_similar(rec, m) for m in cluster[-3:]):
                    cluster.append(rec)
                    placed = True
                    break
            if not placed:
                clusters.append([rec])
        return clusters

    @staticmethod
    def _normalize(vals: list[float]) -> list[float]:
        if not vals:
            return []
        lo, hi = min(vals), max(vals)
        if math.isclose(lo, hi):
            return [0.5 for _ in vals]
        return [(v - lo) / (hi - lo) for v in vals]

    def _score_representatives(self, reps: list[ImageRecord]) -> list[tuple[ImageRecord, float]]:
        # Additional Menna-like dimensions from image content
        contrast_vals: list[float] = []
        composition_vals: list[float] = []
        for r in reps:
            img = self._safe_open_image(r.path)
            contrast_vals.append(self._contrast_score(img))
            composition_vals.append(self._thirds_composition_score(img))

        tech = self._normalize([r.focus_score for r in reps])
        expr = self._normalize([r.brightness_score for r in reps])
        comp = self._normalize(composition_vals)

        # rarity: distance from average embedding among representatives
        if reps:
            center = np.mean([r.embedding for r in reps], axis=0)
            center /= np.linalg.norm(center) + 1e-12
            rarity_raw = [1.0 - self._cosine(r.embedding, center) for r in reps]
        else:
            rarity_raw = []
        rarity = self._normalize(rarity_raw)

        scored: list[tuple[ImageRecord, float]] = []
        for i, r in enumerate(reps):
            # expression: prefer well-exposed but not extreme brightness
            expr_balanced = 1.0 - abs(expr[i] - 0.55)
            score = (
                WEIGHT_TECHNICAL * tech[i]
                + WEIGHT_EXPRESSION * expr_balanced
                + WEIGHT_COMPOSITION * comp[i]
                + WEIGHT_RARITY * rarity[i]
            )
            scored.append((r, float(score)))

        scored.sort(key=lambda x: x[1], reverse=True)
        return scored

    @staticmethod
    def _recommended_best_shot_count(total_representatives: int) -> int:
        if total_representatives < 1:
            return 0
        return max(1, int(round(total_representatives * DEFAULT_BEST_SHOT_RATIO)))

    @staticmethod
    def _ng_tail_count(total_representatives: int) -> int:
        if total_representatives <= 1:
            return 0
        return min(max(1, int(round(total_representatives * NG_TAIL_RATIO))), total_representatives - 1)

    def _select_buckets(
        self, reps: list[ImageRecord], best_shot_count: int | None = None
    ) -> tuple[list[ImageRecord], list[ImageRecord], list[ImageRecord]]:
        if not reps:
            return [], [], []

        if best_shot_count is not None and (not isinstance(best_shot_count, int) or best_shot_count < 1):
            raise ValueError("best_shot_count must be a positive integer.")

        scored = self._score_representatives(reps)
        ordered = [r for r, _ in scored]

        ng_count = self._ng_tail_count(len(ordered))
        ng = ordered[-ng_count:] if ng_count else []
        non_ng = ordered[:-ng_count] if ng_count else ordered

        requested_count = (
            self._recommended_best_shot_count(len(ordered)) if best_shot_count is None else best_shot_count
        )
        final_count = min(requested_count, len(non_ng))
        final = non_ng[:final_count]
        other = non_ng[final_count:]

        return final, ng, other

    @staticmethod
    def _copy(path: Path, dst: Path) -> None:
        dst.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(path, dst)

    def _score_details_for_report(self, reps: list[ImageRecord]) -> dict[Path, dict[str, float]]:
        contrast_vals: list[float] = []
        composition_vals: list[float] = []
        for rec in reps:
            img = self._safe_open_image(rec.path)
            contrast_vals.append(self._contrast_score(img))
            composition_vals.append(self._thirds_composition_score(img))

        tech = self._normalize([rec.focus_score for rec in reps])
        brightness = self._normalize([rec.brightness_score for rec in reps])
        comp = self._normalize(composition_vals)
        if reps:
            center = np.mean([rec.embedding for rec in reps], axis=0)
            center /= np.linalg.norm(center) + 1e-12
            rarity_raw = [1.0 - self._cosine(rec.embedding, center) for rec in reps]
        else:
            rarity_raw = []
        rarity = self._normalize(rarity_raw)

        details: dict[Path, dict[str, float]] = {}
        for i, rec in enumerate(reps):
            expression = 1.0 - abs(brightness[i] - 0.55)
            total = (
                WEIGHT_TECHNICAL * tech[i]
                + WEIGHT_EXPRESSION * expression
                + WEIGHT_COMPOSITION * comp[i]
                + WEIGHT_RARITY * rarity[i]
            )
            details[rec.path] = {
                "total": round(total * 10.0, 2),
                "technical": round(tech[i] * 10.0, 2),
                "expression": round(expression * 10.0, 2),
                "composition": round(comp[i] * 10.0, 2),
                "rarity": round(rarity[i] * 10.0, 2),
            }
        return details

    @staticmethod
    def _capture_time_text(rec: ImageRecord) -> str:
        return rec.capture_time.strftime("%Y-%m-%d %H:%M:%S") if rec.capture_time else ""

    @staticmethod
    def _bold_header(ws) -> None:
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"

    def _write_outputs(
        self,
        output_dir: Path,
        clusters: list[list[ImageRecord]],
        reps: list[ImageRecord],
        final: list[ImageRecord],
        ng: list[ImageRecord],
        other: list[ImageRecord],
        summary: PipelineResult,
    ) -> None:
        sim_dir = output_dir / "similarity_clusters"
        final_dir = output_dir / "final_selected"
        ng_dir = output_dir / "ng_photos"
        other_dir = output_dir / "other_passing"
        for d in [sim_dir, final_dir, ng_dir, other_dir]:
            d.mkdir(parents=True, exist_ok=True)

        rep_paths = {r.path for r in reps}
        cluster_ids: dict[Path, int] = {}
        for i, cluster in enumerate(clusters, start=1):
            cdir = sim_dir / f"cluster_{i:03d}"
            cdir.mkdir(parents=True, exist_ok=True)
            for rec in cluster:
                cluster_ids[rec.path] = i
                prefix = "REP_" if rec.path in rep_paths else ""
                self._copy(rec.path, cdir / f"{prefix}{rec.path.name}")

        for rec in final:
            self._copy(rec.path, final_dir / rec.path.name)
        for rec in ng:
            self._copy(rec.path, ng_dir / rec.path.name)
        for rec in other:
            self._copy(rec.path, other_dir / rec.path.name)

        score_details = self._score_details_for_report(reps)

        wb = Workbook()
        ws_best = wb.active
        ws_best.title = "ベストショット一覧"
        ws_best.append(
            [
                "ファイル名",
                "区分",
                "総合スコア",
                "技術スコア",
                "表情/明るさスコア",
                "構図スコア",
                "希少性スコア",
                "類似グループID",
                "撮影日時",
                "コメント",
                "選定理由",
            ]
        )
        for rec in final:
            scores = score_details.get(rec.path, {})
            ws_best.append(
                [
                    rec.path.name,
                    "ベストショット",
                    scores.get("total"),
                    scores.get("technical"),
                    scores.get("expression"),
                    scores.get("composition"),
                    scores.get("rarity"),
                    cluster_ids.get(rec.path, ""),
                    self._capture_time_text(rec),
                    "総合評価が高く、ベストショット候補として選定されました。",
                    "類似写真グループ内の代表候補として選定され、総合評価が高いため最終選定されました。",
                ]
            )
        self._bold_header(ws_best)

        ws_ng = wb.create_sheet("NG写真一覧")
        ws_ng.append(["ファイル名", "区分", "総合スコア", "NG理由", "コメント", "類似グループID", "撮影日時"])
        for rec in ng:
            scores = score_details.get(rec.path, {})
            ws_ng.append(
                [
                    rec.path.name,
                    "NG写真",
                    scores.get("total"),
                    "代表候補の中で総合評価が低いため",
                    "代表候補の中で総合評価が低いため、NG候補として分類されました。",
                    cluster_ids.get(rec.path, ""),
                    self._capture_time_text(rec),
                ]
            )
        self._bold_header(ws_ng)

        ws_other = wb.create_sheet("未選定写真一覧")
        ws_other.append(["ファイル名", "区分", "総合スコア", "コメント", "未選定理由", "類似グループID", "撮影日時"])
        for rec in other:
            scores = score_details.get(rec.path, {})
            ws_other.append(
                [
                    rec.path.name,
                    "未選定写真",
                    scores.get("total"),
                    "品質基準は満たしていますが、ベストショット選定数の上限により未選定となりました。",
                    "ベストショット選定数の上限により未選定",
                    cluster_ids.get(rec.path, ""),
                    self._capture_time_text(rec),
                ]
            )
        self._bold_header(ws_other)

        ws_summary = wb.create_sheet("統計サマリー")
        ws_summary.append(["項目", "値"])
        ws_summary.append(["■ 全体集計", ""])
        ws_summary.append(["入力写真数", summary.total_input_images])
        ws_summary.append(["類似グループ数", summary.total_clusters])
        ws_summary.append(["代表候補数", summary.total_representative_candidates])
        ws_summary.append(["重複削減率", summary.dedup_reduction_rate])
        ws_summary.append(["ベストショット選定数", summary.final_selected_count])
        ws_summary.append(["NG写真数", summary.ng_count_after_menna])
        ws_summary.append(["未選定写真数", summary.other_passing_count])
        ws_summary.append(["ベストショット指定数", summary.best_shot_count])
        self._bold_header(ws_summary)

        wb.save(output_dir / "snap_pipeline_report.xlsx")

    def run(
        self,
        input_dir: Path,
        output_dir: Path,
        best_shot_count: int | None = None,
    ) -> PipelineResult:
        if best_shot_count is not None and (not isinstance(best_shot_count, int) or best_shot_count < 1):
            raise ValueError("best_shot_count must be a positive integer.")
        image_paths = self._collect_images(input_dir)
        records = self._load_records(image_paths)
        clusters = self._cluster_records(records)

        reps = [max(cluster, key=lambda x: x.focus_score) for cluster in clusters]
        final, ng, other = self._select_buckets(reps, best_shot_count)

        dedup_reduction_rate = 0.0
        if image_paths:
            dedup_reduction_rate = round((1.0 - (len(reps) / len(image_paths))) * 100.0, 2)

        summary = PipelineResult(
            total_input_images=len(image_paths),
            total_clusters=len(clusters),
            total_representative_candidates=len(reps),
            dedup_reduction_rate=dedup_reduction_rate,
            best_shot_count=len(final),
            final_selected_count=len(final),
            ng_count_after_menna=len(ng),
            other_passing_count=len(other),
        )

        if output_dir.exists():
            shutil.rmtree(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        self._write_outputs(output_dir, clusters, reps, final, ng, other, summary)
        return summary


__all__ = ["SnapPipeline", "PipelineResult", "DEFAULT_BEST_SHOT_COUNT"]
