from __future__ import annotations

import base64
import io
import json
import logging
import os
import shutil
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import cv2
import numpy as np
from openpyxl import Workbook
from PIL import ExifTags, Image
from backend.excel_labels import CLUB_SHEET_LABELS, excel_label, translate_display_value

logger = logging.getLogger(__name__)

try:
    from insightface.app import FaceAnalysis
except Exception:
    FaceAnalysis = None

try:
    from openai import OpenAI
except Exception:
    OpenAI = None

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff", ".heic", ".heif"}
RANK_TOKEN = "本"
RANK_PAD = 2
MODEL_VISION = os.getenv("OPENAI_MODEL_VISION", os.getenv("MODEL_VISION", "gpt-5.4"))
EYE_CLOSED_RATIO_THRESHOLD = float(os.getenv("CLUB_EYE_CLOSED_RATIO_THRESHOLD", "0.24"))
LEFT_EYE_LANDMARK_IDX = list(range(33, 43))
RIGHT_EYE_LANDMARK_IDX = list(range(87, 97))

# ── Constants shared with ai_detect.py ──────────────────────────────────────────────
COLOR_CLOSED  = (0,   0, 210)   # BGR deep red   – eyes closed face box
COLOR_OPEN    = (0, 180,  60)   # BGR green       – eyes open face box
COLOR_NO_FACE = (0, 200, 220)   # BGR yellow-cyan – detected but unclear/occluded
FACE_PADDING_PCT = 0.25         # % padding added around bbox before sending crop to GPT

_GPT_EYE_PROMPT = """
You are examining a CROPPED IMAGE of a single FACE. Determine if the eyes are closed.

STEP 1 — has_clear_face: Can you see this person's eye region?
  TRUE  → Eye area is visible (even if eyes are closed, angled, or looking down).
  FALSE → Only back of head / extreme side profile / eyes fully blocked by hair, hands,
          or mask / image too blurry or small to judge.

STEP 2 — eyes_closed: Are the eyes closed?
  TRUE  → The upper eyelid is resting DOWN covering the eye. No iris or pupil visible.
  FALSE → ANY part of the dark iris or pupil is visible, even a small sliver.

THE MOST COMMON MISTAKE — smiling squint:
  Smiling raises the cheeks and narrows the eye opening. This is NOT closed eyes.
  Key test: if you can see ANY dark iris/pupil crescent → eyes are OPEN (FALSE).

RULES for East Asian faces:
  - Naturally narrow or monolid eyes show only a thin iris line when open → FALSE.
  - Smiling/laughing with crescent-shaped openings → still FALSE if iris is visible.
  - Only TRUE if eyelid is fully covering the eye with NO iris visible at all.

Return ONLY this JSON (no markdown, no extra text):
{"has_clear_face": true|false, "eyes_closed": true|false, "confidence": "high"|"medium"|"low"}
""".strip()


def _client():
    """Return an OpenAI client if the API key is set, else None."""
    if OpenAI is None:
        return None
    key = os.getenv("OPENAI_API_KEY")
    if not key:
        return None
    return OpenAI(api_key=key)


def _ask_eye(openai_client, face_crop_bgr) -> tuple[bool, bool, str]:
    """Send a face crop to GPT and return (has_clear_face, eyes_closed, confidence).

    Falls back to (False, False, 'low') on any error so the caller can degrade gracefully.
    The crop is expected to be a BGR numpy array (OpenCV format).
    """
    if openai_client is None:
        return False, False, "low"
    import base64, json as _json
    h, w = face_crop_bgr.shape[:2]
    if w < 30 or h < 30:
        return False, False, "low"
    try:
        _, buf = cv2.imencode(".jpg", face_crop_bgr, [cv2.IMWRITE_JPEG_QUALITY, 90])
        b64 = base64.b64encode(buf).decode("utf-8")
        resp = openai_client.responses.create(
            model=MODEL_VISION,
            input=[{
                "role": "user",
                "content": [
                    {"type": "input_text",  "text": _GPT_EYE_PROMPT},
                    {"type": "input_image",
                     "image_url": f"data:image/jpeg;base64,{b64}",
                     "detail": "high"},
                ],
            }],
            reasoning={"effort": "medium"},
        )
        text = (resp.output_text or "").strip()
        if "```" in text:
            parts = text.split("```")
            text = parts[1] if len(parts) > 1 else parts[0]
            if text.startswith("json"):
                text = text[4:]
        data = _json.loads(text.strip())
        return (
            bool(data.get("has_clear_face", False)),
            bool(data.get("eyes_closed",    False)),
            str(data.get("confidence",      "medium")),
        )
    except Exception as exc:
        logger.debug("_ask_eye GPT error: %s", exc)
        return False, False, "low"


PHOTO_EVAL_PROMPT = """
You are a professional photo editor selecting the best shot for a school club album.
Return STRICT JSON only.
JSON keys must stay English.
ng_reason and short_comment values MUST be Japanese.

Important pose rule:
- Thumbs up is acceptable and must NOT be treated as NG.
- NG hand/pose examples: middle finger, obscene/improper gestures, sexual/improper posing.
- ng_reason and short_comment values MUST be Japanese.

JSON schema:
{
  "formality_score": 0-10,
  "beauty_score": 0-10,
  "expression_score": 0-10,
  "emotion_score": 0-10,
  "people_count_score": 0-10,
  "has_hand_gesture": true|false,
  "has_exaggerated_expression": true|false,
  "eyes_closed": true|false,
  "face_obscured": true|false,
  "subject_falling": true|false,
  "eating_or_mouth_full": true|false,
  "strange_obscene_posing": true|false,
  "bad_pose": true|false,
  "visible_underwear": true|false,
  "severe_framing_issue": true|false,
  "gesture_expression_penalty": 0-5,
  "is_ng": true|false,
  "ng_reason": "short reason",
  "short_comment": "short comment"
}
""".strip()

@dataclass
class FaceDetail:
    face_index: int
    bbox: tuple[int, int, int, int]
    left_eye_ratio: float | None
    right_eye_ratio: float | None
    eye_closed: bool | None


@dataclass
class ClubImageResult:
    club_name: str
    path: Path
    shooting_date: str
    person_count: int
    eyes_closed_photo: bool
    closed_eye_face_count: int
    face_details: list[FaceDetail] = field(default_factory=list)
    eval_data: dict[str, Any] = field(default_factory=dict)
    ng_flag: bool = False
    ng_reason: str = ""
    total_score: float = 0.0
    rank: int = 0
    renamed: str = ""


def _get_client():
    key = os.getenv("OPENAI_API_KEY")
    if not key or OpenAI is None:
        return None
    return OpenAI(api_key=key)


def _collect_club_images(extracted_root: Path):
    """Collect image files grouped by their top-level club folder."""
    out = {}
    detected = [p for p in sorted(extracted_root.rglob("*")) if p.is_file() and p.suffix.lower() in IMAGE_EXTS]
    logger.info("Club pipeline detected image count=%s paths=%s", len(detected), [str(p) for p in detected])
    for image_path in detected:
        rel = image_path.relative_to(extracted_root)
        if len(rel.parts) < 2:
            logger.warning("Ignoring club image without a club folder: %s", image_path)
            continue
        out.setdefault(rel.parts[0], []).append(image_path)
    logger.info("Club pipeline detected club count=%s names=%s", len(out), list(out))
    return out


def _read_image(path: Path) -> np.ndarray | None:
    """Read images from paths that may contain non-ASCII characters."""
    try:
        data = np.fromfile(path, dtype=np.uint8)
        if data.size == 0:
            return None
        return cv2.imdecode(data, cv2.IMREAD_COLOR)
    except Exception:
        logger.exception("Failed to read club image: %s", path)
        return None


def _write_image(path: Path, image: np.ndarray) -> None:
    """Write images to paths that may contain non-ASCII characters."""
    path.parent.mkdir(parents=True, exist_ok=True)
    ext = path.suffix or ".jpg"
    ok, encoded = cv2.imencode(ext, image)
    if not ok:
        raise ValueError(f"Failed to encode output image: {path}")
    encoded.tofile(path)



def _safe_extract_zip(zip_path: Path, dst: Path) -> list[str]:
    dst_root = dst.resolve()
    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()
        logger.info("Club pipeline zip contents=%s", names)
        for member in zf.infolist():
            target = (dst / member.filename).resolve()
            try:
                target.relative_to(dst_root)
            except ValueError as exc:
                raise ValueError("ZIP contains an unsafe path.") from exc
            if member.is_dir():
                target.mkdir(parents=True, exist_ok=True)
                continue
            target.parent.mkdir(parents=True, exist_ok=True)
            with zf.open(member, "r") as src, target.open("wb") as out_file:
                shutil.copyfileobj(src, out_file)
    return names

def _shooting_date(path: Path) -> str:
    try:
        exif = Image.open(path).getexif()
        if exif:
            for tag_id, value in exif.items():
                if ExifTags.TAGS.get(tag_id, tag_id) == "DateTimeOriginal" and value:
                    return str(value)[:10].replace(":", "")
    except Exception:
        pass
    return datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y%m%d")


def _eye_aspect_ratio(landmarks: np.ndarray, indices: list[int]) -> float | None:
    if landmarks is None or landmarks.shape[0] <= max(indices):
        return None
    pts = landmarks[indices]
    width = float(pts[:, 0].max() - pts[:, 0].min())
    height = float(pts[:, 1].max() - pts[:, 1].min())
    if width < 1:
        return None
    return round(height / width, 4)


def _landmark_points_for_log(landmarks: np.ndarray, indices: list[int]) -> list[tuple[int, int]]:
    return [(int(round(x)), int(round(y))) for x, y in landmarks[indices]]


def _gpt_eye_fallback(openai_client, image_bgr, bbox) -> bool | None:
    """Ask GPT whether the eyes are closed for a single face crop.

    Used when InsightFace 106-point landmarks are unavailable (otherwise every face
    degrades to the "unknown / yellow box" state). Mirrors ai_detect's crop+padding so
    the model sees enough context. Returns True/False, or None when GPT can't judge.
    """
    if openai_client is None:
        return None
    h, w = image_bgr.shape[:2]
    x1, y1, x2, y2 = bbox
    pw = int((x2 - x1) * FACE_PADDING_PCT)
    ph = int((y2 - y1) * FACE_PADDING_PCT)
    crop = image_bgr[max(0, y1 - ph):min(h, y2 + ph), max(0, x1 - pw):min(w, x2 + pw)]
    if crop.size == 0:
        return None
    has_face, eyes_closed, _conf = _ask_eye(openai_client, crop)
    return bool(eyes_closed) if has_face else None


def _analyze_faces(image_bgr: np.ndarray, face_app: Any | None, openai_client: Any | None = None):
    img_h, img_w = image_bgr.shape[:2]
    vis = image_bgr.copy(); details=[]
    face_label_scale = max(0.8, img_w / 1400)
    face_label_thickness = max(2, int(img_w / 800))
    if face_app is not None:
        faces = face_app.get(cv2.cvtColor(image_bgr, cv2.COLOR_BGR2RGB))
        logger.debug("Club eye-close detection face count=%s threshold=%s", len(faces), EYE_CLOSED_RATIO_THRESHOLD)
        for i,f in enumerate(faces,1):
            x1,y1,x2,y2=[int(v) for v in f.bbox]
            lmk = getattr(f, "landmark_2d_106", None)
            closed: bool | None = None
            lr: float | None = None
            rr: float | None = None
            if lmk is None or getattr(lmk, "shape", (0,))[0] <= max(RIGHT_EYE_LANDMARK_IDX):
                logger.warning(
                    "Club eye-close landmarks missing face_index=%s bbox=%s landmark_shape=%s",
                    i,
                    (x1, y1, x2, y2),
                    getattr(lmk, "shape", None),
                )
            else:
                lr = _eye_aspect_ratio(lmk, LEFT_EYE_LANDMARK_IDX)
                rr = _eye_aspect_ratio(lmk, RIGHT_EYE_LANDMARK_IDX)
                logger.debug(
                    "Club eye-close landmarks face_index=%s bbox=%s left_points=%s right_points=%s",
                    i,
                    (x1, y1, x2, y2),
                    _landmark_points_for_log(lmk, LEFT_EYE_LANDMARK_IDX),
                    _landmark_points_for_log(lmk, RIGHT_EYE_LANDMARK_IDX),
                )
                if lr is None or rr is None:
                    logger.warning(
                        "Club eye-close ratio unavailable face_index=%s bbox=%s left_ratio=%s right_ratio=%s",
                        i,
                        (x1, y1, x2, y2),
                        lr,
                        rr,
                    )
                else:
                    closed = lr <= EYE_CLOSED_RATIO_THRESHOLD or rr <= EYE_CLOSED_RATIO_THRESHOLD
                logger.debug(
                    "Club eye-close ratios face_index=%s left_ratio=%s right_ratio=%s threshold=%s closed=%s",
                    i,
                    lr,
                    rr,
                    EYE_CLOSED_RATIO_THRESHOLD,
                    closed,
                )
            if closed is None and openai_client is not None:
                closed = _gpt_eye_fallback(openai_client, image_bgr, (x1, y1, x2, y2))
                logger.debug(
                    "Club eye-close GPT fallback face_index=%s bbox=%s closed=%s",
                    i,
                    (x1, y1, x2, y2),
                    closed,
                )
            details.append(FaceDetail(i,(x1,y1,x2,y2),lr,rr,closed))
            color = (0, 0, 255) if closed is True else (0, 180, 0) if closed is False else (0, 200, 220)
            cv2.rectangle(vis,(x1,y1),(x2,y2),color,2)
            if closed is True:
                cv2.putText(vis, "EYES CLOSED", (x1, max(15, y1 - 6)), cv2.FONT_HERSHEY_SIMPLEX, face_label_scale, color, face_label_thickness)
    else:
        g=cv2.cvtColor(image_bgr, cv2.COLOR_BGR2GRAY)
        face = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
        eye = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_eye.xml")
        faces = face.detectMultiScale(g,1.15,5,minSize=(50,50))
        logger.debug("Club eye-close Haar fallback face count=%s", len(faces))
        for i,(x,y,w,h) in enumerate(faces,1):
            eye_hits = eye.detectMultiScale(g[y:y+h, x:x+w],1.1,4,minSize=(10,10))
            closed = len(eye_hits)==0
            logger.debug(
                "Club eye-close Haar fallback face_index=%s bbox=%s eye_hits=%s closed=%s",
                i,
                (x,y,x+w,y+h),
                len(eye_hits),
                closed,
            )
            details.append(FaceDetail(i,(x,y,x+w,y+h),None,None,closed))
            cv2.rectangle(vis,(x,y),(x+w,y+h),(0,0,255) if closed else (0,180,0),2)
    closed_count=sum(1 for d in details if d.eye_closed is True)
    logger.debug(
        "Club eye-close image result face_count=%s closed_face_count=%s closed_photo=%s",
        len(details),
        closed_count,
        closed_count > 0,
    )
    font_scale = max(1.2, img_w / 900)
    thickness = max(3, int(img_w / 500))
    x = max(20, int(img_w * 0.015))
    y = max(50, int(img_h * 0.05))
    cv2.putText(vis, f"People:{len(details)} ClosedEyes:{closed_count}",(x,y),cv2.FONT_HERSHEY_SIMPLEX,font_scale,(255,0,0),thickness)
    return len(details), closed_count, details, vis


def _parse_json_obj(text: str) -> dict[str, Any]:
    t=(text or "").strip(); s=t.find("{"); e=t.rfind("}")
    if s==-1 or e==-1: return {}
    return json.loads(t[s:e+1])


def _evaluate_with_gpt(client: Any, image_path: Path, person_count: int) -> dict[str, Any] | None:
    try:
        img=Image.open(image_path).convert("RGB")
        b=io.BytesIO(); img.save(b, format="JPEG", quality=92)
        b64=base64.b64encode(b.getvalue()).decode("utf-8")
        resp=client.responses.create(model=MODEL_VISION,input=[{"role":"user","content":[{"type":"input_text","text":PHOTO_EVAL_PROMPT+f"\nDetected people count: {person_count}"},{"type":"input_image","image_url":f"data:image/jpeg;base64,{b64}","detail":"high"}]}])
        return _parse_json_obj(resp.output_text)
    except Exception:
        return None


def _fallback_eval(person_count:int, closed_count:int)->dict[str,Any]:
    return {"formality_score":max(0,8-closed_count),"beauty_score":7,"expression_score":max(0,8-closed_count),"emotion_score":7,"people_count_score":min(10, max(1,person_count)),"has_hand_gesture":False,"has_exaggerated_expression":False,"eyes_closed":closed_count>0,"face_obscured":person_count==0,"subject_falling":False,"eating_or_mouth_full":False,"strange_obscene_posing":False,"bad_pose":False,"visible_underwear":False,"severe_framing_issue":False,"gesture_expression_penalty":0,"is_ng":person_count==0,"ng_reason":"no face detected" if person_count==0 else "","short_comment":"heuristic"}


def _score(eval_data: dict[str, Any], closed_count: int)->tuple[float,bool,str]:
    w = (float(eval_data.get("formality_score",0))*0.30 + float(eval_data.get("beauty_score",0))*0.25 + float(eval_data.get("expression_score",0))*0.15 + float(eval_data.get("emotion_score",0))*0.10 + float(eval_data.get("people_count_score",0))*0.20)
    total = w - float(eval_data.get("gesture_expression_penalty",0)) - (closed_count*0.75)
    ng = bool(eval_data.get("is_ng", False)) or bool(eval_data.get("eyes_closed", False) and closed_count>0)
    reason = str(eval_data.get("ng_reason", ""))
    return round(total,3), ng, reason


def run_club_pipeline(input_zip_path: str, output_dir: str) -> dict[str, Any]:
    out=Path(output_dir); extracted=out/"extracted"; club_out=out/"Club_Output"
    ranked=club_out/"ranked_photos"; ranked_marked=club_out/"ranked_photos_marked"; ng_root=club_out/"ng_photos"; marked=club_out/"marked_images"; clean=club_out/"clean_images"
    if out.exists(): shutil.rmtree(out)
    for d in [extracted,ranked,ranked_marked,ng_root,marked,clean]: d.mkdir(parents=True,exist_ok=True)
    logger.info("Club pipeline input zip=%s output_root=%s", input_zip_path, out)
    _safe_extract_zip(Path(input_zip_path), extracted)
    logger.info("Club pipeline extracted input path=%s", extracted)
    clubs=_collect_club_images(extracted)
    if not clubs:
        raise ValueError("No club images were detected in the uploaded ZIP. Images must be inside club folders.")
    face_app=None
    if FaceAnalysis is not None:
        try:
            face_app=FaceAnalysis(name="buffalo_l", providers=["CPUExecutionProvider"]); face_app.prepare(ctx_id=0, det_size=(640, 640))
        except Exception: face_app=None
    client=_get_client(); results=[]
    marked_outputs: dict[tuple[str, str], Path] = {}
    for club,images in clubs.items():
        for p in images:
            img=_read_image(p)
            if img is None:
                logger.warning("Skipping unreadable club image: %s", p)
                continue
            pc,cc,fd,vis=_analyze_faces(img,face_app,client)
            logger.debug(
                "Club eye-close per-image file=%s face_count=%s closed_face_count=%s closed_photo=%s threshold=%s",
                p,
                pc,
                cc,
                cc > 0,
                EYE_CLOSED_RATIO_THRESHOLD,
            )
            (marked/club).mkdir(parents=True,exist_ok=True); (clean/club).mkdir(parents=True,exist_ok=True)
            marked_path = marked / club / p.name
            clean_path = clean / club / p.name
            _write_image(marked_path, vis)
            _write_image(clean_path, img)
            marked_outputs[(club, p.name)] = marked_path
            ev=_evaluate_with_gpt(client,p,pc) if client else None
            if not ev: ev=_fallback_eval(pc,cc)
            total,ng,reason=_score(ev,cc)
            results.append(ClubImageResult(club,p,_shooting_date(p),pc,cc>0,cc,fd,ev,ng,reason,total))
    if not results:
        raise ValueError("Club images were detected, but none could be read or processed.")
    by={}
    for r in results: by.setdefault(r.club_name,[]).append(r)
    for club,rows in by.items():
        rows.sort(key=lambda x:(x.ng_flag,-x.total_score))
        for i,r in enumerate(rows,1):
            r.rank=i; r.renamed=f"{club}_{r.shooting_date}_{r.path.stem}_{RANK_TOKEN}{i:02d}{r.path.suffix.lower()}"
            (ranked/club).mkdir(parents=True,exist_ok=True); (ranked_marked/club).mkdir(parents=True,exist_ok=True)
            shutil.copy2(clean / club / r.path.name, ranked / club / r.renamed)
            marked_source = marked_outputs.get((club, r.path.name), marked / club / r.path.name)
            shutil.copy2(marked_source, ranked_marked / club / r.renamed)
            if r.ng_flag:
                (ng_root/club).mkdir(parents=True,exist_ok=True); shutil.copy2(clean/club/r.path.name, ng_root/club/r.path.name)

    excel=club_out/"club_result.xlsx"; jsonp=club_out/"club_result.json"
    wb=Workbook(); ws=wb.active; ws.title=CLUB_SHEET_LABELS["Summary"]
    ws.append([excel_label("club_count"),excel_label("photo_count"),excel_label("closed_eye_photo_count"),excel_label("closed_eye_face_count"),excel_label("ranked_output_count")])
    ws.append([len(by),len(results),sum(1 for r in results if r.eyes_closed_photo),sum(r.closed_eye_face_count for r in results),len(results)])
    ews=wb.create_sheet(CLUB_SHEET_LABELS["Eye Closure Summary"]); ews.append([excel_label("club"),excel_label("file_name"),excel_label("person_count"),excel_label("closed_eye_faces"),excel_label("eyes_closed_photo")])
    for r in results: ews.append([r.club_name,r.path.name,r.person_count,r.closed_eye_face_count,translate_display_value(r.eyes_closed_photo)])
    fws=wb.create_sheet(CLUB_SHEET_LABELS["Face Detail"]); fws.append([excel_label("club"),excel_label("file_name"),excel_label("face_index"),excel_label("bbox"),excel_label("left_eye_ratio"),excel_label("right_eye_ratio"),excel_label("eye_closed")])
    for r in results:
        for f in r.face_details: fws.append([r.club_name,r.path.name,f.face_index,str(f.bbox),f.left_eye_ratio,f.right_eye_ratio,f.eye_closed])
    rws=wb.create_sheet(CLUB_SHEET_LABELS["Best Shot Ranking"])
    rws.append([excel_label("club"),excel_label("rank"),excel_label("file_name"),excel_label("formality"),excel_label("beauty_score"),excel_label("expression_score"),excel_label("emotion_score"),excel_label("people_count_score"),excel_label("gesture_expression_penalty"),excel_label("is_ng"),excel_label("ng_reason"),excel_label("short_comment"),excel_label("total_score")])
    for r in sorted(results,key=lambda x:(x.club_name,x.rank)):
        ev=r.eval_data
        rws.append([r.club_name,r.rank,r.path.name,ev.get("formality_score"),ev.get("beauty_score"),ev.get("expression_score"),ev.get("emotion_score"),ev.get("people_count_score"),ev.get("gesture_expression_penalty"),translate_display_value(r.ng_flag),translate_display_value(r.ng_reason),translate_display_value(ev.get("short_comment")),r.total_score])
    nws=wb.create_sheet(CLUB_SHEET_LABELS["Rename Output"]); nws.append([excel_label("club"),excel_label("original_file"),excel_label("renamed_file"),excel_label("shooting_date"),excel_label("rank")])
    for r in sorted(results,key=lambda x:(x.club_name,x.rank)): nws.append([r.club_name,r.path.name,r.renamed,r.shooting_date,r.rank])
    ngs=wb.create_sheet("NG写真・要確認")
    ngs.append([excel_label("club"),excel_label("file_name"),excel_label("ng_flag"),excel_label("reason")])
    for r in sorted(results,key=lambda x:(x.club_name,x.rank)):
        if r.ng_flag: ngs.append([r.club_name,r.path.name,translate_display_value(r.ng_flag),translate_display_value(r.ng_reason)])
    wb.save(excel)
    logger.info("Club pipeline output root path=%s", club_out)
    generated_files = [str(p.relative_to(club_out)) for p in sorted(club_out.rglob("*")) if p.is_file()]
    logger.info("Club pipeline generated output files=%s", generated_files)
    # Per-club and per-photo detail surfaced in the job summary so the Review/Output UI
    # (clubReviewBody / clubRenameBody) can render real folders/thumbnails. The landmark
    # eye-detection above already produced everything; reshape it for the frontend.
    summary_rows: list[dict[str, Any]] = []
    rename_rows: list[dict[str, Any]] = []
    for club, rows in by.items():
        ranked = [r for r in rows if not r.ng_flag]
        best = min(rows, key=lambda x: x.rank) if rows else None
        summary_rows.append({
            "group_index": list(by).index(club) + 1,
            "club": club,
            "input_count": len(rows),
            "ranked_count": len(ranked),
            "ng_count": len(rows) - len(ranked),
            "copied_count": len(rows),
            "best_file_name": best.renamed if best else "",
            "best_reason": (best.eval_data.get("short_comment") if best else "") or "",
        })
        for r in sorted(rows, key=lambda x: x.rank):
            rename_rows.append({
                "club": club,
                "rank": r.rank,
                "original_file_name": r.path.name,
                "new_file_name": r.renamed,
                "total_score": r.total_score,
                "is_ng": bool(r.ng_flag),
                "ng_reason": r.ng_reason or "",
                "short_comment": (r.eval_data.get("short_comment") or ""),
            })

    items = []
    for r in sorted(results, key=lambda x: (x.club_name, x.rank)):
        rank_label = f"{RANK_TOKEN}{r.rank:0{RANK_PAD}d}"
        items.append(
            {
                "club_name": r.club_name,
                "original_file": r.path.name,
                "shooting_date": r.shooting_date,
                "rank": r.rank,
                "ai_rank": r.rank,
                "final_rank": r.rank,
                "rank_label": rank_label,
                "ai_rank_label": rank_label,
                "final_rank_label": rank_label,
                "renamed_file": r.renamed,
                "final_renamed_file": r.renamed,
                "ng_flag": r.ng_flag,
                "excluded": False,
                "ng_reason": r.ng_reason,
                "person_count": r.person_count,
                "eyes_closed_photo": r.eyes_closed_photo,
                "closed_eye_face_count": r.closed_eye_face_count,
                "total_score": r.total_score,
                "evaluation": r.eval_data,
                "status": "NG" if r.ng_flag else ("Best Shot" if r.rank == 1 else "Passing"),
                "clean_relative_path": (Path("clean_images") / r.club_name / r.path.name).as_posix(),
                "marked_relative_path": (Path("marked_images") / r.club_name / r.path.name).as_posix(),
                "ranked_relative_path": (Path("ranked_photos") / r.club_name / r.renamed).as_posix(),
                "ranked_marked_relative_path": (Path("ranked_photos_marked") / r.club_name / r.renamed).as_posix(),
            }
        )

    # club_count/photo_count are the original key names; total_clubs/total_images/
    # eyes_closed_count are what the frontend Export screen (clubSummaryBox) reads.
    summary = {
        "club_count": len(by),
        "photo_count": len(results),
        "total_clubs": len(by),
        "total_images": len(results),
        "eyes_closed_count": sum(r.closed_eye_face_count for r in results),
        "ng_count": sum(1 for r in results if r.ng_flag),
        "summary_rows": summary_rows,
        "rename_rows": rename_rows,
    }
    jsonp.write_text(json.dumps({"summary": summary, "items": items}, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"status":"completed","summary":summary,"excel_path":str(excel),"json_path":str(jsonp),"output_dir":str(club_out)}
