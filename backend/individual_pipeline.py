from __future__ import annotations

import csv, json, os, re, shutil, zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any
from dotenv import load_dotenv
from backend.individual.colab_runtime import load_colab_module

IMAGE_EXTS={".jpg",".jpeg",".png",".webp",".bmp",".tif",".tiff",".heic",".heif"}

# manual resolution helpers

def _safe_read_json(path: Path, default: Any) -> Any:
    if not path.exists(): return default
    try: return json.loads(path.read_text(encoding='utf-8'))
    except Exception: return default

def save_error_resolutions(path: Path, resolutions: dict[str, Any]) -> None:
    path.write_text(json.dumps(resolutions, ensure_ascii=False, indent=2), encoding='utf-8')

def load_error_resolutions(path: Path) -> dict[str, Any]:
    return _safe_read_json(path,{})

def resolve_no_card_detected(resolutions: dict[str, Any], error_id: str, class_name: str, student_no: str) -> dict[str, Any]:
    resolutions[error_id]={"action":"assign_student","class_name":class_name,"student_no":student_no}; return resolutions

def resolve_multiple_card_detected(resolutions: dict[str, Any], error_id: str, selected_student_no: str) -> dict[str, Any]:
    resolutions[error_id]={"action":"choose_card","student_no":selected_student_no}; return resolutions

def mark_error_deleted(resolutions: dict[str, Any], error_id: str) -> dict[str, Any]:
    resolutions[error_id]={"action":"deleted"}; return resolutions

def parse_roster_file(roster_file: str | None, output_dir: Path) -> list[dict[str, Any]]:
    if not roster_file: return []
    path=Path(roster_file)
    rows=[]
    try:
        if path.suffix.lower()=='.json':
            data=json.loads(path.read_text(encoding='utf-8'))
            if isinstance(data,list): rows=[x for x in data if isinstance(x,dict)]
        elif path.suffix.lower() in {'.xlsx','.xls'}:
            from openpyxl import load_workbook
            wb=load_workbook(path, data_only=True)
            ws=wb.active
            headers=[str(c.value).strip() if c.value is not None else '' for c in ws[1]]
            for r in ws.iter_rows(min_row=2, values_only=True):
                d={headers[i] if i<len(headers) and headers[i] else f'col_{i}': r[i] for i in range(len(r))}
                rows.append(d)
        elif path.suffix.lower()=='.pdf':
            rows=[{"raw_pdf": path.name, "note": "PDF roster detected; structured parsing deferred."}]
    except Exception as e:
        rows=[{"error":str(e), "roster_file":path.name}]
    (output_dir/'roster.json').write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding='utf-8')
    return rows

def _extract_num(filename:str)->str|None:
    m=re.search(r'(\d{1,3})(?!.*\d)', filename)
    return m.group(1).zfill(2) if m else None

def _class_from_path(p:Path, root:Path)->str|None:
    rel=p.relative_to(root)
    if len(rel.parts)>1:
        top=rel.parts[0]
        if re.search(r'\d', top):
            return top.replace('_','-')
    return None

def apply_error_resolutions_to_class_groups(class_groups:list[dict[str,Any]], error_queue:list[dict[str,Any]], resolutions:dict[str,Any])->list[dict[str,Any]]:
    by={e['error_id']:e for e in error_queue}
    by_gid={g['group_id']:g for g in class_groups}
    for eid,res in resolutions.items():
        err=by.get(eid)
        if not err: continue
        g=by_gid.get(err.get('group_id'))
        if not g: continue
        if res.get('action')=='deleted': g['deleted']=True
        if res.get('class_name'): g['class_name']=res['class_name']
        if res.get('student_no'): g['student_no']=str(res['student_no']).zfill(2)
    return class_groups

def filter_unresolved_error_queue(error_queue:list[dict[str,Any]], resolutions:dict[str,Any])->list[dict[str,Any]]:
    return [e for e in error_queue if e['error_id'] not in resolutions]

def build_exportable_class_groups(class_groups:list[dict[str,Any]], unresolved:list[dict[str,Any]])->list[dict[str,Any]]:
    blocked={e['group_id'] for e in unresolved}
    return [g for g in class_groups if g['group_id'] not in blocked and not g.get('deleted')]

def _run_scaffold_fallback(photos_dir:str, output_dir:str, roster_file:str|None=None, options:dict|None=None)->dict:
    print('[individual] Stage 1: collect images')
    options=options or {}
    photos=Path(photos_dir); out=Path(output_dir); out.mkdir(parents=True, exist_ok=True)
    images=[p for p in sorted(photos.rglob('*')) if p.is_file() and p.suffix.lower() in IMAGE_EXTS]

    print('[individual] Stage 2: parse roster')
    roster=parse_roster_file(roster_file, out) if roster_file else []

    print('[individual] Stage 3: run face grouping / card detection')
    class_groups=[]; error_queue=[]
    class_counter=defaultdict(int)
    for i,p in enumerate(images,1):
        class_name=_class_from_path(p, photos) or 'unclassified'
        student_no=_extract_num(p.stem)
        tag='本01'
        if 'card' in p.stem.lower() or '札' in p.stem: tag='札01'
        if not student_no:
            class_counter[class_name]+=1
            student_no=str(class_counter[class_name]).zfill(2)
            error_queue.append({"error_id":f"e{i:04d}","group_id":f"g{i:04d}","error_type":"student_number_missing","message":f"No student number detected from filename: {p.name}"})
        class_groups.append({"group_id":f"g{i:04d}","source_path":str(p),"class_name":class_name,"student_no":student_no,"tag":tag,"filename":p.name})

    print('[individual] Stage 4: detect errors')
    if not images:
        error_queue.append({"error_id":"e0000","group_id":"none","error_type":"no_images_found","message":"No image files found"})
    if all(g['class_name']=='unclassified' for g in class_groups) and class_groups:
        error_queue.append({"error_id":"e0001","group_id":"all","error_type":"class_not_detected","message":"No class folders/cards detected; exported under unclassified."})

    (out/'error_queue.json').write_text(json.dumps(error_queue, ensure_ascii=False, indent=2), encoding='utf-8')

    resolutions=load_error_resolutions(out/'error_resolutions.json')
    class_groups=apply_error_resolutions_to_class_groups(class_groups,error_queue,resolutions)
    unresolved=filter_unresolved_error_queue(error_queue,resolutions)
    exportable=build_exportable_class_groups(class_groups,unresolved)

    print('[individual] Stage 5: export class folders')
    exported=0
    school=options.get('school_name') or 'school'
    year=(options.get('year') or '26')[-2:]
    for g in exportable:
        cdir=out/g['class_name']; cdir.mkdir(parents=True, exist_ok=True)
        src=Path(g['source_path'])
        class_digits=''.join(ch for ch in g['class_name'] if ch.isdigit()) or '00'
        sid=f"{class_digits}{g['student_no']}"
        dst_name=f"{year}_{school}_{src.stem}_{sid}_{g['tag']}{src.suffix.lower()}"
        shutil.copy2(src, cdir/dst_name)
        exported+=1

    # logs
    (out/'error_log.json').write_text(json.dumps(error_queue, ensure_ascii=False, indent=2), encoding='utf-8')
    with (out/'error_log.csv').open('w',encoding='utf-8',newline='') as f:
        w=csv.DictWriter(f,fieldnames=['error_id','group_id','error_type','message']); w.writeheader(); [w.writerow({k:e.get(k,'') for k in w.fieldnames}) for e in error_queue]

    out_images=[p for p in out.rglob('*') if p.is_file() and p.suffix.lower() in IMAGE_EXTS]
    status='ok'
    if exported==0 and images: status='warning'
    if not images: status='error'
    pipeline_mode='scaffold_fallback'
    summary={
        'status':status,'pipeline_mode':pipeline_mode,
        'total_classes':len({g['class_name'] for g in class_groups}) if class_groups else 0,
        'total_student_groups':len(class_groups),'identified_students':len(class_groups),
        'need_review':len(unresolved),'exported':exported,'unresolved_errors':len(unresolved),
        'error_summary':dict(Counter(e['error_type'] for e in unresolved)),'roster_provided':bool(roster_file),
        'total_images_found':len(images),'images_processed':len(class_groups),'class_folders_found':len({g['class_name'] for g in class_groups if g['class_name']!='unclassified'}),
        'class_groups_detected':len(class_groups),'groups_with_best_shot':len(class_groups),'groups_with_card_images':sum(1 for g in class_groups if g['tag']=='札01'),'groups_with_portraits':sum(1 for g in class_groups if g['tag']!='札01'),
        'exported_files_count':exported,'output_image_files_count':len(out_images)
    }
    (out/'manifest.json').write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding='utf-8')
    return summary


def run_individual_pipeline(photos_dir:str, output_dir:str, roster_file:str|None=None, options:dict|None=None)->dict:
    load_dotenv()
    options = options or {}
    allow_fallback = bool(options.get("allow_fallback", False))
    try:
        m = load_colab_module()
        required = ["parse_roster", "FaceGrouper", "detect_pipeline_errors", "export_all_classes"]
        if not all(hasattr(m, x) for x in required):
            raise RuntimeError(f"Missing required symbols: {required}")

        out = Path(output_dir)
        out.mkdir(parents=True, exist_ok=True)
        print("PHASE 1: Scanning images for class separators")
        school_name = options.get("school_name")
        year = options.get("year")
        openai_api_key = os.getenv("OPENAI_API_KEY")
        roster = m.parse_roster(roster_file, school_name=school_name, openai_api_key=openai_api_key) if roster_file else []
        if hasattr(m, "save_roster_json"):
            m.save_roster_json(roster, str(out / "roster.json"))
        else:
            (out / "roster.json").write_text(json.dumps(roster, ensure_ascii=False, indent=2), encoding="utf-8")

        print("PHASE 2: Detecting faces + reading cards")
        valid_numbers = set()
        scoring = options.get("scoring", "local")
        grouper = m.FaceGrouper(valid_numbers=valid_numbers, scoring=scoring, openai_client=None, roster=roster)
        class_groups = grouper.process_folder(photos_dir)

        print("PHASE 3: Building person groups from face clusters")
        error_queue = m.detect_pipeline_errors(class_groups, roster, photos_dir) if hasattr(m, "detect_pipeline_errors") else []
        (out / "error_queue.json").write_text(json.dumps(error_queue, ensure_ascii=False, indent=2), encoding="utf-8")

        print("PHASE 4: Assigning groups to classes")
        resolutions = load_error_resolutions(out / "error_resolutions.json")
        if hasattr(m, "apply_error_resolutions_to_class_groups"):
            class_groups = m.apply_error_resolutions_to_class_groups(class_groups, resolutions)
        unresolved = m.filter_unresolved_error_queue(error_queue, resolutions) if hasattr(m, "filter_unresolved_error_queue") else error_queue
        exportable = m.build_exportable_class_groups(class_groups, unresolved) if hasattr(m, "build_exportable_class_groups") else class_groups

        print("PHASE 5: Exporting class folders")
        if hasattr(m, "export_error_items"):
            m.export_error_items(unresolved, str(out))
        if hasattr(m, "write_error_log_json"):
            m.write_error_log_json(error_queue, str(out / "error_log.json"))
        else:
            (out / "error_log.json").write_text(json.dumps(error_queue, ensure_ascii=False, indent=2), encoding="utf-8")
        if hasattr(m, "write_error_log_csv"):
            m.write_error_log_csv(error_queue, str(out / "error_log.csv"))
        else:
            with (out / "error_log.csv").open("w", encoding="utf-8", newline="") as f:
                w = csv.DictWriter(f, fieldnames=["error_id", "group_id", "error_type", "message"]); w.writeheader()
                for e in error_queue: w.writerow({k:e.get(k,"") for k in w.fieldnames})

        class_mapping = options.get("class_mapping")
        max_backups = int(options.get("max_backups", 5))
        m.export_all_classes(
            exportable,
            roster,
            str(out),
            school_name=school_name,
            year=year,
            class_mapping=class_mapping,
            max_backups=max_backups,
        )
        if hasattr(m, "process_manifest_offsets"):
            try:
                m.process_manifest_offsets(package_root=str(out))
            except Exception:
                pass

        out_images=[p for p in out.rglob("*") if p.is_file() and p.suffix.lower() in IMAGE_EXTS]
        summary = {
            "status": "ok" if out_images else "warning",
            "pipeline_mode": "real_colab",
            "total_classes": len({g.get("class_name","unclassified") for g in class_groups}) if isinstance(class_groups, list) else 0,
            "total_student_groups": len(class_groups) if isinstance(class_groups, list) else 0,
            "identified_students": len(class_groups) if isinstance(class_groups, list) else 0,
            "need_review": len(unresolved) if isinstance(unresolved, list) else 0,
            "exported": len(out_images),
            "unresolved_errors": len(unresolved) if isinstance(unresolved, list) else 0,
            "error_summary": dict(Counter(e.get("error_type","unknown") for e in unresolved)) if isinstance(unresolved, list) else {},
            "roster_provided": bool(roster_file),
        }
        (out / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
        return summary
    except Exception as e:
        reason = str(e)
        dep_names = ["mediapipe", "easyocr", "insightface", "onnxruntime"]
        missing = next((d for d in dep_names if d in reason.lower()), None)
        if allow_fallback:
            print(f"[individual] real_colab path failed, fallback scaffold: {reason}")
            s = _run_scaffold_fallback(photos_dir, output_dir, roster_file, options)
            s["pipeline_mode"] = "scaffold_fallback"
            s["reason"] = reason
            return s
        return {
            "status": "error",
            "pipeline_mode": "real_colab_failed",
            "reason": f"missing dependency: {missing}" if missing else reason,
            "exported": 0,
            "unresolved_errors": 0,
            "error_summary": {},
        }

def zip_output_dir(output_dir:Path, zip_path:Path)->Path:
    if zip_path.exists(): zip_path.unlink()
    with zipfile.ZipFile(zip_path,'w',zipfile.ZIP_DEFLATED) as zf:
        for p in output_dir.rglob('*'):
            if p.is_file(): zf.write(p,p.relative_to(output_dir))
    return zip_path
