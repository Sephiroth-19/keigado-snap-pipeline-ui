from __future__ import annotations

"""
card_detector.py – Card detection + OCR for school photography pipeline
========================================================================

Simple approach:
  1. Find the white card region in the image
  2. Determine what kind of card it is:
     - Class separator (e.g. "3-1") → marks start of a new class
     - Teacher card (Japanese text: name, class, role) → teacher photo
     - Student number card (digit 1-50) → student attendance number
  3. Return the result

That's it. No multi-layer validation. The images are high quality and
the task is straightforward.
"""

import json
import logging
import os
import re
import threading
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

import cv2
import numpy as np

logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────
# Constants
# ──────────────────────────────────────────────────────────────────

# Regex: class separator "3-1", "3 - 3", "3=7", "3–2"
CLASS_SEP_RE = re.compile(r"(\d)\s*[-=\u2013\u2014]\s*(\d+)")

# Full-width → half-width digits/hyphen
_FW2HW = str.maketrans("０１２３４５６７８９－", "0123456789-")

# Number regex
_NUMBER_RE = re.compile(r"\d+")

# Filter template text on teacher cards
_INSTRUCTION_RE = re.compile(
    r"フルネーム|ご記入ください|記入してください"
)

# Japanese name regex (before 先生)
_NAME_BEFORE_SENSEI_RE = re.compile(
    r"([\u4e00-\u9fff][\u4e00-\u9fff\s\u3000]{1,9}[\u4e00-\u9fff])\s*先生|"
    r"([\u4e00-\u9fff]{2,10})\s*先生"
)


# ──────────────────────────────────────────────────────────────────
# Data classes
# ──────────────────────────────────────────────────────────────────

@dataclass
class OCRCandidate:
    value: int
    confidence: float
    bbox: list | None = None
    source: str = "original"


@dataclass
class ClassSeparatorInfo:
    grade: int
    class_number: int
    raw_text: str = ""
    confidence: float = 0.0
    student_number: int | None = None

    @property
    def label(self) -> str:
        return f"{self.grade}-{self.class_number}"


@dataclass
class TeacherInfo:
    name: str = ""
    class_label: str = ""
    role: str = ""
    raw_text: str = ""
    confidence: float = 0.0


@dataclass
class DetectionResult:
    image_path: str
    has_card: bool = False
    is_class_separator: bool = False
    class_separator: ClassSeparatorInfo | None = None
    is_teacher_card: bool = False
    teacher_info: TeacherInfo | None = None
    accepted_number: int | None = None
    all_candidates: list[OCRCandidate] = field(default_factory=list)
    validation_layer: str = ""
    review_needed: bool = False
    notes: str = ""


# ──────────────────────────────────────────────────────────────────
# Image I/O (Unicode-safe for Windows)
# ──────────────────────────────────────────────────────────────────

def imread_safe(path: str) -> np.ndarray | None:
    try:
        buf = np.fromfile(path, dtype=np.uint8)
        return cv2.imdecode(buf, cv2.IMREAD_COLOR)
    except Exception:
        return None


# ──────────────────────────────────────────────────────────────────
# Image helpers
# ──────────────────────────────────────────────────────────────────

def _resize(img: np.ndarray, max_dim: int = 1600) -> np.ndarray:
    h, w = img.shape[:2]
    s = min(1.0, max_dim / max(h, w))
    if s < 1.0:
        img = cv2.resize(img, (int(w * s), int(h * s)), interpolation=cv2.INTER_AREA)
    return img


def detect_card_region(img: np.ndarray) -> tuple[np.ndarray, bool]:
    """Find the white card in the lower portion of the image."""
    h, w = img.shape[:2]
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)

    # White mask (tight: sat < 40, val > 170)
    mask = cv2.inRange(hsv, (0, 0, 170), (180, 40, 255))
    k = np.ones((15, 15), np.uint8)
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, k)
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, k)

    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    best, best_area = None, 0
    for cnt in contours:
        area = cv2.contourArea(cnt)
        if area < h * w * 0.008 or area > h * w * 0.25:
            continue
        x, y, cw, ch = cv2.boundingRect(cnt)
        aspect = cw / ch if ch else 0
        if not (1.2 < aspect < 3.5):
            continue
        if y < h * 0.30:
            continue
        margin = w * 0.05
        if x < margin or (x + cw) > (w - margin):
            continue
        rect_fill = area / (cw * ch) if cw * ch else 0
        if rect_fill < 0.65:
            continue
        if area > best_area:
            best_area = area
            best = (x, y, cw, ch)

    if best:
        x, y, cw, ch = best
        pad = 12
        crop = img[max(0, y - pad):min(h, y + ch + pad),
                    max(0, x - pad):min(w, x + cw + pad)]
        return crop, True

    # Fallback: lower-centre crop
    return img[int(h * 0.45):int(h * 0.85), int(w * 0.10):int(w * 0.90)], False


def _prepare_card(card_img: np.ndarray) -> list[tuple[str, np.ndarray]]:
    """Return original + enhanced versions for OCR."""
    h, w = card_img.shape[:2]
    if h < 120:
        s = 120 / h
        card_img = cv2.resize(card_img, (int(w * s), int(h * s)),
                               interpolation=cv2.INTER_CUBIC)

    gray = cv2.cvtColor(card_img, cv2.COLOR_BGR2GRAY)
    enhanced = cv2.convertScaleAbs(gray, alpha=2.0, beta=0)

    return [
        ("original", card_img),
        ("enhanced", cv2.cvtColor(enhanced, cv2.COLOR_GRAY2BGR)),
    ]


# ──────────────────────────────────────────────────────────────────
# OCR engines
# ──────────────────────────────────────────────────────────────────

class _OCREngine:
    """EasyOCR wrapper for reading numbers (English digits)."""

    def __init__(self):
        self._reader = None
        self._lock = threading.Lock()

    def _ensure(self):
        if self._reader is not None:
            return
        with self._lock:
            if self._reader is not None:
                return
            import easyocr
            try:
                import torch
                gpu = torch.cuda.is_available()
            except ImportError:
                gpu = False
            self._reader = easyocr.Reader(["en"], gpu=gpu, verbose=False)
            logger.info("OCR backend: EasyOCR")

    def read_numbers(self, img_bgr: np.ndarray) -> list[OCRCandidate]:
        """OCR the image for digits, return candidates."""
        self._ensure()
        candidates = []
        try:
            results = self._reader.readtext(img_bgr, detail=1, allowlist="0123456789")
            for bbox, text, conf in results:
                for ns in _NUMBER_RE.findall(text):
                    val = int(ns)
                    if val > 0:
                        candidates.append(OCRCandidate(val, float(conf), bbox))
        except Exception as e:
            logger.error("OCR failed: %s", e)
        return candidates

    def read_texts(self, img_bgr: np.ndarray) -> list[tuple[str, float]]:
        """OCR the image for any text, return (text, conf) list."""
        self._ensure()
        try:
            results = self._reader.readtext(img_bgr, detail=1)
            return [(text, conf) for _, text, conf in results]
        except Exception:
            return []


class _OCREngineJapanese:
    """EasyOCR wrapper for reading Japanese text (teacher cards)."""

    def __init__(self):
        self._reader = None

    def _ensure(self):
        if self._reader is not None:
            return
        import easyocr
        try:
            import torch
            gpu = torch.cuda.is_available()
        except ImportError:
            gpu = False
        self._reader = easyocr.Reader(["ja", "en"], gpu=gpu, verbose=False)
        logger.info("OCR (Japanese): EasyOCR [ja,en]")

    def read_texts(self, img_bgr: np.ndarray) -> list[tuple[str, float]]:
        """Run Japanese OCR, return (text, confidence) list."""
        self._ensure()
        try:
            results = self._reader.readtext(img_bgr, detail=1)
            return [(text, float(conf)) for _, text, conf in results]
        except Exception:
            return []


# Singletons
_ocr = _OCREngine()
_ocr_ja = _OCREngineJapanese()



# ──────────────────────────────────────────────────────────────────
# Detection: class separator
# ──────────────────────────────────────────────────────────────────

def detect_class_separator(img: np.ndarray) -> ClassSeparatorInfo | None:
    """Check if the image is a class separator card like '3-1'."""
    h, w = img.shape[:2]
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    mask = cv2.inRange(hsv, (0, 0, 180), (180, 40, 255))
    white_ratio = mask.sum() / 255 / (h * w)
    if white_ratio < 0.04:
        return None

    # Find largest white contour → crop → OCR
    k = np.ones((15, 15), np.uint8)
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, k)
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, k)
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    best_crop, best_area = None, 0
    for cnt in contours:
        area = cv2.contourArea(cnt)
        if area < h * w * 0.02 or area > h * w * 0.50:
            continue
        x, y, cw, ch = cv2.boundingRect(cnt)
        aspect = cw / ch if ch else 0
        if not (1.5 < aspect < 6.0):
            continue
        if area > best_area:
            best_area = area
            pad = 10
            best_crop = img[max(0, y - pad):min(h, y + ch + pad),
                            max(0, x - pad):min(w, x + cw + pad)]

    if best_crop is None:
        best_crop = img[int(h * 0.25):int(h * 0.75), int(w * 0.15):int(w * 0.85)]

    texts = _ocr.read_texts(best_crop)

    sep = None
    student_num = None
    for text, conf in texts:
        clean = text.strip()
        m = CLASS_SEP_RE.search(clean)
        if m and sep is None:
            g, c = int(m.group(1)), int(m.group(2))
            if 1 <= g <= 6 and 1 <= c <= 20:
                sep = ClassSeparatorInfo(grade=g, class_number=c,
                                         raw_text=clean, confidence=conf)
        else:
            for ns in re.findall(r"\b(\d{1,2})\b", clean):
                v = int(ns)
                if 1 <= v <= 50:
                    if student_num is None or v < student_num:
                        student_num = v

    if sep and student_num:
        sep.student_number = student_num
    return sep


# ──────────────────────────────────────────────────────────────────
# Detection: teacher card
# ──────────────────────────────────────────────────────────────────

def detect_teacher_card(card_crop: np.ndarray) -> TeacherInfo | None:
    """
    Run Japanese OCR on the card crop. If it contains teacher keywords
    (先生, 担任, 教科+class), parse the teacher info.
    """
    # OCR on original + enhanced
    variants = _prepare_card(card_crop)
    seen: set[str] = set()
    texts: list[tuple[str, float]] = []

    for _, vimg in variants:
        for text, conf in _ocr_ja.read_texts(vimg):
            k = text.strip()
            if not k or conf < 0.10:
                continue
            if k not in seen:
                seen.add(k)
                texts.append((text, conf))

    if not texts:
        return None

    # Filter instruction lines
    filtered = [(t, c) for t, c in texts if not _INSTRUCTION_RE.search(t)]
    full = " ".join(t.strip() for t, _ in filtered)

    # Check for teacher keywords
    has_sensei = "先生" in full
    has_tanin = "担任" in full
    has_kyouka_class = False
    if "教科" in full:
        for t, _ in filtered:
            clean = t.strip().translate(_FW2HW)
            clean = re.sub(r"^教科\s*", "", clean)
            if CLASS_SEP_RE.search(clean):
                has_kyouka_class = True
                break

    if not (has_sensei or has_tanin or has_kyouka_class):
        return None

    info = TeacherInfo()

    # Class (e.g. 3-1)
    for t, c in filtered:
        clean = t.strip().translate(_FW2HW)
        clean = re.sub(r"^教科\s*", "", clean)
        m = CLASS_SEP_RE.search(clean)
        if m:
            g, cn = int(m.group(1)), int(m.group(2))
            if 1 <= g <= 6 and 1 <= cn <= 20:
                info.class_label = f"{g}-{cn}"
                break

    # Role (副担任 before 担任)
    for t, c in filtered:
        if "副担任" in t:
            info.role = "副担任"
            break
        if "担任" in t:
            info.role = "担任"
            break

    # Name extraction
    name_candidates: list[str] = []

    # Strategy 1: text before 先生 on same line
    for t, _ in filtered:
        if "先生" in t:
            part = t.split("先生")[0].strip()
            if len(re.findall(r"[\u4e00-\u9fff]", part)) >= 2:
                name_candidates.append(re.sub(r"[\s\u3000]+", " ", part))

    # Strategy 2: previous line when 先生 is separate box
    for i, (t, _) in enumerate(filtered):
        if "先生" in t:
            before = t.split("先生")[0].strip()
            if len(re.findall(r"[\u4e00-\u9fff]", before)) < 2 and i > 0:
                prev = filtered[i - 1][0].strip()
                if (len(re.findall(r"[\u4e00-\u9fff]", prev)) >= 2
                        and 2 <= len(prev) <= 15
                        and not re.search(r"教科|担任|先生", prev)):
                    name_candidates.append(re.sub(r"[\s\u3000]+", " ", prev))

    # Strategy 3: regex
    m = _NAME_BEFORE_SENSEI_RE.search(full)
    if m:
        name = (m.group(1) or m.group(2) or "").strip()
        if name:
            name_candidates.append(re.sub(r"[\s\u3000]+", " ", name))

    # Strategy 4: longest kanji-rich line
    for t, _ in filtered:
        s = t.strip()
        if re.search(r"教科|担任|先生", s):
            continue
        if len(re.findall(r"[\u4e00-\u9fff]", s)) >= 2 and 2 <= len(s) <= 12:
            name_candidates.append(re.sub(r"[\s\u3000]+", " ", s))

    if name_candidates:
        best = max(name_candidates, key=len)
        # Reject garbage (has ASCII, digits, weird symbols)
        if not re.search(r"[_|\\/*+?!@#$%^&~`\[\]{}=<>0-9a-zA-Z]", best):
            info.name = best

    if info.name or info.class_label or info.role:
        info.confidence = sum(c for _, c in filtered) / max(len(filtered), 1)
        return info
    return None


# ──────────────────────────────────────────────────────────────────
# Main detector
# ──────────────────────────────────────────────────────────────────

class CardDetector:
    """
    Detects cards in portrait photos:
      - Class separator cards (e.g. '3-1')
      - Teacher cards (Japanese name + class + role)
      - Student number cards (digit)
    """

    def __init__(
        self,
        valid_numbers: set[int] | None = None,
        absent_numbers: set[int] | None = None,
        ocr_backend: str = "easyocr",
    ):
        self.valid_numbers = valid_numbers or set()
        self.absent_numbers = absent_numbers or set()
        self.ocr_backend = ocr_backend
        self._prev_number: int | None = None
        self._review_queue: list[dict] = []

    def reset_sequence(self):
        self._prev_number = None

    @property
    def review_queue(self) -> list[dict]:
        return self._review_queue

    def detect(self, image_path: str) -> DetectionResult:
        """
        Analyse one image. Returns what's on the card:
          - Class separator → result.is_class_separator + result.class_separator
          - Teacher → result.is_teacher_card + result.teacher_info
          - Student number → result.accepted_number
          - No card / portrait → result.has_card = False
        """
        result = DetectionResult(image_path=image_path)

        img = imread_safe(image_path)
        if img is None:
            result.notes = "Could not read image"
            return result

        img = _resize(img)
        fname = os.path.basename(image_path)

        if self.ocr_backend == "none":
            result.notes = "OCR disabled"
            return result

        # 1. Class separator?
        sep = detect_class_separator(img)
        if sep is not None:
            result.has_card = True
            result.is_class_separator = True
            result.class_separator = sep
            if sep.student_number is not None:
                result.accepted_number = sep.student_number
                result.notes = f"Class {sep.label} + student #{sep.student_number}"
                logger.info("  🏫 CLASS %s + student #%d  ← %s",
                            sep.label, sep.student_number, fname)
            else:
                result.notes = f"Class separator: {sep.label}"
                logger.info("  🏫 Class separator: %s  ← %s", sep.label, fname)
            return result

        # 2. Find card region
        card_crop, card_found = detect_card_region(img)
        result.has_card = card_found

        # 3. Teacher card?
        try:
            teacher = detect_teacher_card(card_crop)
            if teacher is not None:
                result.has_card = True
                result.is_teacher_card = True
                result.teacher_info = teacher
                result.notes = (f"Teacher: {teacher.name or '?'} "
                                f"class {teacher.class_label or '?'} "
                                f"{teacher.role or ''}")
                logger.info("  📋 Teacher %s [%s] %s  ← %s",
                            teacher.name or "?", teacher.class_label or "?",
                            teacher.role, fname)
                return result
        except Exception as e:
            logger.debug("Teacher OCR unavailable: %s", e)

        # 4. Student number — OCR the card crop
        all_candidates: list[OCRCandidate] = []
        for vname, vimg in _prepare_card(card_crop):
            for c in _ocr.read_numbers(vimg):
                c.source = vname
                all_candidates.append(c)

        result.all_candidates = all_candidates

        if not all_candidates:
            result.notes = "No numbers found"
            result.has_card = False
            return result

        # If no card was found by contour detection, require high confidence
        best_conf = max(c.confidence for c in all_candidates)
        threshold = 0.50 if card_found else 0.80
        if best_conf < threshold:
            logger.debug("  ↷ OCR conf %.2f < %.2f — portrait: %s",
                         best_conf, threshold, fname)
            result.has_card = False
            result.notes = f"Low confidence ({best_conf:.2f})"
            return result

        # Pick best candidate: highest confidence, prefer 1-50 range
        good = [c for c in all_candidates if c.confidence >= threshold]
        if not good:
            good = all_candidates

        # Prefer values in valid roster if available
        if self.valid_numbers:
            in_roster = [c for c in good if c.value in self.valid_numbers]
            if in_roster:
                good = in_roster

        # Prefer values in 1-50 range
        in_range = [c for c in good if 1 <= c.value <= 50]
        if in_range:
            good = in_range

        winner = max(good, key=lambda c: c.confidence)
        result.accepted_number = winner.value
        result.has_card = True
        result.validation_layer = "best-confidence"

        # Track sequence
        if result.accepted_number is not None:
            self._prev_number = result.accepted_number

        self._review_queue.append({
            "image": image_path,
            "candidates": [{"value": c.value, "confidence": round(c.confidence, 3),
                            "source": c.source} for c in all_candidates],
            "accepted": result.accepted_number,
            "previous_number": self._prev_number,
            "status": "auto",
        })

        return result

    def try_read_placard_number(
        self,
        image_path: str,
        current_number: int | None,
    ) -> int | None:
        """If a portrait shows a card, try to read the number on it."""
        if self.ocr_backend == "none":
            return None
        img = imread_safe(image_path)
        if img is None or not self.has_placard(img):
            return None
        img = _resize(img)
        card_crop, _ = detect_card_region(img)
        candidates = []
        for _, vimg in _prepare_card(card_crop):
            candidates.extend(_ocr.read_numbers(vimg))
        if not candidates:
            return None
        valid = [c for c in candidates if 1 <= c.value <= 50 and c.confidence >= 0.20]
        if self.valid_numbers:
            roster_valid = [c for c in valid if c.value in self.valid_numbers]
            if roster_valid:
                valid = roster_valid
        if not valid:
            return None
        best = max(valid, key=lambda c: c.confidence)
        if current_number is not None and best.value == current_number:
            return None
        return best.value

    @staticmethod
    def has_placard(img) -> bool:
        """Quick CV check: does image contain a handheld white card?"""
        if isinstance(img, (str, Path)):
            img = imread_safe(str(img))
            if img is None:
                return False
        h, w = img.shape[:2]
        hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
        mask = cv2.inRange(hsv, (0, 0, 190), (180, 35, 255))
        k = np.ones((20, 20), np.uint8)
        mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, k)
        mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, k)
        contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        for cnt in contours:
            area = cv2.contourArea(cnt)
            if not (h * w * 0.015 < area < h * w * 0.20):
                continue
            x, y, cw, ch = cv2.boundingRect(cnt)
            aspect = cw / ch if ch else 0
            if not (0.5 < aspect < 2.5):
                continue
            if y < h * 0.35:
                continue
            margin_x = w * 0.07
            if x < margin_x or (x + cw) > (w - margin_x):
                continue
            if (y + ch) > h * 0.95:
                continue
            fill = area / (cw * ch) if cw * ch else 0
            if fill < 0.55:
                continue
            return True
        return False

    def save_review_queue(self, path: str | Path):
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(self._review_queue, f, ensure_ascii=False, indent=2)
        logger.info("Saved %d review items → %s", len(self._review_queue), path)

    def load_review_resolutions(self, path: str | Path):
        path = Path(path)
        if not path.exists():
            return {}
        with open(path, encoding="utf-8") as f:
            items = json.load(f)
        return {
            item["image"]: item["resolved_number"]
            for item in items
            if item.get("status") == "resolved" and "resolved_number" in item
        }

"""### Module 2 — `photo_scoring`
This module defines the core data structures and scoring mechanisms for evaluating portrait photos. It includes:
- **Data Classes**: `ScoredPhoto`, `StudentPhotoGroup`, `TeacherPhotoGroup`, and `ClassPhotoGroup` for organizing photo metadata, student/teacher details, and class information.
- **Portrait Scoring**:
  - `score_photo_local`: Utilizes CV heuristics based on `InsightFace` (buffalo_l model) for face detection, pose analysis, eye openness (Eye Aspect Ratio), sharpness, brightness, and gesture detection. It also includes a Haar cascade fallback.
  - `score_photo_openai`: Leverages the `GPT-4o` vision model for advanced scoring, assessing formality, suitability, beauty, expression, and identifying 'NG' conditions. This method requires an OpenAI API key.
"""

"""
photo_scoring.py – Data classes and portrait scoring
=====================================================

Contents:
  - `ScoredPhoto`, `StudentPhotoGroup`, `TeacherPhotoGroup`, `ClassPhotoGroup`
  - `score_photo_local(path)`  : InsightFace primary scorer + Haar fallback
  - `score_photo_openai(path)` : GPT-4o vision scorer (optional)
"""


import logging
import re
from dataclasses import dataclass, field
from typing import Any

import cv2
import numpy as np



logger = logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────────────
# Data structures
# ──────────────────────────────────────────────────────────────────

@dataclass
class ScoredPhoto:
    path: str
    score: float = 0.0
    is_ng: bool = False
    ng_reason: str = ""
    comment: str = ""


@dataclass
class StudentPhotoGroup:
    """All photos for one student from a single shooting session."""
    attendance_number: int | None = None
    card_images: list[str] = field(default_factory=list)  # paths to 札 shots (may be multiple)
    portraits: list[ScoredPhoto] = field(default_factory=list)
    best_shot: ScoredPhoto | None = None


@dataclass
class TeacherPhotoGroup:
    """All photos for one teacher (先生 card + portraits)."""
    teacher_info: TeacherInfo | None = None
    card_images: list[str] = field(default_factory=list)  # paths to 札 shots (may be multiple)
    portraits: list[ScoredPhoto] = field(default_factory=list)
    best_shot: ScoredPhoto | None = None


@dataclass
class ClassPhotoGroup:
    """
    All student groups for one class, identified by a class separator card.
    May include a teacher photo group if a teacher card was shot for this class.
    """
    grade: int | None = None
    class_number: int | None = None
    class_label: str = ""           # e.g. "3-7" or derived letter "A"
    separator_image: str | None = None
    students: list[StudentPhotoGroup] = field(default_factory=list)
    teacher: TeacherPhotoGroup | None = None

    @property
    def id(self) -> str:
        """Unique ID like '3-7' or 'unknown'."""
        if self.grade is not None and self.class_number is not None:
            return f"{self.grade}-{self.class_number}"
        return self.class_label or "unknown"


# ──────────────────────────────────────────────────────────────────
# InsightFace scorer  (lazy singleton — shared model, no circular import)
# ──────────────────────────────────────────────────────────────────

class _IFScorer:
    """
    Lazy-loaded InsightFace scorer.
    Uses buffalo_l to get pose (pitch/yaw/roll) and 2D-106 landmarks
    for every face in an image.
    Kept separate from face_grouper._FaceEngine to avoid circular imports.
    """

    def __init__(self):
        self._app = None
        self._available: bool | None = None   # None = not yet tried

    def _ensure(self) -> bool:
        """Try to load the model; return True on success."""
        if self._available is not None:
            return self._available
        try:
            from insightface.app import FaceAnalysis
            self._app = FaceAnalysis(
                name="buffalo_l",
                providers=["CUDAExecutionProvider", "CPUExecutionProvider"],
            )
            self._app.prepare(ctx_id=0, det_size=(640, 640))
            self._available = True
            logger.debug("InsightFace scorer ready (buffalo_l).")
        except Exception as e:
            logger.warning("InsightFace unavailable for scoring: %s", e)
            self._available = False
        return self._available

    def score_face(self, img: np.ndarray) -> dict | None:
        """
        Run InsightFace on *img* and return a dict with quality signals
        for the **largest / most confident** face found.

        Returns
        -------
        dict with keys:
          det_score   float   detection confidence 0-1
          pose        list    [pitch, yaw, roll] in degrees
          lmk         ndarray (106, 2) or None
          bbox        list    [x1, y1, x2, y2]
        or None if no face detected or model unavailable.
        """
        if not self._ensure():
            return None
        try:
            faces = self._app.get(img)
        except Exception as e:
            logger.debug("InsightFace.get() failed: %s", e)
            return None
        if not faces:
            return None
        # Pick the face with the highest detection score
        face = max(faces, key=lambda f: float(getattr(f, "det_score", 0)))
        return {
            "det_score": float(getattr(face, "det_score", 1.0)),
            "pose":      list(getattr(face, "pose",      [0.0, 0.0, 0.0])),
            "lmk":       getattr(face, "landmark_2d_106", None),
            "bbox":      face.bbox.tolist(),
        }


_if_scorer = _IFScorer()


# ── InsightFace-based signal helpers ──────────────────────────────

# Eye landmark indices inside the 2D-106 point model (buffalo_l).
# Left  eye outline: 33-42  Right eye outline: 87-96
_LEFT_EYE_IDX  = list(range(33, 43))
_RIGHT_EYE_IDX = list(range(87, 97))


def _eye_aspect_ratio(lmk: np.ndarray, indices: list[int]) -> float:
    """
    Eye Aspect Ratio (EAR) from a subset of 2D-106 landmarks.
    EAR = vertical_span / horizontal_span.
    Open eye: ~0.25-0.50.  Closed eye: < 0.15.
    """
    pts = lmk[indices]           # (N, 2)
    w = float(pts[:, 0].max() - pts[:, 0].min())
    h = float(pts[:, 1].max() - pts[:, 1].min())
    if w < 1:
        return 0.0
    return h / w


def _eye_close_penalty(lmk: np.ndarray | None) -> float:
    """
    Returns 0.0 (eyes open) → 1.0 (both eyes clearly closed).
    Uses EAR on the 2D-106 landmark eye regions.

    Calibrated from real portrait data:
      Open eyes   : avg EAR ≈ 0.31 – 0.41  → 0 penalty
      Squinting   : avg EAR ≈ 0.20 – 0.30  → partial penalty
      Closed eyes : avg EAR ≤ 0.16         → NG-level penalty

    Thresholds
    ----------
    open_thr   = 0.30  → EAR ≥ 0.30 gives 0 penalty
    closed_thr = 0.16  → EAR ≤ 0.16 gives full penalty (1.0)
    """
    if lmk is None or lmk.shape[0] < 97:
        return 0.0
    left_ear  = _eye_aspect_ratio(lmk, _LEFT_EYE_IDX)
    right_ear = _eye_aspect_ratio(lmk, _RIGHT_EYE_IDX)
    avg_ear   = (left_ear + right_ear) / 2.0

    OPEN   = 0.30
    CLOSED = 0.16
    if avg_ear >= OPEN:
        return 0.0
    return float(np.clip((OPEN - avg_ear) / (OPEN - CLOSED), 0.0, 1.0))


def _pose_penalty(pose: list[float]) -> tuple[float, list[str]]:
    """
    Derive a 0-1 penalty and flag strings from [pitch, yaw, roll].

    Thresholds (degrees)
    --------------------
    pitch  : +ve = looking down, -ve = looking up
      > 18  → head pitched forward (like the eyes-closed smile shot)
      < -15 → head tilted back
    yaw    : |yaw| > 25 → turned sideways
    roll   : |roll| > 20 → head tilt
    """
    pitch, yaw, roll = pose[0], pose[1], pose[2]
    flags: list[str] = []
    penalties: list[float] = []

    # Pitch (looking down)
    if pitch > 18:
        p = float(np.clip((pitch - 18) / 20, 0.0, 1.0))
        penalties.append(p)
        flags.append(f"pitch-down({pitch:.0f}°)")
    elif pitch < -15:
        p = float(np.clip((-pitch - 15) / 20, 0.0, 1.0))
        penalties.append(p * 0.5)   # looking up is less bad
        flags.append(f"pitch-up({pitch:.0f}°)")

    # Yaw (turned sideways)
    abs_yaw = abs(yaw)
    if abs_yaw > 25:
        p = float(np.clip((abs_yaw - 25) / 25, 0.0, 1.0))
        penalties.append(p)
        flags.append(f"yaw({yaw:.0f}°)")

    # Roll (head tilt)
    abs_roll = abs(roll)
    if abs_roll > 20:
        p = float(np.clip((abs_roll - 20) / 20, 0.0, 1.0)) * 0.6
        penalties.append(p)
        flags.append(f"roll({roll:.0f}°)")

    combined = float(min(1.0, sum(penalties)))  # accumulate, cap at 1
    return combined, flags


# ──────────────────────────────────────────────────────────────────
# Local scoring heuristics (no API needed)
# ──────────────────────────────────────────────────────────────────

def _sharpness(gray: np.ndarray) -> float:
    """Laplacian variance — higher = sharper."""
    return cv2.Laplacian(gray, cv2.CV_64F).var()


def _brightness(gray: np.ndarray) -> float:
    """Mean pixel value (0-255)."""
    return float(gray.mean())


def _detect_face_bbox_haar(
    img: np.ndarray,
) -> tuple[int, int, int, int] | None:
    """
    Run Haar frontal-face cascade and return the largest detected face as
    (x, y, w, h), or None if no face found.
    """
    try:
        cascade_path = cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
        face_cascade = cv2.CascadeClassifier(cascade_path)
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.1, 5, minSize=(60, 60))
        if len(faces) == 0:
            return None
        return tuple(max(faces, key=lambda f: f[2] * f[3]))
    except Exception:
        return None


def _skin_mask(img: np.ndarray) -> np.ndarray:
    """
    HSV-based skin-colour mask that covers Asian and general skin tones.
    Returns an 8-bit single-channel binary mask.
    """
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    # Primary warm range (yellow → red skin)
    m1 = cv2.inRange(
        hsv,
        np.array([0,  15,  80], np.uint8),
        np.array([25, 170, 255], np.uint8),
    )
    # Wraparound reddish tones
    m2 = cv2.inRange(
        hsv,
        np.array([165, 15,  80], np.uint8),
        np.array([180, 170, 255], np.uint8),
    )
    mask = cv2.bitwise_or(m1, m2)
    # Remove small noise
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel)
    return mask


def _gesture_penalty(
    img: np.ndarray,
    face_bbox: tuple[int, int, int, int] | None,
) -> float:
    """
    Detect extended arms / raised hands outside the face+neck region.
    Returns 0.0 (no gesture) → 1.0 (strong gesture).

    Strategy
    --------
    1. Build an HSV skin mask for the whole image.
    2. Zero-out the face + neck bounding box so the face itself doesn't
       contribute.
    3. Two complementary checks in the upper 70 % of the frame:
       a) Area penalty  – total non-face skin fraction (catches bare arms,
          shoulders, multiple exposed skin regions).
       b) Isolated-blob penalty – even a small isolated skin blob (e.g. a
          single fist or hand, even partially sleeved) signals a gesture;
          the largest disconnected blob triggers this penalty.
    4. Return max(area_penalty, blob_penalty) capped at 1.0.

    Profiles that get penalised:
      - Pointing at camera  → fist/hand as isolated blob in centre frame
      - Raised arm          → skin blob above or beside the face
      - Bare arm reach      → large skin area streak in the ROI
      - Sleeved arm gesture → isolated hand/fist blob detection
    """
    h, w = img.shape[:2]
    mask = _skin_mask(img)

    if face_bbox is not None:
        fx, fy, fw, fh = face_bbox
        # Expand: ±40 % horizontally, 20 % above, 80 % below (covers neck)
        pad_x   = int(fw * 0.40)
        pad_top = int(fh * 0.20)
        pad_bot = int(fh * 0.80)
        x1 = max(0, fx - pad_x)
        y1 = max(0, fy - pad_top)
        x2 = min(w, fx + fw + pad_x)
        y2 = min(h, fy + fh + pad_bot)
        mask[y1:y2, x1:x2] = 0

    # ROI: upper 70 % of the frame
    roi = mask[: int(h * 0.70), :]
    total = roi.shape[0] * roi.shape[1]
    skin_px = int(roi.sum()) // 255
    ratio = skin_px / max(total, 1)

    # ── Area-based penalty (tighter thresholds) ──────────────────
    # ≤2 %  → no visible arms → 0 penalty
    # ≥8 %  → clear arm/hand gesture → full penalty (1.0)
    if ratio <= 0.02:
        area_penalty = 0.0
    else:
        area_penalty = float(min(1.0, (ratio - 0.02) / 0.06))

    # ── Isolated-blob penalty ─────────────────────────────────────
    # Even a small isolated skin blob (e.g. a single fist) signals a
    # hand gesture. Find connected components in the ROI; penalise if any
    # blob is large enough to be a hand yet DISCONNECTED from the
    # face+neck region.
    blob_penalty = 0.0
    try:
        n_labels, labels, stats, _ = cv2.connectedComponentsWithStats(
            roi, connectivity=8,
        )
        min_blob_px = max(int(total * 0.003), 50)   # ≥0.3 % of ROI or ≥50px
        max_blob_px = int(total * 0.25)              # sanity cap (no whole-body blob)
        big_blobs = [
            i for i in range(1, n_labels)
            if min_blob_px <= stats[i, cv2.CC_STAT_AREA] <= max_blob_px
        ]
        if big_blobs:
            # Largest isolated blob as fraction of ROI
            biggest = max(stats[i, cv2.CC_STAT_AREA] for i in big_blobs)
            blob_ratio = biggest / total
            # 0.3 % → minor; 3 % → maximum isolated-blob contribution
            blob_penalty = float(min(0.70, blob_ratio / 0.03 * 0.70))
    except Exception:
        pass

    return float(min(1.0, max(area_penalty, blob_penalty)))


def _frontal_score(
    face_bbox: tuple[int, int, int, int] | None,
    img_shape: tuple,
) -> float:
    """
    Estimate how "portrait-appropriate" the face orientation is.
    Returns 0.0 (bad / turned) → 1.0 (ideal frontal, centred).

    Checks
    ------
    - Face detected at all by Haar (frontal cascade misses profiles).
    - Face width-to-height ratio: frontal ~0.75-0.95; profile < 0.55.
    - Horizontal centering: face centre should be within the middle 60 %.
    """
    if face_bbox is None:
        # Haar frontal cascade found nothing → likely turned / profile
        return 0.2

    x, y, w, h = face_bbox
    ih, iw = img_shape[:2]

    # Width-to-height ratio check
    wh_ratio = w / max(h, 1)
    ratio_score = float(np.clip((wh_ratio - 0.45) / (0.90 - 0.45), 0.0, 1.0))

    # Horizontal centering (face centre should be within 30 % of image centre)
    face_cx = (x + w / 2) / iw
    center_dev = abs(face_cx - 0.5)
    centering_score = float(np.clip(1.0 - center_dev / 0.30, 0.0, 1.0))

    return ratio_score * 0.6 + centering_score * 0.4


def score_photo_local(path: str) -> ScoredPhoto:
    """
    Score a portrait using InsightFace signals (primary) +
    fast OpenCV heuristics (always computed, used as fallback).

    Primary signals (InsightFace buffalo_l)
    ----------------------------------------
    - det_score     face detection confidence           weight 0.10
    - pose_score    head orientation (pitch/yaw/roll)   weight 0.30
    - eye_score     eye openness via EAR on landmarks   weight 0.30
    - sharp_score   Laplacian variance                  weight 0.15
    - bright_score  exposure quality                    weight 0.15
    Multipliers applied on top:
    - gesture_penalty  (skin-blob arm/hand detection)  max −0.80×
    - eye_close_penalty already baked into eye_score

    Fallback (when InsightFace is unavailable)
    ------------------------------------------
    Pure Haar + skin-blob scoring (previous behaviour).
    """
    sp = ScoredPhoto(path=path)
    img = imread_safe(path)
    if img is None:
        sp.is_ng = True
        sp.ng_reason = "unreadable"
        return sp

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # NG: placard still visible
    if CardDetector.has_placard(img):
        sp.is_ng = True
        sp.ng_reason = "placard visible"
        return sp

    sharp  = _sharpness(gray)
    bright = _brightness(gray)

    # Always compute Haar bbox for gesture detection (cheap)
    face_bbox_haar = _detect_face_bbox_haar(img)
    gesture_penalty = _gesture_penalty(img, face_bbox_haar)

    # ── Try InsightFace ──────────────────────────────────────────
    if_data = _if_scorer.score_face(img)

    if if_data is not None:
        pose             = if_data["pose"]
        lmk              = if_data["lmk"]
        det_score        = if_data["det_score"]
        x1, y1, x2, y2  = if_data["bbox"]
        face_r = ((x2 - x1) * (y2 - y1)) / max(img.shape[0] * img.shape[1], 1)

        # Eye openness score (0 = closed, 1 = open)
        eye_close_pen = _eye_close_penalty(lmk)
        eye_score     = (1.0 - eye_close_pen) * 10

        # Head pose score
        pose_pen, pose_flags = _pose_penalty(pose)
        pose_score = (1.0 - pose_pen) * 10

        # Base quality
        sharp_score  = min(sharp / 200, 1.0) * 10
        bright_score = (1.0 - abs(bright - 140) / 140) * 10
        det_s        = min(det_score, 1.0) * 10

        raw = (
            det_s        * 0.10 +
            pose_score   * 0.30 +
            eye_score    * 0.30 +
            sharp_score  * 0.15 +
            bright_score * 0.15
        )

        gesture_mult = 1.0 - gesture_penalty * 0.80
        sp.score = round(raw * gesture_mult, 2)

        # NG rules: eyes clearly closed (EAR penalty ≥ 0.80 on the new 0.30/0.16 scale)
        if eye_close_pen >= 0.80:
            sp.is_ng    = True
            sp.ng_reason = "eyes closed"

        flags = list(pose_flags)
        if eye_close_pen > 0.40:
            flags.append(f"eyes-closed(EAR:{eye_close_pen:.2f})")
        if gesture_penalty > 0.30:
            flags.append(f"gesture({gesture_penalty:.2f})")
        if sharp < 30:
            flags.append(f"low-sharp({sharp:.0f})")
        flag_str = (" " + " ".join(flags)) if flags else ""
        sp.comment = (
            f"[IF] sharp={sharp:.0f} bright={bright:.0f} face={face_r:.3f} "
            f"pose={[round(p,1) for p in pose]} "
            f"EAR_pen={eye_close_pen:.2f} gesture_pen={gesture_penalty:.2f}{flag_str}"
        )
        return sp

    # ── Fallback: pure Haar heuristics ──────────────────────────
    face_r = 0.0
    if face_bbox_haar is not None:
        fx, fy, fw, fh = face_bbox_haar
        face_r = (fw * fh) / (img.shape[0] * img.shape[1])

    frontal = _frontal_score(face_bbox_haar, img.shape)

    sharp_score  = min(sharp / 200, 1.0) * 10
    bright_score = (1.0 - abs(bright - 140) / 140) * 10
    face_score   = min(face_r / 0.08, 1.0) * 10
    frontal_s    = frontal * 10

    raw = (
        sharp_score  * 0.25 +
        bright_score * 0.20 +
        face_score   * 0.20 +
        frontal_s    * 0.35
    )
    gesture_mult = 1.0 - gesture_penalty * 0.80
    sp.score = round(raw * gesture_mult, 2)

    flags = []
    if sharp < 30:
        flags.append(f"low-sharp({sharp:.0f})")
    if bright < 50:
        flags.append(f"dark({bright:.0f})")
    if gesture_penalty > 0.30:
        flags.append(f"gesture({gesture_penalty:.2f})")
    if frontal < 0.50:
        flags.append(f"not-frontal({frontal:.2f})")
    flag_str = (" " + " ".join(flags)) if flags else ""
    sp.comment = (
        f"[Haar] sharp={sharp:.0f} bright={bright:.0f} face={face_r:.3f} "
        f"frontal={frontal:.2f} gesture_pen={gesture_penalty:.2f}{flag_str}"
    )
    return sp

# ──────────────────────────────────────────────────────────────────
# OpenAI scoring (optional, mirrors _embed.ipynb logic)
# ──────────────────────────────────────────────────────────────────

def score_photo_openai(path: str, client: Any = None, max_retries: int = 3) -> ScoredPhoto:
    """
    Score a portrait using GPT-4o vision.
    Requires `openai` package and a valid client.
    Falls back to local scoring if API fails.
    """
    import base64
    import json
    import time

    sp = ScoredPhoto(path=path)
    img = imread_safe(path)
    if img is None:
        sp.is_ng = True
        sp.ng_reason = "unreadable"
        return sp

    if CardDetector.has_placard(img):
        sp.is_ng = True
        sp.ng_reason = "placard visible"
        return sp

    if client is None:
        return score_photo_local(path)

    PROMPT = """You are a professional photo editor selecting the BEST FORMAL PORTRAIT for a graduation album.
The goal is to select a standard, professional-looking portrait.

Evaluate the photo using the criteria below.

CRITERIA:
1) Formality & Suitability (CRITICAL)
   - Standard posture (straight, facing forward).
   - Natural, professional expression (natural smile or neutral is best).
   - NO hand gestures (e.g., peace signs, hearts, thumbs up).
   - Standard framing (person centered, head not cut off).

2) Beauty & Quality
   - Lighting (bright, clear, not dark).
   - Sharpness (in focus, not blurry).
   - Background (clean blue background).

3) NG (MUST eliminate if ANY apply - is_ng should be true)
   - Hand gestures visible (e.g., finger hearts, peace signs).
   - Subject failing, stumbling, or eyes closed.
   - Mouth full or eating.
   - Strange or obscene pouting/poses.
   - Severe framing (e.g., top of head significantly cut off).
   - Person holding a placard/number card (this should be marked NG).

OUTPUT FORMAT (STRICT JSON ONLY):
{
  "suitability_score": 0-10,
  "beauty_score": 0-10,
  "expression_score": 0-10,
  "is_ng": true | false,
  "ng_reason": "string or null",
  "short_comment": "one-sentence professional comment explaining why this is or isn't a good formal portrait"
}

Rules:
- If there are hand gestures or "peace/heart" poses, set is_ng=true.
- Prioritize a "Formal Portrait" over a "Fun Pose".
- Output JSON only. No explanations."""

    # Encode
    h, w = img.shape[:2]
    scale = min(1.0, 1024 / max(h, w))
    if scale < 1.0:
        img = cv2.resize(img, (int(w * scale), int(h * scale)))
    _, buf = cv2.imencode(".jpg", img, [cv2.IMWRITE_JPEG_QUALITY, 85])
    b64 = base64.b64encode(buf.tobytes()).decode()

    for attempt in range(1, max_retries + 1):
        try:
            resp = client.chat.completions.create(
                model="gpt-4o",
                max_tokens=300,
                messages=[{
                    "role": "user",
                    "content": [
                        {"type": "text", "text": PROMPT},
                        {"type": "image_url",
                         "image_url": {"url": f"data:image/jpeg;base64,{b64}", "detail": "low"}},
                    ],
                }],
            )
            raw = resp.choices[0].message.content
            if not raw or not raw.strip():
                time.sleep(2 * attempt)
                continue
            raw = raw.strip()
            raw = re.sub(r"^```[a-z]*\n?", "", raw)
            raw = re.sub(r"\n?```$", "", raw)
            data = json.loads(raw)

            if data.get("is_ng"):
                sp.is_ng = True
                sp.ng_reason = data.get("ng_reason", "")
                return sp

            suit_s = data.get("suitability_score", 0)
            beauty_s = data.get("beauty_score", 0)
            expr_s = data.get("expression_score", 0)

            sp.score = round(
                suit_s * 0.50 +
                beauty_s * 0.30 +
                expr_s * 0.20,
                2,
            )
            sp.comment = data.get("short_comment", "")
            return sp

        except Exception as e:
            logger.warning("GPT-4o attempt %d failed: %s", attempt, e)
            time.sleep(2 * attempt)

    # Fall back to local
    return score_photo_local(path)

"""### Module 3 — `roster_parser`
Parses school rosters from Excel Format A (wide HR-block), Format B (paired-class sheets), or scanned PDF (GPT-5 vision).
"""

"""
roster_parser.py – Universal roster → roster.json converter
============================================================
Supports Excel (.xlsx) and scanned PDF rosters.

Excel formats:
  Format A ("千早高-style")
    • Sheet named "名票" (or "アルバム委員")
    • HR blocks repeated every 11 columns
    • Row 2 contains HR labels (HR31, HR32 …)
    • Data rows start at row 6

  Format B ("船橋啓明高-style")
    • Paired-class sheets ("3AB", "3CD", …)
    • Row 1: "３Ａ担任：Teacher Name"  |  "３Ｂ担任：…"
    • Row 2: Headers (No, 氏名, ふりがな, 性)
    • Data rows start at row 3

PDF format:
  Uses GPT-4o vision to extract structured data from scanned roster images.
  Requires an OpenAI API key (passed via --openai-key or OPENAI_API_KEY env var).

Output schema (roster.json):
{
  "school": "...",
  "year": 3,
  "classes": {
     "A": {
        "teacher": "片江 佳奈子",
        "student_count": 39,
        "students": [
           {"number": 1, "name": "青木 伸篤", "furigana": "あおき のぶあつ", "gender": "男"},
           ...
        ]
     },
     ...
  }
}
"""


import base64
import io
import json
import logging
import os
import re
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────

_FULL_TO_HALF = str.maketrans(
    "０１２３４５６７８９ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ"
    "ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ",
    "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    "abcdefghijklmnopqrstuvwxyz",
)


def _normalise(text: str) -> str:
    """Full-width → half-width, strip whitespace."""
    if text is None:
        return ""
    return str(text).translate(_FULL_TO_HALF).strip()


def _extract_class_id(raw: str) -> str | None:
    """
    Extract the class identifier, preserving the original format.
    '３Ａ担任：…' → 'A'  (letter-based school)
    'HR31'      → '1'   (number-based school, class 1)
    '3AB'       → parse later.
    """
    raw = _normalise(raw)
    # "3A担任" pattern — already a letter
    m = re.search(r"(\d)\s*([A-Za-z])\s*担任", raw)
    if m:
        return m.group(2).upper()
    # "HR3X" pattern  (千早高 uses numeric suffixes: HR31→class1, HR32→class2 …)
    m = re.search(r"HR\d(\d)", raw)
    if m:
        return m.group(1)  # preserve original number as string
    return None


def _extract_teacher(raw: str) -> str | None:
    """'３Ａ担任：片江　佳奈子' → '片江 佳奈子'"""
    raw = _normalise(raw)
    m = re.search(r"担任[：:]\s*(.+)", raw)
    if m:
        name = m.group(1).strip()
        # collapse multiple spaces / full-width spaces
        name = re.sub(r"[\s　]+", " ", name)
        return name
    return None


# ──────────────────────────────────────────────────────────────────
# Format B — Paired-class sheets  (船橋啓明高-style)
# ──────────────────────────────────────────────────────────────────

_PAIRED_SHEET_RE = re.compile(r"^(\d)([A-Za-z]{2})$")


def _detect_format_b(wb: openpyxl.Workbook) -> bool:
    """Return True if any sheet name matches '3AB', '3CD' etc."""
    for name in wb.sheetnames:
        if _PAIRED_SHEET_RE.match(_normalise(name)):
            return True
    return False


def _parse_format_b(wb: openpyxl.Workbook) -> dict:
    classes: dict[str, dict] = {}

    for sheet_name in wb.sheetnames:
        norm = _normalise(sheet_name)
        m = _PAIRED_SHEET_RE.match(norm)
        if not m:
            continue
        ws = wb[sheet_name]
        _year = int(m.group(1))
        pair = m.group(2).upper()  # "AB", "CD" …

        # ── Left class (cols A-D) ──
        left_header = str(ws.cell(1, 1).value or "")
        left_letter = _extract_class_id(left_header)
        left_teacher = _extract_teacher(left_header)
        left_count_raw = ws.cell(1, 5).value  # col E
        left_count = int(left_count_raw) if left_count_raw else None

        # ── Right class (cols G-J) ──
        right_header = str(ws.cell(1, 7).value or "")
        right_letter = _extract_class_id(right_header)
        right_teacher = _extract_teacher(right_header)
        right_count_raw = ws.cell(1, 11).value  # col K
        right_count = int(right_count_raw) if right_count_raw else None

        # Fallback letters from sheet name
        if not left_letter:
            left_letter = pair[0]
        if not right_letter:
            right_letter = pair[1]

        # ── Read students ──
        for letter, num_col, name_col, furi_col, gender_col, teacher, count in [
            (left_letter, 1, 2, 3, 4, left_teacher, left_count),
            (right_letter, 7, 8, 9, 10, right_teacher, right_count),
        ]:
            students = []
            seen_numbers: set[int] = set()
            consecutive_empty = 0
            for row in range(3, ws.max_row + 1):
                num_val = ws.cell(row, num_col).value
                name_val = ws.cell(row, name_col).value
                if num_val is None or name_val is None:
                    consecutive_empty += 1
                    if consecutive_empty >= 3:
                        break  # stop after 3 consecutive empty rows (end of roster)
                    continue
                consecutive_empty = 0
                try:
                    num = int(num_val)
                except (ValueError, TypeError):
                    continue
                if num < 1:
                    continue
                if num in seen_numbers:
                    continue  # skip duplicate attendance numbers
                seen_numbers.add(num)
                students.append({
                    "number": num,
                    "name": re.sub(r"[\s　]+", " ", str(name_val).strip()),
                    "furigana": re.sub(r"[\s　]+", " ", str(ws.cell(row, furi_col).value or "").strip()),
                    "gender": str(ws.cell(row, gender_col).value or "").strip(),
                })
            students.sort(key=lambda s: s["number"])
            classes[letter] = {
                "teacher": teacher,
                "student_count": count or len(students),
                "students": students,
            }

    return {"year": _year if classes else 3, "classes": classes}


# ──────────────────────────────────────────────────────────────────
# Format A — Wide HR-block layout  (千早高-style)
# ──────────────────────────────────────────────────────────────────

def _detect_format_a(wb: openpyxl.Workbook) -> bool:
    return "名票" in wb.sheetnames or "アルバム委員" in wb.sheetnames


def _parse_format_a(wb: openpyxl.Workbook) -> dict:
    ws = wb["名票"] if "名票" in wb.sheetnames else wb["アルバム委員"]

    # Read the column stride from "定義" if available
    stride = 11  # default
    if "定義" in wb.sheetnames:
        dws = wb["定義"]
        for row in dws.iter_rows(min_row=1, max_row=20, values_only=False):
            for cell in row:
                if cell.value and "１クラスあたり列数" in str(cell.value):
                    # value is in the cell 2 cols to the right
                    val = dws.cell(cell.row, cell.column + 2).value
                    if val:
                        stride = int(val)

    # Scan row 2 for HR labels → determine block start columns
    blocks: list[tuple[str, int]] = []  # (class_letter, start_col_0based)
    for col in range(1, ws.max_column + 1):
        val = ws.cell(2, col).value
        if val and re.search(r"HR\d+", str(val)):
            letter = _extract_class_id(str(val))
            if letter:
                # The number column is one col before the HR label col
                # In the data: E=number, F=name  where F is the HR-label column
                blocks.append((letter, col))

    # Also try to find teacher from デザイン sheet
    teachers: dict[str, str] = {}
    if "デザイン" in wb.sheetnames:
        dws = wb["デザイン"]
        for row in dws.iter_rows(min_row=1, max_row=dws.max_row, values_only=False):
            for cell in row:
                if cell.value and "担任" in str(cell.value):
                    # Teacher name might be nearby — check next row same col, or adjacent
                    pass  #千早高 only has the label "担任", not the actual name here

    classes: dict[str, dict] = {}

    for letter, name_col in blocks:
        # name_col is where the HR label sits = also the Name column in data rows
        num_col = name_col - 1         # attendance number
        gender_col = name_col + 1      # gender
        furi_col = name_col + 2        # furigana

        students = []
        for row in range(3, ws.max_row + 1):
            num_val = ws.cell(row, num_col).value
            name_val = ws.cell(row, name_col).value
            if num_val is None or name_val is None:
                continue
            try:
                num = int(num_val)
            except (ValueError, TypeError):
                continue
            if num < 1:
                continue
            students.append({
                "number": num,
                "name": re.sub(r"[\s　]+", " ", str(name_val).strip()),
                "furigana": re.sub(r"[\s　]+", " ", str(ws.cell(row, furi_col).value or "").strip()),
                "gender": str(ws.cell(row, gender_col).value or "").strip(),
            })

        students.sort(key=lambda s: s["number"])
        classes[letter] = {
            "teacher": teachers.get(letter),
            "student_count": len(students),
            "students": students,
        }

    return {"year": 3, "classes": classes}


# ──────────────────────────────────────────────────────────────────
# Generic fallback — scan all sheets for (number, name) columns
# ──────────────────────────────────────────────────────────────────

def _parse_generic(wb: openpyxl.Workbook) -> dict:
    """
    Best-effort: look for sheets that contain columns with headers
    matching No/番号 and 氏名/名前.
    """
    classes: dict[str, dict] = {}
    class_idx = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 3:
            continue

        # Try to find header row (first 5 rows)
        num_col = name_col = furi_col = gender_col = None
        header_row = None
        for r in range(1, min(6, ws.max_row + 1)):
            for c in range(1, min(30, ws.max_column + 1)):
                val = _normalise(str(ws.cell(r, c).value or ""))
                if val in ("no", "no.", "番号"):
                    num_col = c
                    header_row = r
                elif "氏" in val and "名" in val:
                    name_col = c
                elif "ふりがな" in val or "フリガナ" in val.upper():
                    furi_col = c
                elif val in ("性", "性別"):
                    gender_col = c

        if num_col is None or name_col is None or header_row is None:
            continue

        # Check row 1 for teacher
        teacher = None
        for c in range(1, min(15, ws.max_column + 1)):
            t = _extract_teacher(str(ws.cell(1, c).value or ""))
            if t:
                teacher = t
                break

        letter_from_header = None
        for c in range(1, min(15, ws.max_column + 1)):
            l = _extract_class_id(str(ws.cell(1, c).value or ""))
            if l:
                letter_from_header = l
                break

        letter = letter_from_header or str(class_idx + 1)
        class_idx += 1

        students = []
        for row in range(header_row + 1, ws.max_row + 1):
            num_val = ws.cell(row, num_col).value
            name_val = ws.cell(row, name_col).value
            if num_val is None or name_val is None:
                continue
            try:
                num = int(num_val)
            except (ValueError, TypeError):
                continue
            if num < 1:
                continue
            students.append({
                "number": num,
                "name": re.sub(r"[\s　]+", " ", str(name_val).strip()),
                "furigana": re.sub(r"[\s　]+", " ", str(ws.cell(row, furi_col).value or "").strip()) if furi_col else "",
                "gender": str(ws.cell(row, gender_col).value or "").strip() if gender_col else "",
            })

        if students:
            students.sort(key=lambda s: s["number"])
            classes[letter] = {
                "teacher": teacher,
                "student_count": len(students),
                "students": students,
            }

    return {"year": 3, "classes": classes}


# ──────────────────────────────────────────────────────────────────
# PDF parsing via GPT-4o vision
# ──────────────────────────────────────────────────────────────────

_PDF_ROSTER_PROMPT = """
You are given a scanned image of a Japanese school student roster (名票 / 学年名票).

Extract ALL classes and ALL students visible in this image.

Each class section typically shows:
- A header like "第3学年1組" or "3年A組" (grade + class number/letter)
- A homeroom teacher line like "担任 山田 太郎"
- A numbered list of students with: attendance number, full name (kanji), furigana (hiragana reading)
- Optionally a gender column (男/女)

Rules:
- Extract EVERY student without exception. Never skip, merge, or omit any row.
- Attendance numbers are sequential integers (1, 2, 3, ...). If a number cell is empty, infer it from position.
- Keep original kanji names EXACTLY as printed, including correct spacing between surname and given name.
- Furigana MUST match the kanji name: provide both surname reading AND given name reading separated by a space.
  Example: if name is "山田 太郎", furigana must be "やまだ たろう" (not just "たろう").
- For class identification: preserve the ORIGINAL identifier from the document.
  If the class uses a NUMBER (e.g. "第3学年1組" or "3年1組"), use that number as
  the JSON key: "1", "2", "3", etc.
  If the class uses a LETTER (e.g. "3年A組"), use that letter: "A", "B", etc.
- Gender: ONLY populate this field if the document has an explicit gender column (男/女 printed next to each student).
  If there is no gender column, set gender to "" for every student. Do NOT infer gender from names.
- If furigana is not visible or unclear, use "" (empty string) — do NOT guess or fabricate.
- Be careful with rare or unusual kanji. Transcribe exactly what is printed.

Return STRICT JSON only. No markdown, no explanation, no ``` blocks.

{
  "classes": {
    "1": {
      "teacher": "teacher full name or null",
      "students": [
        {"number": 1, "name": "山田 太郎", "furigana": "やまだ たろう", "gender": "男"},
        {"number": 2, "name": "鈴木 花子", "furigana": "すずき はなこ", "gender": "女"},
        ...
      ]
    },
    "2": { ... },
    ...
  }
}
""".strip()


def _render_pdf_pages(pdf_path: Path, dpi: int = 200) -> list[bytes]:
    """Render each PDF page as a PNG byte buffer."""
    import fitz  # PyMuPDF

    doc = fitz.open(str(pdf_path))
    pages: list[bytes] = []
    zoom = dpi / 72.0
    mat = fitz.Matrix(zoom, zoom)
    for page in doc:
        pix = page.get_pixmap(matrix=mat)
        pages.append(pix.tobytes("png"))
    doc.close()
    return pages


def _image_to_data_url(png_bytes: bytes) -> str:
    """Convert PNG bytes to a data URL for the OpenAI API."""
    b64 = base64.b64encode(png_bytes).decode("utf-8")
    return f"data:image/png;base64,{b64}"


def _repair_truncated_json(raw: str) -> dict | None:
    """
    Attempt to repair a truncated JSON string from GPT by closing
    open brackets/braces. Works for the common case where the response
    was cut mid-array or mid-object.
    """
    # Strip trailing incomplete key-value pairs or strings
    # Remove trailing comma or partial entry
    cleaned = raw.rstrip()
    # Remove trailing incomplete string (unclosed quote)
    if cleaned.count('"') % 2 != 0:
        # Find last complete quote pair
        last_quote = cleaned.rfind('"')
        if last_quote > 0:
            cleaned = cleaned[:last_quote]
            # Now remove the incomplete entry back to the last comma or bracket
            last_good = max(cleaned.rfind(','), cleaned.rfind('['), cleaned.rfind('{'))
            if last_good > 0:
                cleaned = cleaned[:last_good + 1]

    # Remove trailing comma
    cleaned = cleaned.rstrip().rstrip(',')

    # Count open brackets/braces and close them
    open_braces = cleaned.count('{') - cleaned.count('}')
    open_brackets = cleaned.count('[') - cleaned.count(']')

    # Close arrays first, then objects
    cleaned += ']' * max(0, open_brackets)
    cleaned += '}' * max(0, open_braces)

    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass

    # More aggressive: find the outermost { and try to parse just that
    brace_start = raw.find('{')
    if brace_start >= 0:
        attempt = raw[brace_start:]
        attempt = attempt.rstrip().rstrip(',')
        if attempt.count('"') % 2 != 0:
            last_q = attempt.rfind('"')
            attempt = attempt[:last_q]
            last_good = max(attempt.rfind(','), attempt.rfind('['), attempt.rfind('{'))
            if last_good > 0:
                attempt = attempt[:last_good + 1]
        attempt = attempt.rstrip().rstrip(',')
        ob = attempt.count('{') - attempt.count('}')
        obrk = attempt.count('[') - attempt.count(']')
        attempt += ']' * max(0, obrk)
        attempt += '}' * max(0, ob)
        try:
            return json.loads(attempt)
        except json.JSONDecodeError:
            pass

    return None


def _parse_pdf_with_vision(
    pdf_path: Path,
    openai_api_key: str | None = None,
    model: str = "gpt-4o",
) -> dict:
    """
    Parse a scanned PDF roster using GPT-5 vision.

    Renders each page, sends to GPT-5, merges results.
    """
    from openai import OpenAI

    # Suppress noisy HTTP debug logs from the OpenAI / httpx client
    for _noisy in ("openai", "httpx", "httpcore"):
        logging.getLogger(_noisy).setLevel(logging.WARNING)

    api_key = openai_api_key or os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise ValueError(
            "OpenAI API key required for PDF roster parsing. "
            "Pass --openai-key or set OPENAI_API_KEY environment variable."
        )

    client = OpenAI(api_key=api_key)
    pages = _render_pdf_pages(pdf_path, dpi=300)
    logger.info("Rendered %d PDF pages from %s", len(pages), pdf_path.name)

    all_classes: dict[str, dict] = {}

    for page_num, png_bytes in enumerate(pages, start=1):
        logger.info("  Sending page %d/%d to %s ...", page_num, len(pages), model)

        content = [
            {"type": "text", "text": _PDF_ROSTER_PROMPT + f"\n\nThis is page {page_num} of {len(pages)}."},
            {"type": "image_url", "image_url": {
                "url": _image_to_data_url(png_bytes),
                "detail": "high",
            }},
        ]

        resp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": "You extract structured data from school documents. Return only valid JSON."},
                {"role": "user", "content": content},
            ],
            #temperature=0.0,
            max_tokens=8192,
            response_format={"type": "json_object"}

        )

        raw = resp.choices[0].message.content or ""
        finish_reason = resp.choices[0].finish_reason
        # Strip markdown fences if present
        raw = re.sub(r"^```(?:json)?\s*", "", raw.strip())
        raw = re.sub(r"\s*```$", "", raw)

        if finish_reason == "length":
            logger.warning("  Page %d: response was truncated (max_tokens). "
                           "Attempting to repair JSON...", page_num)

        try:
            page_data = json.loads(raw)
        except json.JSONDecodeError:
            # Try to repair truncated JSON by closing open structures
            repaired = _repair_truncated_json(raw)
            if repaired is not None:
                page_data = repaired
                logger.info("  Page %d: successfully repaired truncated JSON", page_num)
            else:
                logger.warning("  Page %d: failed to parse GPT response as JSON", page_num)
                logger.debug("  Raw response (first 500 chars): %s", raw[:500])
                continue

        page_classes = page_data.get("classes", {})
        for cls_letter, cls_data in page_classes.items():
            cls_letter = cls_letter.upper()
            if cls_letter not in all_classes:
                all_classes[cls_letter] = {
                    "teacher": cls_data.get("teacher"),
                    "students": [],
                }
            # Merge students (avoid duplicates by attendance number)
            existing_nums = {s["number"] for s in all_classes[cls_letter]["students"]}
            for student in cls_data.get("students", []):
                num = student.get("number")
                if num is not None and num not in existing_nums:
                    all_classes[cls_letter]["students"].append({
                        "number": int(num),
                        "name": student.get("name", ""),
                        "furigana": student.get("furigana", ""),
                        "gender": student.get("gender", ""),
                    })
                    existing_nums.add(num)

            # Update teacher if not set
            if not all_classes[cls_letter]["teacher"] and cls_data.get("teacher"):
                all_classes[cls_letter]["teacher"] = cls_data["teacher"]

        logger.info("  Page %d: extracted %d classes", page_num,
                    len(page_classes))

    # Sort students by number within each class
    for cls_data in all_classes.values():
        cls_data["students"].sort(key=lambda s: s["number"])
        cls_data["student_count"] = len(cls_data["students"])

    return {
        "year": 3,  # Will be overridden or inferred
        "classes": all_classes,
    }


# ──────────────────────────────────────────────────────────────────
# Public API
# ──────────────────────────────────────────────────────────────────

def parse_roster(
    roster_path: str | Path,
    school_name: str | None = None,
    openai_api_key: str | None = None,
    vision_model: str = "gpt-4o",
) -> dict:
    """
    Parse a school roster file (.xlsx or .pdf) into a normalised dictionary.

    Parameters
    ----------
    roster_path : path to .xlsx or .pdf file
    school_name : optional override; auto-detected from filename if omitted
    openai_api_key : required for PDF parsing (GPT-4o vision)
    vision_model : OpenAI model for PDF vision extraction (default: gpt-4o)

    Returns
    -------
    dict  – see module docstring for schema
    """
    roster_path = Path(roster_path)
    if not roster_path.exists():
        raise FileNotFoundError(f"Roster file not found: {roster_path}")

    if school_name is None:
        # Strip prefix numbers and extension
        stem = roster_path.stem
        stem = re.sub(r"^\d+", "", stem).strip("_ ")
        school_name = stem or "Unknown"

    ext = roster_path.suffix.lower()

    if ext == ".pdf":
        # ── PDF: GPT-5 vision extraction ──
        logger.info("Detected PDF roster — using %s vision for extraction", vision_model)
        result = _parse_pdf_with_vision(
            roster_path,
            openai_api_key=openai_api_key,
            model=vision_model,
        )
        result["format_detected"] = f"PDF (via {vision_model} vision)"

    elif ext in (".xlsx", ".xls", ".xlsm"):
        # ── Excel parsing ──
        wb = openpyxl.load_workbook(str(roster_path), data_only=True, read_only=False)
        try:
            if _detect_format_b(wb):
                result = _parse_format_b(wb)
                result["format_detected"] = "B (paired-class sheets)"
            elif _detect_format_a(wb):
                result = _parse_format_a(wb)
                result["format_detected"] = "A (wide HR-block)"
            else:
                result = _parse_generic(wb)
                result["format_detected"] = "generic (auto-detect)"
        finally:
            wb.close()

    else:
        raise ValueError(f"Unsupported roster file format: {ext}. Use .xlsx or .pdf")

    result["school"] = school_name
    result["source_file"] = str(roster_path)

    # ── Validation ──
    warnings = []
    for cls_letter, cls_data in result.get("classes", {}).items():
        students = cls_data.get("students", [])
        numbers = [s["number"] for s in students]
        if numbers:
            expected = list(range(1, max(numbers) + 1))
            missing = sorted(set(expected) - set(numbers))
            if missing:
                warnings.append(f"Class {cls_letter}: missing attendance numbers {missing}")
            dupes = sorted(n for n in numbers if numbers.count(n) > 1)
            if dupes:
                warnings.append(f"Class {cls_letter}: duplicate numbers {list(set(dupes))}")
    result["warnings"] = warnings

    return result


def save_roster_json(roster: dict, output_path: str | Path) -> Path:
    """Write roster dict to a JSON file."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(roster, f, ensure_ascii=False, indent=2)
    return output_path


def get_valid_numbers(roster: dict, class_letter: str) -> set[int]:
    """Return the set of valid attendance numbers for a class."""
    cls = roster.get("classes", {}).get(class_letter, {})
    return {s["number"] for s in cls.get("students", [])}


def lookup_student(roster: dict, class_letter: str, number: int) -> dict | None:
    """Look up a student by class and attendance number."""
    cls = roster.get("classes", {}).get(class_letter, {})
    for s in cls.get("students", []):
        if s["number"] == number:
            return s
    return None

"""### Module 4 — `face_grouper`
Core engine: InsightFace buffalo_l model detects and clusters faces across all photos. Assigns student numbers via OCR, scores portraits, assigns groups to classes.
"""

"""
face_grouper.py – Face-recognition based photo grouping
=========================================================

Replaces the sequential approach (sequence_processor.py) with:
  1. Scan all images → detect class separators (pure card shots)
  2. Use InsightFace to detect + embed faces in all remaining images
  3. For EVERY image, also try card detection via contour + OCR
     (same logic as card_detector.py — NOT gated by has_placard)
  4. Cluster photos by face similarity (same person = same group)
  5. For each person-cluster, find the card image → assign number
  6. Score portraits → pick best shot

This solves:
  - Missed students (face grouping doesn't depend on card-reading order)
  - Same person not detected (face embeddings match across photos)
  - Class separator + first student # handled properly
"""


import logging
import os
from typing import Any

import cv2
import numpy as np


# Reuse data structures from sequence_processor for compatibility with package_exporter

logger = logging.getLogger(__name__)

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff"}

# Face similarity thresholds (from _embed.ipynb)
SAME_PERSON_THRESHOLD = 0.55   # cosine sim to match faces


# ──────────────────────────────────────────────────────────────────
# InsightFace engine (lazy singleton)
# ──────────────────────────────────────────────────────────────────

class _FaceEngine:
    """Lazy-loaded InsightFace wrapper."""

    def __init__(self):
        self._app = None

    def _ensure(self):
        if self._app is not None:
            return
        from insightface.app import FaceAnalysis
        logger.info("Loading InsightFace model (buffalo_l)...")
        self._app = FaceAnalysis(
            name="buffalo_l",
            providers=["CUDAExecutionProvider", "CPUExecutionProvider"],
        )
        self._app.prepare(ctx_id=0, det_size=(640, 640))
        logger.info("InsightFace ready.")

    def detect_faces(self, img: np.ndarray) -> list[dict]:
        """
        Detect faces and extract embeddings.
        Returns list of { 'bbox': [x1,y1,x2,y2], 'embedding': np.array, 'score': float }
        """
        self._ensure()
        faces = self._app.get(img)
        results = []
        for face in faces:
            emb = face.normed_embedding.astype(np.float32)
            norm = np.linalg.norm(emb)
            if norm > 0:
                emb = emb / norm
            results.append({
                "bbox": face.bbox.tolist(),
                "embedding": emb,
                "score": float(getattr(face, "det_score", 1.0)),
            })
        return results


_face_engine = _FaceEngine()


# ──────────────────────────────────────────────────────────────────
# Centroid-based face index (from _embed.ipynb)
# ──────────────────────────────────────────────────────────────────

class _FaceIndex:
    """Groups faces by cosine similarity using running centroids."""

    def __init__(self):
        self.group_ids: list[int] = []
        self.centroids: list[np.ndarray] = []
        self.counts: list[int] = []
        self._next_id = 0

    def find_or_create(self, emb: np.ndarray) -> int:
        """Find the closest matching group or create a new one."""
        if self.centroids:
            sims = np.dot(np.stack(self.centroids), emb)
            best_idx = int(np.argmax(sims))
            best_sim = float(sims[best_idx])

            if best_sim >= SAME_PERSON_THRESHOLD:
                # Update centroid with running average
                n = self.counts[best_idx]
                new_centroid = (self.centroids[best_idx] * n + emb) / (n + 1)
                norm = np.linalg.norm(new_centroid)
                if norm > 0:
                    new_centroid = new_centroid / norm
                self.centroids[best_idx] = new_centroid
                self.counts[best_idx] = n + 1
                return self.group_ids[best_idx]

        # New group
        gid = self._next_id
        self._next_id += 1
        self.group_ids.append(gid)
        self.centroids.append(emb.copy())
        self.counts.append(1)
        return gid



# ──────────────────────────────────────────────────────────────────
# Card reading
# ──────────────────────────────────────────────────────────────────

def _resize_for_ocr(img: np.ndarray, max_dim: int = 2400) -> np.ndarray:
    """Resize for OCR — needs larger images than face detection."""
    h, w = img.shape[:2]
    s = min(1.0, max_dim / max(h, w))
    if s < 1.0:
        img = cv2.resize(img, (int(w * s), int(h * s)), interpolation=cv2.INTER_AREA)
    return img

def _read_teacher_card_gpt(card_img: np.ndarray, client) -> TeacherInfo | None:
    """
    Use GPT-4o Vision to read a handwritten teacher card.

    Improvements:
    - PNG encoding (better handwriting preservation)
    - Character-by-character transcription
    - Return family_name, given_name, and spaced full name
    - Two GPT passes with majority vote
    - Robust JSON parsing
    """

    import base64
    import json as _json

    # -----------------------------
    # Encode image as PNG (lossless)
    # -----------------------------
    _, buf = cv2.imencode(".png", card_img)
    b64 = base64.b64encode(buf).decode("utf-8")

    prompt = (
        "This is a photo of a teacher holding a handwritten card.\n\n"
        "STEP 1 — Transcribe ALL visible text exactly as written.\n"
        "Do NOT translate or interpret.\n\n"
        "STEP 2 — Teacher name character separation:\n"
        "Identify each handwritten character individually.\n"
        "Return the characters in order as a list.\n\n"
        "Example:\n"
        '["植","村","優","里","香"]\n\n'
        "IMPORTANT:\n"
        "Japanese handwriting may place characters very close together.\n"
        "Adjacent characters may visually touch.\n"
        "Do NOT merge two characters into one.\n\n"
        "STEP 3 — Extract the following fields:\n"
        "1. family_name (last name kanji only)\n"
        "2. given_name (first name kanji only)\n"
        "3. name (family_name + space + given_name)\n"
        "4. class_label (example: 3-5)\n"
        "5. role (example: 担任 or 副担任)\n\n"
        "Respond ONLY with valid JSON using this format:\n"
        '{'
        '"characters": [],'
        '"family_name": "",'
        '"given_name": "",'
        '"name": "",'
        '"class_label": "",'
        '"role": "",'
        '"full_text": ""'
        '}\n\n'
        'If any field cannot be read, return an empty string "".'
    )

    def _call_gpt():
        resp = client.chat.completions.create(
            model="gpt-4o",
            temperature=0,
            max_tokens=250,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/png;base64,{b64}",
                                "detail": "high",
                            },
                        },
                    ],
                }
            ],
        )

        text = resp.choices[0].message.content.strip()
        if text.startswith("```"):
            text = text.split("\n", 1)[-1].rsplit("```", 1)[0].strip()
        return _json.loads(text)

    def format_name(family_name: str, given_name: str, name_field: str) -> str:
        """Return full name with space if missing."""
        if name_field:
            return name_field
        return f"{family_name} {given_name}".strip()

    try:
        # -----------------------
        # Run GPT twice for reliability
        # -----------------------
        data1 = _call_gpt()
        data2 = _call_gpt()

        # Family / given names
        family1 = data1.get("family_name", "")
        family2 = data2.get("family_name", "")
        given1 = data1.get("given_name", "")
        given2 = data2.get("given_name", "")

        # Class and role
        class1 = data1.get("class_label", "")
        class2 = data2.get("class_label", "")
        role1 = data1.get("role", "")
        role2 = data2.get("role", "")

        # -----------------------
        # Majority vote / fallback logic
        # -----------------------
        family_name = family1 if family1 == family2 else max([family1, family2], key=len)
        given_name  = given1  if given1 == given2   else max([given1, given2], key=len)
        class_label = class1  if class1 else class2
        role        = role1   if role1  else role2

        # Full name with space
        name_field  = data1.get("name", "")
        full_name   = format_name(family_name, given_name, name_field)

        info = TeacherInfo(
            name=full_name,
            class_label=class_label,
            role=role,
            raw_text=_json.dumps({"run1": data1, "run2": data2}, ensure_ascii=False),
            confidence=0.97 if (family1==family2 and given1==given2) else 0.90,
        )

        return info if (info.name or info.class_label) else None

    except Exception as e:
        logger.warning("GPT teacher card reading failed: %s", e)
        return None

# def _read_teacher_card_gpt(card_img: np.ndarray, client) -> TeacherInfo | None:
#     """
#     Use GPT-5 Vision to read a handwritten teacher card.
#     The card typically has: name, 先生, 教科 X-Y, 担任/副担任.
#     Returns TeacherInfo or None.
#     """
#     import base64
#     import json as _json

#     # Encode image as JPEG base64
#     _, buf = cv2.imencode(".jpg", card_img, [cv2.IMWRITE_JPEG_QUALITY, 95])
#     b64 = base64.b64encode(buf).decode("utf-8")

#     prompt = (
#         "This is a Japanese school portrait photo. The person is holding a white card "
#         "with handwritten text. The card may be difficult to see if the person is wearing "
#         "white clothing — look carefully for the rectangular card in the image.\n\n"
#         "The card contains:\n"
#         "  1. The teacher's full name. Read it exactly as written.\n"
#         "  2. A class assignment such as '教科 3-5', '3-5 担任', or similar.\n"
#         "  3. Optionally the printed word '先生' near the name.\n"
#         "  4. Optionally a role: '担任' (homeroom) or '副担任' (assistant homeroom).\n\n"
#         "CRITICAL — handwritten character boundaries:\n"
#         "  Japanese pen/brush handwriting often places characters very close together.\n"
#         "  Adjacent characters can visually merge and look like a single complex character.\n"
#         "  A typical Japanese name structure is: [family name 2-3 kanji] + [given name 1-3 kanji].\n"
#         "  Trace each character individually before writing it — do NOT fuse two adjacent "
#         "characters into one. For example, 里香 (two characters: 里 + 香) must not be "
#         "read as 瑠 (one character).\n"
#         "  Also: ignore stray pen strokes, punctuation marks, or decorative flourishes "
#         "that appear at the end of the name — do not include them in the transcription.\n\n"
#         "IMPORTANT: Transcribe every character exactly as it appears — do not normalise, "
#         "translate, or substitute characters. Preserve the original script and spelling.\n\n"
#         "Respond with JSON only (no markdown, no explanation):\n"
#         '{"name": "<name as written>", "class_label": "<e.g. 3-5>", "role": "<role or empty>"}\n\n'
#         "If there is no teacher card visible in the photo, respond exactly with: "
#         '{"not_teacher": true}\n'
#         "If any field cannot be determined, use an empty string \"\"."
#     )

#     try:
#         resp = client.chat.completions.create(
#             model="gpt-5",
#             messages=[{
#                 "role": "user",
#                 "content": [
#                     {"type": "text", "text": prompt},
#                     {"type": "image_url", "image_url": {
#                         "url": f"data:image/jpeg;base64,{b64}",
#                         "detail": "high",
#                     }},
#                 ],
#             }],
#             response_format={"type": "json_object"},
#             max_completion_tokens=6000,
#         )

#         choice = resp.choices[0]
#         raw_content = choice.message.content

#         # Handle refusal (model declined to answer)
#         refusal = getattr(choice.message, "refusal", None)
#         if refusal:
#             logger.warning("GPT refused teacher card request: %s", refusal)
#             return None

#         # Handle empty / None content
#         if not raw_content or not raw_content.strip():
#             finish = getattr(choice, "finish_reason", "unknown")
#             logger.warning(
#                 "GPT returned empty content for teacher card (finish_reason=%s)", finish
#             )
#             return None

#         text = raw_content.strip()

#         # Strip markdown fences if present
#         if text.startswith("```"):
#             text = text.split("\n", 1)[-1].rsplit("```", 1)[0].strip()

#         if not text:
#             logger.warning("GPT response was empty after stripping fences")
#             return None

#         logger.debug("GPT raw teacher card response: %s", text)
#         data = _json.loads(text)

#         if data.get("not_teacher"):
#             return None

#         info = TeacherInfo(
#             name=data.get("name", ""),
#             class_label=data.get("class_label", ""),
#             role=data.get("role", ""),
#             raw_text=text,
#             confidence=0.95,
#         )
#         return info if (info.name or info.class_label) else None

#     except _json.JSONDecodeError as e:
#         logger.warning("GPT teacher card: JSON parse failed (%s). Raw: %r", e, locals().get("text", ""))
#         return None
#     except Exception as e:
#         logger.warning("GPT teacher card reading failed: %s", e)
#         return None


def _high_white_ratio(img: np.ndarray, threshold: float = 0.25) -> bool:
    """
    Returns True when a large portion of the image is near-white.
    Used to detect white-clothing + white-card scenarios where contour
    detection fails because there is no contrast at the card border.
    """
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    # Near-white: low saturation, high value
    mask = cv2.inRange(hsv, (0, 0, 180), (180, 60, 255))
    ratio = float(np.count_nonzero(mask)) / (img.shape[0] * img.shape[1])
    return ratio >= threshold


def _has_dense_text(card_crop: np.ndarray) -> bool:
    """
    Quick heuristic: teacher cards have many text contours (name, 先生,
    教科, class, 担任).  Student cards have a single large number.
    Returns True when the card crop contains dense text (likely teacher).
    """
    gray = cv2.cvtColor(card_crop, cv2.COLOR_BGR2GRAY) if len(card_crop.shape) == 3 else card_crop
    binary = cv2.adaptiveThreshold(
        gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV, 21, 5,
    )
    contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    # Filter noise – only contours big enough to be a character
    sig = [c for c in contours if cv2.contourArea(c) > 50]
    return len(sig) > 20          # student cards: 1-15 contours; teacher: 35+


def _try_teacher_detection(card_crop, openai_client):
    """Try teacher card detection: GPT first, EasyOCR Japanese fallback."""
    teacher = None
    if openai_client:
        teacher = _read_teacher_card_gpt(card_crop, openai_client)
        if teacher:
            logger.info("    GPT read teacher: %s [%s] %s",
                        teacher.name, teacher.class_label, teacher.role)
    if teacher is None:
        try:
            teacher = detect_teacher_card(card_crop)
        except Exception:
            pass
    return teacher


def _read_card_from_image(
    img_original: np.ndarray,
    valid_numbers: set[int] | None = None,
    openai_client=None,
) -> dict:
    """
    Read card info from image:
      1. If card region has dense text → try teacher detection FIRST
         (prevents false student-number matches on "教科 3-1")
      2. Try student number OCR (EasyOCR digits — fast, reliable)
      3. If no number found & card region detected → try GPT for teacher

    Returns dict with keys:
      is_card_image: bool
      card_number: int|None
      teacher_info: TeacherInfo|None
    """
    result = {
        "is_card_image": False,
        "card_number": None,
        "teacher_info": None,
    }

    # Use 2400px for OCR (small cards unreadable at 1600px)
    img_ocr = _resize_for_ocr(img_original, 2400)

    # Contour-based card crop
    card_crop, card_found = detect_card_region(img_ocr)

    # ── Pre-check: dense text on card → likely teacher card ──
    if card_found and _has_dense_text(card_crop):
        teacher = _try_teacher_detection(card_crop, openai_client)
        if teacher is not None:
            result["teacher_info"] = teacher
            result["is_card_image"] = True
            return result
        # GPT said "not_teacher" – fall through to student OCR

    # ── Student number OCR ──
    all_candidates = []
    for vname, vimg in _prepare_card(card_crop):
        for c in _ocr.read_numbers(vimg):
            c.source = f"contour_{vname}"
            all_candidates.append(c)

    # Direct OCR on full image (catches cards missed by contour)
    for c in _ocr.read_numbers(img_ocr):
        c.source = "direct_full"
        all_candidates.append(c)

    good = [c for c in all_candidates if c.confidence >= 0.40 and 1 <= c.value <= 50]
    if good:
        if valid_numbers:
            in_roster = [c for c in good if c.value in valid_numbers]
            if in_roster:
                good = in_roster
        winner = max(good, key=lambda c: c.confidence)
        result["card_number"] = winner.value
        result["is_card_image"] = True
        return result

    # ── Still nothing & card found → try teacher as last resort ──
    if card_found:
        teacher = _try_teacher_detection(card_crop, openai_client)
        if teacher is not None:
            result["teacher_info"] = teacher
            result["is_card_image"] = True


    # ── White-clothing fallback: no contour found but image is mostly white ──
    # Teacher wears white → white card merges into clothing → contour fails.
    # Only send full image to GPT when there is strong evidence of white content.
    if not result["is_card_image"] and not card_found and openai_client:
        if _high_white_ratio(img_ocr):
            logger.info("  ↪ White-clothing fallback: sending full image to GPT")
            teacher = _try_teacher_detection(img_ocr, openai_client)  # full image
            if teacher is not None:
                result["teacher_info"] = teacher
                result["is_card_image"] = True

    return result


# ──────────────────────────────────────────────────────────────────
# Main processor
# ──────────────────────────────────────────────────────────────────

class FaceGrouper:
    """
    Process a camera folder using face recognition to group photos.

    Steps:
      1. Scan images → find class separators (and their student numbers)
      2. Detect faces → cluster by similarity
      3. For EVERY non-separator image, try card detection via contour+OCR
      4. For each person-cluster, find card image → assign number
      5. Score portraits → pick best
      6. Return ClassPhotoGroup dict (same interface as SequenceProcessor)
    """

    def __init__(
        self,
        valid_numbers: set[int] | None = None,
        scoring: str = "local",
        openai_client: Any = None,
        roster: dict | None = None,
    ):
        self.valid_numbers = valid_numbers or set()
        self.scoring = scoring
        self.openai_client = openai_client
        self.roster = roster
        self._review_queue: list[dict] = []

    @property
    def review_queue(self) -> list[dict]:
        return self._review_queue

    def _list_images(self, folder: str) -> list[str]:
        paths = []
        for f in sorted(os.listdir(folder)):
            if os.path.splitext(f)[1].lower() in IMAGE_EXTENSIONS:
                paths.append(os.path.join(folder, f))
        return paths

    def _score(self, path: str) -> ScoredPhoto:
        if self.scoring == "openai":
            return score_photo_openai(path, self.openai_client)
        return score_photo_local(path)

    def process_folder(self, folder: str) -> dict[str, ClassPhotoGroup]:
        """
        Process all images in folder using face recognition.

        Returns dict of { class_id: ClassPhotoGroup }.
        """
        images = self._list_images(folder)
        if not images:
            logger.warning("No images found in %s", folder)
            return {}

        logger.info("=" * 55)
        logger.info("PHASE 1: Scanning %d images for class separators", len(images))
        logger.info("=" * 55)

        # ── Phase 1: Find class separators ──
        # A separator might also contain a student number (e.g. "3-1" + "#1").
        # If the separator image also has a face, we include it in face grouping too.
        separators: list[tuple[int, str, ClassSeparatorInfo]] = []
        non_separator_images: list[tuple[int, str]] = []

        for idx, img_path in enumerate(images):
            img = imread_safe(img_path)
            if img is None:
                continue
            img_r = _resize(img)

            sep = detect_class_separator(img_r)
            if sep is not None:
                separators.append((idx, img_path, sep))
                if sep.student_number:
                    logger.info("  🏫 CLASS %s + Student #%d  ← %s",
                                sep.label, sep.student_number,
                                os.path.basename(img_path))
                else:
                    logger.info("  🏫 CLASS %s  ← %s",
                                sep.label, os.path.basename(img_path))

                # Check if this separator image also has a face
                # (combined separator+student card photo)
                faces = _face_engine.detect_faces(img_r)
                if faces:
                    non_separator_images.append((idx, img_path))
                    logger.info("    (also has face — will be face-grouped)")
            else:
                non_separator_images.append((idx, img_path))

        logger.info("Found %d class separators, %d images for face grouping",
                     len(separators), len(non_separator_images))

        # ── Phase 2: Detect faces + Read cards for ALL non-separator images ──
        logger.info("")
        logger.info("=" * 55)
        logger.info("PHASE 2: Detecting faces + reading cards")
        logger.info("=" * 55)

        face_index = _FaceIndex()
        image_info: dict[str, dict] = {}

        for idx, img_path in non_separator_images:
            fname = os.path.basename(img_path)
            img = imread_safe(img_path)
            if img is None:
                continue

            img_resized = _resize(img)

            # Detect faces
            faces = _face_engine.detect_faces(img_resized)

            # Try card detection on EVERY image
            # Pass original (not face-resized) so OCR gets enough resolution
            card_result = _read_card_from_image(img, self.valid_numbers, self.openai_client)

            info = {
                "index": idx,
                "is_card_image": card_result["is_card_image"],
                "card_number": card_result["card_number"],
                "teacher_info": card_result["teacher_info"],
                "face_detected": len(faces) > 0,
                "group_id": None,
            }

            # Assign to face group (use largest face)
            if faces:
                faces_sorted = sorted(
                    faces,
                    key=lambda f: (f["bbox"][2] - f["bbox"][0]) * (f["bbox"][3] - f["bbox"][1]),
                    reverse=True,
                )
                main_face = faces_sorted[0]
                group_id = face_index.find_or_create(main_face["embedding"])
                info["group_id"] = group_id

                if card_result["card_number"]:
                    logger.info("  🔢 #%d (face group %d)  ← %s",
                                card_result["card_number"], group_id, fname)
                elif card_result["teacher_info"]:
                    t = card_result["teacher_info"]
                    logger.info("  📋 Teacher %s (face group %d)  ← %s",
                                t.name or "?", group_id, fname)
                else:
                    logger.debug("  📸 Portrait (face group %d)  ← %s",
                                 group_id, fname)
            else:
                if card_result["card_number"]:
                    logger.info("  🔢 #%d (no face)  ← %s",
                                card_result["card_number"], fname)
                elif card_result["teacher_info"]:
                    logger.info("  📋 Teacher %s (no face)  ← %s",
                                card_result["teacher_info"].name or "?", fname)
                else:
                    logger.debug("  ❓ No face, no card  ← %s", fname)

            image_info[img_path] = info

        # ── Phase 3: Build person groups from face clusters ──
        logger.info("")
        logger.info("=" * 55)
        logger.info("PHASE 3: Building person groups from face clusters")
        logger.info("=" * 55)

        groups_by_face: dict[int, list[str]] = {}
        ungrouped: list[str] = []

        for img_path, info in image_info.items():
            gid = info["group_id"]
            if gid is not None:
                groups_by_face.setdefault(gid, []).append(img_path)
            else:
                ungrouped.append(img_path)

        logger.info("Face groups: %d, Ungrouped images: %d",
                     len(groups_by_face), len(ungrouped))

        # For each face group, find the card number / teacher info
        person_groups: list[dict] = []

        for gid, group_images in sorted(groups_by_face.items()):
            group_images.sort()

            group_number = None
            group_teacher = None
            card_image_paths: list[str] = []

            for img_path in group_images:
                info = image_info[img_path]
                if info["card_number"] is not None:
                    if group_number is None:
                        group_number = info["card_number"]
                    card_image_paths.append(img_path)
                if info["teacher_info"] is not None:
                    if group_teacher is None:
                        group_teacher = info["teacher_info"]
                    card_image_paths.append(img_path)

            pg = {
                "group_id": gid,
                "images": group_images,
                "number": group_number,
                "teacher_info": group_teacher,
                "card_images": card_image_paths,
            }
            person_groups.append(pg)

            if group_teacher:
                logger.info("  Group %d: TEACHER %s [%s] (%d photos)",
                            gid, group_teacher.name or "?",
                            group_teacher.class_label or "?", len(group_images))
            elif group_number:
                logger.info("  Group %d: Student #%d (%d photos)",
                            gid, group_number, len(group_images))
            else:
                logger.info("  Group %d: Unknown (%d photos)", gid, len(group_images))

        # ── Attach ungrouped card-only images to nearest face group ──
        for img_path in ungrouped:
            info = image_info[img_path]
            if info["card_number"] is not None or info["teacher_info"] is not None:
                best_group = self._find_nearest_group(
                    img_path, person_groups, images
                )
                if best_group is not None:
                    if info["card_number"] and best_group["number"] is None:
                        best_group["number"] = info["card_number"]
                        best_group["card_images"].append(img_path)
                        logger.info("  Attached card #%d to group %d (by proximity)",
                                    info["card_number"], best_group["group_id"])
                    elif info["teacher_info"] and best_group["teacher_info"] is None:
                        best_group["teacher_info"] = info["teacher_info"]
                        best_group["card_images"].append(img_path)
                        logger.info("  Attached teacher card to group %d (by proximity)",
                                    best_group["group_id"])

        # ── Assign separator student numbers to first unnumbered group ──
        for sep_idx, sep_path, sep_info in separators:
            if sep_info.student_number is None:
                continue

            # Skip if already assigned to some group
            already_assigned = any(
                pg["number"] == sep_info.student_number for pg in person_groups
            )
            if already_assigned:
                continue

            # Find first unnumbered group right after this separator
            best_pg = None
            best_min_idx = float("inf")
            for pg in person_groups:
                if pg["number"] is not None or pg["teacher_info"] is not None:
                    continue
                min_idx = min(image_info[p]["index"] for p in pg["images"])
                if min_idx > sep_idx and min_idx < best_min_idx:
                    best_min_idx = min_idx
                    best_pg = pg

            if best_pg is not None:
                best_pg["number"] = sep_info.student_number
                logger.info("  Assigned separator student #%d to group %d",
                            sep_info.student_number, best_pg["group_id"])

        # ── Phase 4: Assign groups to classes ──
        logger.info("")
        logger.info("=" * 55)
        logger.info("PHASE 4: Assigning groups to classes")
        logger.info("=" * 55)

        # Build class ranges from separator positions
        class_ranges: list[tuple[int, int, ClassSeparatorInfo, str]] = []
        for i, (sep_idx, sep_path, sep_info) in enumerate(separators):
            if i + 1 < len(separators):
                end_idx = separators[i + 1][0]
            else:
                end_idx = len(images)
            class_ranges.append((sep_idx, end_idx, sep_info, sep_path))

        all_class_groups: dict[str, ClassPhotoGroup] = {}

        for pg in person_groups:
            # Determine which class by median image index
            class_label = "unknown"
            sep_info = None
            sep_path = None

            # Teachers: prefer the class_label from their card (GPT / OCR)
            if pg["teacher_info"] is not None and pg["teacher_info"].class_label:
                teacher_cls = pg["teacher_info"].class_label  # e.g. "3-3"
                # Find the matching separator to get sep_info
                for start_idx, end_idx, cr_sep, cr_path in class_ranges:
                    if cr_sep.label == teacher_cls:
                        class_label = cr_sep.label
                        sep_info = cr_sep
                        sep_path = cr_path
                        break
                if class_label == "unknown":
                    # No matching separator – use teacher card value directly
                    class_label = teacher_cls

            # Students / unknown teachers: determine by median image index
            if class_label == "unknown":
                indices = [image_info[p]["index"] for p in pg["images"]]
                if indices:
                    median_idx = sorted(indices)[len(indices) // 2]
                    for start_idx, end_idx, cr_sep, cr_path in class_ranges:
                        if start_idx <= median_idx < end_idx:
                            class_label = cr_sep.label
                            sep_info = cr_sep
                            sep_path = cr_path
                            break

                if class_label == "unknown" and separators and indices:
                    if max(indices) < separators[0][0]:
                        class_label = "pre-separator"

            # Get or create class group
            if class_label not in all_class_groups:
                cpg = ClassPhotoGroup(class_label=class_label)
                if sep_info:
                    cpg.grade = sep_info.grade
                    cpg.class_number = sep_info.class_number
                    cpg.separator_image = sep_path
                all_class_groups[class_label] = cpg
            cpg = all_class_groups[class_label]

            # Build StudentPhotoGroup or TeacherPhotoGroup
            # IMPORTANT: card images (is_card_image=True) are NEVER
            # included as portraits/backups — they are card-only (_札).
            if pg["teacher_info"] is not None:
                tpg = TeacherPhotoGroup(
                    teacher_info=pg["teacher_info"],
                    card_images=pg["card_images"],
                )
                for img_path in pg["images"]:
                    info = image_info[img_path]
                    if not info["is_card_image"]:
                        scored = self._score(img_path)
                        tpg.portraits.append(scored)
                self._pick_best_teacher(tpg)
                cpg.teacher = tpg
                logger.info("  Teacher %s → class %s",
                            pg["teacher_info"].name or "?", class_label)
            else:
                spg = StudentPhotoGroup(
                    attendance_number=pg["number"],
                    card_images=pg["card_images"],
                )
                for img_path in pg["images"]:
                    info = image_info[img_path]
                    # Only non-card images become portraits/backups
                    if not info["is_card_image"]:
                        scored = self._score(img_path)
                        spg.portraits.append(scored)
                self._pick_best(spg)
                cpg.students.append(spg)

                if pg["number"]:
                    logger.info("  Student #%d → class %s (%d portraits)",
                                pg["number"], class_label, len(spg.portraits))
                else:
                    logger.info("  Unknown student → class %s (%d portraits)",
                                class_label, len(spg.portraits))

        # ── Summary ──
        logger.info("")
        logger.info("=" * 55)
        logger.info("RESULTS SUMMARY")
        logger.info("=" * 55)
        for cid, cg in sorted(all_class_groups.items()):
            n_students = len(cg.students)
            n_id = sum(1 for s in cg.students if s.attendance_number is not None)
            n_review = n_students - n_id
            teacher_note = ""
            if cg.teacher and cg.teacher.teacher_info:
                teacher_note = f" + teacher ({cg.teacher.teacher_info.name})"
            logger.info("  Class %s: %d students (%d identified, %d unknown)%s",
                        cid, n_students, n_id, n_review, teacher_note)

        return all_class_groups

    def _find_nearest_group(
        self,
        card_path: str,
        person_groups: list[dict],
        all_images: list[str],
    ) -> dict | None:
        """Find the person group closest to this card image by filename order."""
        try:
            card_idx = all_images.index(card_path)
        except ValueError:
            return None

        best_group = None
        best_dist = float("inf")

        for pg in person_groups:
            if pg["number"] is not None and pg["teacher_info"] is not None:
                continue
            for img_path in pg["images"]:
                try:
                    img_idx = all_images.index(img_path)
                    dist = abs(img_idx - card_idx)
                    if dist < best_dist:
                        best_dist = dist
                        best_group = pg
                except ValueError:
                    continue

        if best_dist <= 5:
            return best_group
        return None

    def _pick_best(self, group: StudentPhotoGroup):
        valid = [p for p in group.portraits if not p.is_ng]
        if valid:
            group.best_shot = max(valid, key=lambda p: p.score)
        elif group.portraits:
            group.best_shot = max(group.portraits, key=lambda p: p.score)
            if group.best_shot:
                group.best_shot.comment += " [all-NG fallback]"

    def _pick_best_teacher(self, group: TeacherPhotoGroup):
        valid = [p for p in group.portraits if not p.is_ng]
        if valid:
            group.best_shot = max(valid, key=lambda p: p.score)
        elif group.portraits:
            group.best_shot = max(group.portraits, key=lambda p: p.score)
            if group.best_shot:
                group.best_shot.comment += " [all-NG fallback]"

    def save_review_queue(self, path):
        """Compatibility method."""
        import json
        from pathlib import Path
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(self._review_queue, f, ensure_ascii=False, indent=2)

    def process_folder_flat(self, folder: str) -> list[StudentPhotoGroup]:
        """Legacy flat mode."""
        class_groups = self.process_folder(folder)
        flat: list[StudentPhotoGroup] = []
        for cg in class_groups.values():
            flat.extend(cg.students)
        return flat

"""### Module 5 — `package_exporter`
Copies and renames the best shots into the standardised output folder structure expected by the InDesign script.
"""

"""
package_exporter.py – Export standardised "Ready-for-InDesign" folder
=====================================================================

Takes the output of SequenceProcessor (list of StudentPhotoGroup) plus the
roster data, and produces:

  output_folder/
    3-1/                                    ← one sub-folder per class
      26_上水高_IMG_1234_3101_札01.jpg        ← ID plate shot (single)
      26_上水高_IMG_1234_3101_札02.jpg        ← ID plate shot (2nd, if multiple)
      26_上水高_IMG_5678_3101_本01.jpg        ← Best shot
      26_上水高_IMG_5679_3101_本02.jpg        ← Backup shot  (next best)
      26_上水高_IMG_2345_3102_札01.jpg
      26_上水高_IMG_6789_3102_本01.jpg
      ...
    manifest.json                           ← machine-readable summary for JSX

Naming convention:
  [Year]_[SchoolName]_[OriginalFileName]_[Grade][Class][ID]_[Tag].ext
  Tags: 札01 = ID plate, 札02 = 2nd ID plate, 本01 = best shot, 本02 = 2nd best, ...
  Example: 26_上水高_IMG_2337_3105_札01.jpg  (grade 3, class 1, no. 5)
"""


import json
import logging
import os
import shutil
from pathlib import Path
from typing import Any


logger = logging.getLogger(__name__)

# Grade-class → roster letter mapping (configurable per school)
# e.g. {"3-1": "A", "3-2": "B", ...} or auto-generate
def _grade_class_to_roster_id(grade: int, class_num: int, mapping: dict | None = None) -> str:
    """Convert grade-class (e.g. 3, 7) to a roster class key.

    Uses explicit mapping if provided, otherwise returns the class number
    as a string (preserving the original identifier).
    """
    key = f"{grade}-{class_num}"
    if mapping and key in mapping:
        return mapping[key]
    # Default: use class_num as-is (matches roster keys like "1", "2", "7")
    return str(class_num)


def _sanitise_filename(name: str) -> str:
    """Remove characters illegal in Windows filenames."""
    return "".join(c for c in name if c not in r'<>:"/\|?*').strip()



def _build_filename(year: str, school_name: str, orig_path: str, id_str: str, tag: str) -> str:
    """Build standardised filename.

    Examples:
      '26_上水高_IMG_2337_3105_札01.jpg'
      '26_上水高_IMG_2338_3105_本01.jpg'

    Parameters
    ----------
    year : 2-digit year string (e.g. '26')
    school_name : school name provided by user
    orig_path : original source file path (stem is extracted)
    id_str : formatted ID segment (e.g. '3105' for grade 3, class 1, no. 5)
    tag : '札01', '本01', '本02', etc.
    """
    safe_school = _sanitise_filename(school_name)
    orig_stem = _sanitise_filename(Path(orig_path).stem)
    ext = os.path.splitext(orig_path)[1]
    return f"{year}_{safe_school}_{orig_stem}_{id_str}_{tag}{ext}"



def _make_student_id(grade_str: str, class_str: str, num: int | str) -> str:
    """Build the ID segment: grade + class + zero-padded attendance number.

    Example: grade_str='3', class_str='1', num=5  →  '3105'
    """
    num_part = str(num).zfill(2) if isinstance(num, int) else str(num)
    return f"{grade_str}{class_str}{num_part}"


def export_package(
    groups: list[StudentPhotoGroup],
    roster: dict | None,
    class_id: str,
    output_dir: str | Path,
    school_name: str = "",
    year: str = "",
    max_backups: int = 0,
    roster_class_letter: str | None = None,
    teacher_group: TeacherPhotoGroup | None = None,
) -> dict:
    """
    Export student photo groups into standardised folder structure.

    Parameters
    ----------
    groups : list of StudentPhotoGroup for ONE class
    roster : parsed roster dict (from roster_parser), or None if no roster
    class_id : class identifier for folder/file naming (e.g. '3-7' or 'A')
    output_dir : root output directory
    school_name : school name for file naming (e.g. '千早高')
    year : 2-digit year string for file naming (e.g. '26'). Defaults to ''.
    max_backups : max backup portraits to include. 0 = no limit (default).
    roster_class_letter : roster class letter for student lookup (e.g. 'G')
                          If None, uses class_id directly.

    Returns
    -------
    dict : manifest data (also saved as manifest.json)
    """
    lookup_letter = roster_class_letter or class_id
    # Parse grade and class number from class_id (e.g. "3-1" → grade_str="3", class_str="1")
    if '-' in class_id:
        _parts = class_id.split('-', 1)
        grade_str, class_str = _parts[0], _parts[1]
    else:
        grade_str, class_str = '', class_id
    output_dir = Path(output_dir)
    class_dir = output_dir / class_id
    class_dir.mkdir(parents=True, exist_ok=True)

    manifest_entries: list[dict] = []
    exported = 0
    skipped = 0

    # ── Export teacher photos first (if any) ──
    if teacher_group is not None and teacher_group.teacher_info is not None:
        ti = teacher_group.teacher_info
        teacher_files: dict[str, str] = {}

        teacher_id = f"{grade_str}{class_str}先生"
        for card_i, card_path in enumerate(teacher_group.card_images, start=1):
            if not os.path.isfile(card_path):
                continue
            tag = f"札{str(card_i).zfill(2)}"
            dst_name = _build_filename(year, school_name, card_path, teacher_id, tag)
            shutil.copy2(card_path, class_dir / dst_name)
            teacher_files[tag] = dst_name
            logger.debug("  📋 (teacher) %s", dst_name)

        if teacher_group.best_shot and os.path.isfile(teacher_group.best_shot.path):
            dst_name = _build_filename(year, school_name, teacher_group.best_shot.path, teacher_id, "本01")
            shutil.copy2(teacher_group.best_shot.path, class_dir / dst_name)
            teacher_files["本01"] = dst_name
            logger.debug("  ⭐ (teacher) %s", dst_name)

        # ── Teacher backup shots (本02, 本03, ...) — same as students ──
        best_path = teacher_group.best_shot.path if teacher_group.best_shot else ""
        teacher_backups = sorted(
            [p for p in teacher_group.portraits if p.path != best_path],
            key=lambda p: p.score,
            reverse=True,
        )
        t_limit = max_backups if max_backups > 0 else len(teacher_backups)
        for i, photo in enumerate(teacher_backups[:t_limit], start=2):
            if not os.path.isfile(photo.path):
                continue
            tag = f"本{str(i).zfill(2)}"
            dst_name = _build_filename(year, school_name, photo.path, teacher_id, tag)
            shutil.copy2(photo.path, class_dir / dst_name)
            teacher_files[tag] = dst_name
            logger.debug("  📸 (teacher) %s  (score=%.2f)", dst_name, photo.score)

        manifest_entries.append({
            "class": class_id,
            "number": 0,
            "name": ti.name or "先生",
            "role": ti.role or "担任",
            "is_teacher": True,
            "files": teacher_files,
        })
        exported += 1

    for group in groups:
        num = group.attendance_number
        if num is None:
            # Export with placeholder so photos aren't lost (e.g. pre-card portraits)
            if group.portraits or group.best_shot or group.card_images:
                entry: dict[str, Any] = {
                    "class": class_id,
                    "number": 0,
                    "name": "未確認",
                    "furigana": "",
                    "files": {},
                    "needs_review": True,
                }
                unk_id = f"{grade_str}{class_str}00"
                for card_i, card_path in enumerate(group.card_images, start=1):
                    if os.path.isfile(card_path):
                        tag = f"札{str(card_i).zfill(2)}"
                        dst_name = _build_filename(year, school_name, card_path, unk_id, tag)
                        shutil.copy2(card_path, class_dir / dst_name)
                        entry["files"][tag] = dst_name
                if group.best_shot and os.path.isfile(group.best_shot.path):
                    dst_name = _build_filename(year, school_name, group.best_shot.path, unk_id, "本01")
                    shutil.copy2(group.best_shot.path, class_dir / dst_name)
                    entry["files"]["本01"] = dst_name
                best_path_unk = group.best_shot.path if group.best_shot else ""
                backups = sorted(
                    [p for p in group.portraits if p.path != best_path_unk],
                    key=lambda p: p.score,
                    reverse=True,
                )
                b_limit = max_backups if max_backups > 0 else len(backups)
                for i, photo in enumerate(backups[:b_limit], start=2):
                    if os.path.isfile(photo.path):
                        tag = f"本{str(i).zfill(2)}"
                        dst_name = _build_filename(year, school_name, photo.path, unk_id, tag)
                        shutil.copy2(photo.path, class_dir / dst_name)
                        entry["files"][tag] = dst_name
                manifest_entries.append(entry)
                exported += 1
                logger.debug("  ⚠️ Exported unknown group (needs review)")
            else:
                skipped += 1
            continue

        # Look up student name from roster
        name = f"生徒{num}"
        furigana = ""
        if roster:
            student = lookup_student(roster, lookup_letter, num)
            if student:
                name = student["name"]
                furigana = student.get("furigana", "")
            else:
                logger.warning("Student #%d not found in roster for class %s",
                               num, lookup_letter)

        entry: dict[str, Any] = {
            "class": class_id,
            "number": num,
            "name": name,
            "furigana": furigana,
            "files": {},
        }

        # ── Copy card(s) (札01, 札02, …) ──
        student_id = _make_student_id(grade_str, class_str, num)
        for card_i, card_path in enumerate(group.card_images, start=1):
            if not os.path.isfile(card_path):
                continue
            tag = f"札{str(card_i).zfill(2)}"
            dst_name = _build_filename(year, school_name, card_path, student_id, tag)
            shutil.copy2(card_path, class_dir / dst_name)
            entry["files"][tag] = dst_name
            logger.debug("  📋 %s", dst_name)

        # ── Copy best shot (本01) ──
        if group.best_shot and os.path.isfile(group.best_shot.path):
            dst_name = _build_filename(year, school_name, group.best_shot.path, student_id, "本01")
            dst_path = class_dir / dst_name
            shutil.copy2(group.best_shot.path, dst_path)
            entry["files"]["本01"] = dst_name
            entry["best_score"] = group.best_shot.score
            logger.debug("  ⭐ %s  (score=%.2f)", dst_name, group.best_shot.score)

        # ── Copy backup shots (本02, 本03, …) — ALL portraits, no filtering ──
        best_path = group.best_shot.path if group.best_shot else ""
        backups = sorted(
            [p for p in group.portraits if p.path != best_path],
            key=lambda p: p.score,
            reverse=True,
        )
        limit = max_backups if max_backups > 0 else len(backups)
        for i, photo in enumerate(backups[:limit], start=2):
            tag = f"本{str(i).zfill(2)}"
            dst_name = _build_filename(year, school_name, photo.path, student_id, tag)
            dst_path = class_dir / dst_name
            shutil.copy2(photo.path, dst_path)
            entry["files"][tag] = dst_name
            logger.debug("  📸 %s  (score=%.2f)", dst_name, photo.score)

        manifest_entries.append(entry)
        exported += 1

    # ── Detect absent students ──
    all_numbers: set[int] = set()
    if roster:
        cls_data = roster.get("classes", {}).get(lookup_letter, {})
        all_numbers = {s["number"] for s in cls_data.get("students", [])}
    else:
        cls_data = {}
    present_numbers = {e["number"] for e in manifest_entries}
    absent_numbers = sorted(all_numbers - present_numbers)

    for num in absent_numbers:
        student = lookup_student(roster, lookup_letter, num) if roster else None
        manifest_entries.append({
            "class": class_id,
            "number": num,
            "name": student["name"] if student else f"生徒{num}",
            "furigana": student.get("furigana", "") if student else "",
            "files": {},
            "absent": True,
        })

    # Sort manifest by number
    manifest_entries.sort(key=lambda e: e["number"])

    manifest = {
        "class": class_id,
        "roster_letter": lookup_letter,
        "teacher": cls_data.get("teacher") if cls_data else None,
        "total_students": len(all_numbers),
        "exported": exported,
        "absent": absent_numbers,
        "skipped_unknown": skipped,
        "entries": manifest_entries,
    }

    manifest_path = output_dir / "manifest.json"
    # Merge with existing manifest if processing multiple classes
    if manifest_path.exists():
        with open(manifest_path, encoding="utf-8") as f:
            existing = json.load(f)
        if "classes" not in existing:
            existing = {"classes": {}}
        existing["classes"][class_id] = manifest
        with open(manifest_path, "w", encoding="utf-8") as f:
            json.dump(existing, f, ensure_ascii=False, indent=2)
    else:
        full_manifest = {"classes": {class_id: manifest}}
        with open(manifest_path, "w", encoding="utf-8") as f:
            json.dump(full_manifest, f, ensure_ascii=False, indent=2)

    logger.info("✅ Class %s: exported %d, absent %d, skipped %d → %s",
                class_id, exported, len(absent_numbers), skipped, class_dir)

    return manifest


def export_all_classes(
    class_groups: dict[str, ClassPhotoGroup],
    roster: dict | None,
    output_dir: str | Path,
    school_name: str = "",
    year: str = "",
    class_mapping: dict[str, str] | None = None,
    max_backups: int = 0,
) -> dict:
    """
    Export all detected classes from a camera folder.

    Parameters
    ----------
    class_groups : output from SequenceProcessor.process_folder()
    roster : parsed roster dict (or None)
    output_dir : root output directory
    school_name : school name for file naming (e.g. '千早高')
    year : 2-digit year string for file naming (e.g. '26'). Defaults to ''.
    class_mapping : optional { "3-7": "G", "3-6": "F", ... } for roster lookup
    max_backups : how many backup portraits to include

    Returns
    -------
    dict : full manifest with all classes
    """
    all_manifests: dict[str, dict] = {}

    for class_id, cpg in sorted(class_groups.items()):
        # Determine roster letter for student lookup
        roster_letter = None
        if class_mapping and class_id in class_mapping:
            roster_letter = class_mapping[class_id]
        elif cpg.class_number is not None:
            roster_letter = _grade_class_to_roster_id(
                cpg.grade or 3, cpg.class_number
            )

        logger.info("Exporting class %s (roster letter: %s, %d students)",
                    class_id, roster_letter or '?', len(cpg.students))

        manifest = export_package(
            groups=cpg.students,
            roster=roster,
            class_id=class_id,
            output_dir=output_dir,
            school_name=school_name,
            year=year,
            max_backups=max_backups,
            roster_class_letter=roster_letter,
            teacher_group=cpg.teacher,
        )
        all_manifests[class_id] = manifest

    return all_manifests

"""### Module 6 — `face_offset_calculator`
Detects the chin (landmark 152) and eye positions (landmarks 33, 263) in each best-shot portrait using MediaPipe Face Mesh. It then computes **per-student** `offsetX`, `offsetY`, and `scaleFactor` based on a **chin-anchor + eye-distance scale** strategy (v3). This ensures all faces are consistently sized and positioned relative to the chin line, regardless of hair or forehead visibility.

Writes results into `manifest.json` as `face_offsets` for each entry.
"""

"""
face_offset_calculator.py — Per-image face alignment for InDesign
=================================================================
v6: Fixed coordinate math — scale anchors at graphic centre, not frame origin.

Root causes fixed from v5:
  - compute_offsets() was computing chin position as if scale S was applied
    from the top-left corner of the fitted image (origin_y).
    In reality, InDesign scales around the GRAPHIC CENTRE.
    The correct chin-in-frame formula after centering+scaling is:
      chin_in_frame = graphic_centre_y + (chin_y - 0.5) * fitted_h * S
    And the required offsetY to land it at target_chin_mm:
      offsetY = target_chin_mm - chin_in_frame
    The old formula omitted the (chin_y - 0.5) centring step, so it
    effectively computed everything from the image top, giving offsets
    that were far too small (heads landed near the top of every frame).

  - Similarly for X: the graphic centre after centering is frame_w/2,
    so face_center_x shift from centre = (face_center_x - 0.5) * fitted_w * S
    and offsetX = target_x_mm - [frame_w/2 + (face_cx - 0.5)*fitted_w*S]

  - The JSX placeWithTransform counterpart is also fixed: it now calls
    frame.fit() once more AFTER scaling to recentre, then applies the
    mm offsets. See AutoPlacePhotosAndNames_v14.jsx.

All other logic (frame config loading, manifest processing, scale clamp
derivation) is unchanged from v5.
"""
import json, logging, os, shutil
from pathlib import Path
from typing import Optional
import cv2, mediapipe as mp, numpy as np

logger = logging.getLogger(__name__)

CHIN_IDX       = 152
NOSE_IDX       = 1
LEFT_EYE_IDX   = 33
RIGHT_EYE_IDX  = 263
_mp_face_mesh  = mp.solutions.face_mesh


# ══════════════════════════════════════════════════════════════════
# FRAME CONFIG — loads everything Python needs from the JSX export
# ══════════════════════════════════════════════════════════════════

_HARDCODED_FALLBACK = {
    "student": {
        "frame_w_mm": 36.0, "frame_h_mm": 44.0,
        "guide_ratios": {"top_ratio": 0.10, "bottom_ratio": 0.86},
        "scale_clamp":  {"min": 125, "max": 145},
    },
    "teacher": {
        "frame_w_mm": 46.0, "frame_h_mm": 54.0,
        "guide_ratios": {"top_ratio": 0.10, "bottom_ratio": 0.78},
        "scale_clamp":  {"min": 124, "max": 144},
    },
}


def load_frame_config(config_path) -> dict:
    path = Path(config_path)
    if not path.is_file():
        logger.warning(
            "frame_config.json not found at %s\n"
            "  → Run ExportFrameDimensions.jsx on your template first.\n"
            "  → Using built-in fallback values (may not match your template).",
            path,
        )
        return _HARDCODED_FALLBACK.copy()

    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    for role in ("student", "teacher"):
        if role not in data:
            data[role] = _HARDCODED_FALLBACK[role].copy()
            logger.warning("frame_config.json missing '%s' key — using fallback", role)
            continue
        if "guide_ratios" not in data[role]:
            data[role]["guide_ratios"] = _HARDCODED_FALLBACK[role]["guide_ratios"]
            logger.warning(
                "frame_config.json['%s'] has no guide_ratios — "
                "re-run ExportFrameDimensions.jsx for accurate placement", role
            )
        if "scale_clamp" not in data[role]:
            data[role]["scale_clamp"] = _HARDCODED_FALLBACK[role]["scale_clamp"]

    src = data.get("source_document", "unknown")
    gen = data.get("generated_at",    "unknown")
    print(f"  Frame config : {path.name}  (from '{src}', {gen})")
    for role in ("student", "teacher"):
        d = data[role]
        gr = d.get("guide_ratios")
        cl = d.get("scale_clamp", {})
        print(f"    {role:8s}: {d['frame_w_mm']} × {d['frame_h_mm']} mm"
              f"  guide_chin={gr['bottom_ratio'] if gr else '?'}"
              f"  scale_clamp=[{cl.get('min','?')}–{cl.get('max','?')}%]")
    return data


def get_target_params(role_config: dict, target_eye_dist_override: Optional[float] = None):
    """
    Extract target parameters for compute_offsets():
      target_chin_y   — ratio of frame height where chin should land
                        (from guide_ratios.bottom_ratio, already in [0,1])
      target_eye_dist — eye span as fraction of frame width
      scale_clamp     — (min, max) tuple

    NOTE: target_chin_y here is a RATIO (0–1 range = within the frame).
    The old code was reading values like 0.864 correctly from frame_config,
    but compute_offsets() was applying them incorrectly (see that function).
    """
    gr = role_config.get("guide_ratios")
    cl = role_config.get("scale_clamp", {"min": 125, "max": 145})
    fw = role_config["frame_w_mm"]
    fh = role_config["frame_h_mm"]

    # target_chin_y: ratio within the frame (e.g. 0.864 = 86.4% down from top)
    if gr:
        target_chin_y = gr["bottom_ratio"]
    else:
        target_chin_y = 0.86   # safe fallback (within frame)

    # target_eye_dist: eye span as fraction of frame width
    if target_eye_dist_override is not None:
        target_eye_dist = target_eye_dist_override
    elif gr:
        face_zone_h_mm  = (gr["bottom_ratio"] - gr["top_ratio"]) * fh
        # eye_dist ≈ face_zone_h * 0.55 (empirical for passport-style crops),
        # expressed as fraction of frame width
        target_eye_dist = round((face_zone_h_mm * 0.55) / fw, 3)
        target_eye_dist = max(0.16, min(0.32, target_eye_dist))
    else:
        target_eye_dist = 0.22

    return {
        "target_chin_y":    target_chin_y,
        "target_eye_dist":  target_eye_dist,
        "scale_clamp_min":  cl["min"],
        "scale_clamp_max":  cl["max"],
    }


# ══════════════════════════════════════════════════════════════════
# FACE DETECTION
# ══════════════════════════════════════════════════════════════════

def detect_face_landmarks(image_path: str, confidence: float = 0.5) -> Optional[dict]:
    img = cv2.imread(image_path)
    if img is None:
        logger.warning("Could not read image: %s", image_path)
        return None
    h, w = img.shape[:2]
    rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

    mp_image = mp.Image(image_format=mp.ImageFormat.SRGB, data=rgb)

    with _mp_face_mesh.FaceMesh(
        static_image_mode=True, max_num_faces=1,
        refine_landmarks=True, min_detection_confidence=confidence,
    ) as face_mesh:
        results = face_mesh.process(mp_image.numpy_view())

    if not results.multi_face_landmarks:
        logger.warning("No face detected: %s", os.path.basename(image_path))
        return None

    lm        = results.multi_face_landmarks[0].landmark
    chin      = lm[CHIN_IDX]
    left_eye  = lm[LEFT_EYE_IDX]
    right_eye = lm[RIGHT_EYE_IDX]
    nose      = lm[NOSE_IDX]

    eye_center_y = (left_eye.y + right_eye.y) / 2
    eye_center_x = (left_eye.x + right_eye.x) / 2
    eye_dist     = abs(right_eye.x - left_eye.x)

    return {
        "chin_y":        chin.y,
        "eye_center_y":  eye_center_y,
        "eye_center_x":  eye_center_x,
        "eye_dist":      eye_dist,
        "face_center_x": nose.x,
        "img_width":     w,
        "img_height":    h,
        "landmarks": {
            "chin":      (chin.x,      chin.y),
            "nose":      (nose.x,      nose.y),
            "left_eye":  (left_eye.x,  left_eye.y),
            "right_eye": (right_eye.x, right_eye.y),
        },
    }


# ══════════════════════════════════════════════════════════════════
# OFFSET COMPUTATION  — v6: scale anchored at graphic centre
# ══════════════════════════════════════════════════════════════════

def _simulate_fill_proportionally(img_w, img_h, frame_w_mm, frame_h_mm):
    """
    Replicate InDesign's FILL_PROPORTIONALLY + CENTER_CONTENT.

    Returns:
      fitted_w, fitted_h  — size of the scaled image (may exceed frame)
      origin_x, origin_y  — top-left of the fitted image relative to frame origin
                            (negative if image overflows the frame)
    """
    img_aspect   = img_w / img_h
    frame_aspect = frame_w_mm / frame_h_mm
    if img_aspect > frame_aspect:
        # Image is wider → fit height, overflow width
        fitted_h = frame_h_mm
        fitted_w = frame_h_mm * img_aspect
    else:
        # Image is taller → fit width, overflow height
        fitted_w = frame_w_mm
        fitted_h = frame_w_mm / img_aspect
    # After centering, origin is negative when image overflows
    origin_x = (frame_w_mm - fitted_w) / 2
    origin_y = (frame_h_mm - fitted_h) / 2
    return fitted_w, fitted_h, origin_x, origin_y


def compute_offsets(
    face_data: dict,
    frame_w_mm: float,
    frame_h_mm: float,
    target_face_center_x: float = 0.50,
    target_chin_y: float        = 0.86,   # ratio within frame (0=top, 1=bottom)
    target_eye_dist: float      = 0.22,
    scale_clamp_min: float      = 125,
    scale_clamp_max: float      = 145,
) -> dict:
    """
    Compute per-image offsetX, offsetY, scaleFactor for InDesign.

    Coordinate model (v6 — corrected):
    ─────────────────────────────────
    After frame.fit(FILL_PROPORTIONALLY) + frame.fit(CENTER_CONTENT):
      • The graphic's centre is at (frame_w/2, frame_h/2).
      • origin_x/y = (frame_w - fitted_w)/2  (negative = image overflows).

    After applying scale factor S (= scaleFactor/100):
      InDesign scales the graphic around its OWN centre (frame_w/2, frame_h/2).
      So every point P_image moves to:
        P_frame_x = frame_w/2 + (origin_x + px * fitted_w  - frame_w/2) * S
        P_frame_y = frame_h/2 + (origin_y + py * fitted_h  - frame_h/2) * S

    Simplifying for the chin point (px=face_cx, py=chin_y):
        chin_in_frame_y = frame_h/2 + (origin_y + chin_y*fitted_h - frame_h/2) * S

    The required offsetY to move chin to target_chin_mm (= target_chin_y * frame_h):
        offsetY = target_chin_mm - chin_in_frame_y

    The JSX side (placeWithTransform) must apply the move() AFTER scaling,
    starting from the already-centred position — see v14 JSX for the matching fix.
    """
    img_w = face_data["img_width"]
    img_h = face_data["img_height"]
    fitted_w, fitted_h, origin_x, origin_y = _simulate_fill_proportionally(
        img_w, img_h, frame_w_mm, frame_h_mm
    )

    face_cx  = face_data["face_center_x"]
    chin_y   = face_data["chin_y"]
    eye_dist = face_data["eye_dist"]

    # ── 1. Determine scale from eye distance ──────────────────────────────
    eye_dist_mm   = eye_dist * fitted_w                        # eyes span in fitted image
    desired_ed_mm = target_eye_dist * frame_w_mm              # desired span in frame coords
    scale_factor  = (desired_ed_mm / eye_dist_mm * 100) if eye_dist_mm > 0 else 100.0
    scale_factor  = float(np.clip(scale_factor, scale_clamp_min, scale_clamp_max))
    S = scale_factor / 100.0

    # ── 2. Where does the graphic centre sit? (always frame centre) ───────
    gc_x = frame_w_mm / 2.0   # graphic centre X in frame coords
    gc_y = frame_h_mm / 2.0   # graphic centre Y in frame coords

    # ── 3. Where does the chin land after scale? ──────────────────────────
    # In un-scaled graphic coordinates (relative to graphic's own origin):
    chin_raw_y   = origin_y + chin_y   * fitted_h   # chin Y from frame top, S=1
    face_raw_x   = origin_x + face_cx  * fitted_w   # face centre X from frame left, S=1

    # After scaling around the graphic centre:
    chin_in_frame_y = gc_y + (chin_raw_y - gc_y) * S
    face_in_frame_x = gc_x + (face_raw_x - gc_x) * S

    # ── 4. Compute offsets to hit target positions ─────────────────────────
    target_chin_mm = target_chin_y        * frame_h_mm   # e.g. 0.864 * 44 = 38.0 mm
    target_x_mm   = target_face_center_x * frame_w_mm   # e.g. 0.50  * 36 = 18.0 mm

    offset_y = target_chin_mm - chin_in_frame_y
    offset_x = target_x_mm   - face_in_frame_x

    return {
        "offsetX":       round(offset_x,     2),
        "offsetY":       round(offset_y,     2),
        "scaleFactor":   round(scale_factor, 1),
        # diagnostics
        "chin_y":         round(chin_y,   4),
        "eye_dist":       round(eye_dist, 4),
        "face_center_x":  round(face_cx,  4),
        "chin_in_frame_y_mm": round(chin_in_frame_y, 3),
        "target_chin_mm":     round(target_chin_mm,  3),
    }


# ══════════════════════════════════════════════════════════════════
# MANIFEST PROCESSOR
# ══════════════════════════════════════════════════════════════════

def process_manifest_offsets(
    manifest_path,
    package_root,
    target_cx:         float = 0.50,
    teacher_target_cx: float = 0.50,
    target_chin_y_override:           Optional[float] = None,
    target_eye_dist_override:         Optional[float] = None,
    teacher_target_chin_y_override:   Optional[float] = None,
    teacher_target_eye_dist_override: Optional[float] = None,
    base_offset_x:   float = -3.5,
    base_offset_y:   float = 12.0,
    base_scale:      int   = 127,
    frame_w_mm:         Optional[float] = None,
    frame_h_mm:         Optional[float] = None,
    teacher_frame_w_mm: Optional[float] = None,
    teacher_frame_h_mm: Optional[float] = None,
    classes:    Optional[list] = None,
    confidence: float = 0.5,
):
    """
    Read manifest.json, run face detection, compute offsets, write back.

    Key change in v6:
      compute_offsets() now correctly models InDesign's scale-around-centre
      behaviour. Offset values will be larger (more negative for Y when the
      chin needs to move down) but will result in correct face placement.
    """
    manifest_path = Path(manifest_path)
    package_root  = Path(package_root)

    frame_config_path = manifest_path.parent / "frame_config.json"
    frame_cfg = load_frame_config(frame_config_path)

    s_cfg = frame_cfg["student"]
    t_cfg = frame_cfg["teacher"]

    student_fw = frame_w_mm         or s_cfg["frame_w_mm"]
    student_fh = frame_h_mm         or s_cfg["frame_h_mm"]
    teacher_fw = teacher_frame_w_mm or t_cfg["frame_w_mm"]
    teacher_fh = teacher_frame_h_mm or t_cfg["frame_h_mm"]

    s_cfg_eff = dict(s_cfg); s_cfg_eff["frame_w_mm"] = student_fw; s_cfg_eff["frame_h_mm"] = student_fh
    t_cfg_eff = dict(t_cfg); t_cfg_eff["frame_w_mm"] = teacher_fw; t_cfg_eff["frame_h_mm"] = teacher_fh

    student_params = get_target_params(s_cfg_eff, target_eye_dist_override)
    teacher_params = get_target_params(t_cfg_eff, teacher_target_eye_dist_override)

    if target_chin_y_override is not None:
        student_params["target_chin_y"] = target_chin_y_override
    if teacher_target_chin_y_override is not None:
        teacher_params["target_chin_y"] = teacher_target_chin_y_override

    print(f"\n  Placement parameters (v6 — scale-around-centre model):")
    print(f"    Student : frame={student_fw}×{student_fh}mm"
          f"  chin_y={student_params['target_chin_y']}"
          f"  eye_dist={student_params['target_eye_dist']}"
          f"  scale=[{student_params['scale_clamp_min']}–{student_params['scale_clamp_max']}%]")
    print(f"    Teacher : frame={teacher_fw}×{teacher_fh}mm"
          f"  chin_y={teacher_params['target_chin_y']}"
          f"  eye_dist={teacher_params['target_eye_dist']}"
          f"  scale=[{teacher_params['scale_clamp_min']}–{teacher_params['scale_clamp_max']}%]")

    with open(manifest_path, encoding="utf-8") as f:
        manifest = json.load(f)

    classes_to_do = classes or list(manifest.get("classes", {}).keys())
    all_results = {}
    updated = 0; fallback = 0; failed = []

    for class_id in classes_to_do:
        class_data = manifest["classes"].get(class_id)
        if not class_data:
            print(f"  Class {class_id} not in manifest, skipping.")
            continue
        class_folder = package_root / class_id
        entries      = class_data.get("entries", [])
        class_results = []
        print(f"\n  Class {class_id}: {len(entries)} entries")

        for entry in entries:
            num        = entry.get("number", 0)
            name       = entry.get("name", "")
            files      = entry.get("files", {})
            is_teacher = entry.get("is_teacher", False)
            is_absent  = entry.get("absent", False)
            best_shot  = files.get("本01") or files.get("本_01")
            label      = "Teacher" if is_teacher else f"#{num:02d}"

            if is_teacher:
                fw, fh = teacher_fw, teacher_fh
                tcx    = teacher_target_cx
                params = teacher_params
            else:
                fw, fh = student_fw, student_fh
                tcx    = target_cx
                params = student_params

            def _fb(method, _e=entry):
                _e["face_offsets"] = {
                    "offsetX":     base_offset_x,
                    "offsetY":     base_offset_y,
                    "scaleFactor": base_scale,
                    "method":      method,
                }

            if not best_shot or is_absent:
                _fb("fallback_no_image"); fallback += 1
                class_results.append({"number": num, "name": name, "status": "skipped"})
                continue

            img_path = str(class_folder / best_shot)
            if not os.path.isfile(img_path):
                print(f"    {label} ({name}): file not found")
                _fb("fallback_missing_file")
                failed.append((class_id, num, name)); fallback += 1
                continue

            face = detect_face_landmarks(img_path, confidence)
            if face is None:
                print(f"    {label} ({name}): NO FACE DETECTED")
                _fb("fallback_no_detection")
                failed.append((class_id, num, name)); fallback += 1
                class_results.append({"number": num, "name": name, "status": "no_face"})
                continue

            offsets = compute_offsets(
                face, fw, fh,
                target_face_center_x = tcx,
                target_chin_y        = params["target_chin_y"],
                target_eye_dist      = params["target_eye_dist"],
                scale_clamp_min      = params["scale_clamp_min"],
                scale_clamp_max      = params["scale_clamp_max"],
            )
            entry["face_offsets"] = {
                "offsetX":     offsets["offsetX"],
                "offsetY":     offsets["offsetY"],
                "scaleFactor": offsets["scaleFactor"],
                "chin_y":      offsets["chin_y"],
                "eye_dist":    offsets["eye_dist"],
                "method":      "chin_anchor_eye_scale_v6_centre_pivot",
                "frame_w_mm":  round(fw, 2),
                "frame_h_mm":  round(fh, 2),
            }
            updated += 1
            print(f"    {label} ({name}) [{fw:.0f}×{fh:.0f}mm]:  "
                  f"chin={offsets['chin_y']:.3f}  eye_d={offsets['eye_dist']:.3f}  "
                  f"→ offY={offsets['offsetY']:.1f}  scale={offsets['scaleFactor']:.0f}%  "
                  f"[chin lands at {offsets['chin_in_frame_y_mm']:.1f}mm, target {offsets['target_chin_mm']:.1f}mm]")
            class_results.append({
                "number": num, "name": name, "status": "ok",
                "offsets": offsets, "face": face, "image_path": img_path,
            })

        all_results[class_id] = class_results

    backup = str(manifest_path).replace(".json", "_pre_offsets.json")
    shutil.copy2(manifest_path, backup)
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
    print(f"\n  ✅ {updated} computed, {fallback} fallback, {len(failed)} failed")
    print(f"  Backup : {backup}")
    print(f"  Saved  : {manifest_path}")
    return all_results


# =========================
# Error handling helpers
# Minimal patch extension for Stage 2 / Stage 3
# Current version:
# - no teacher/student split
# - remove person_changed_without_new_tag
# - keep duplicate_number + missing_tag_shot
# - missing_tag_shot is HARD: exclude from normal export
# - copy error files into _errors (do NOT modify testing input files)
# - if a duplicated group contains a class separator image, keep it in class export
# =========================

import os
import json
import csv
import shutil
import copy
from pathlib import Path
from collections import defaultdict, Counter


# -------------------------
# Utilities
# -------------------------

def _sorted_image_paths(folder: str | Path) -> list[str]:
    folder = Path(folder)
    items = []
    for f in sorted(os.listdir(folder)):
        p = folder / f
        if p.is_file() and p.suffix.lower() in IMAGE_EXTENSIONS:
            items.append(str(p))
    return items


def _group_key(class_id: str, group_idx: int) -> str:
    return f"{class_id}::{group_idx}"


def _parse_group_key(group_key: str) -> tuple[str, int]:
    cid, idx = group_key.split("::", 1)
    return cid, int(idx)


def _photo_exists(path: str | None) -> bool:
    return bool(path) and os.path.isfile(path)


def _ensure_list(x):
    if x is None:
        return []
    if isinstance(x, list):
        return x
    return [x]


def _dedup_keep_order(seq):
    seen = set()
    out = []
    for x in seq:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


def _group_all_paths(sg: StudentPhotoGroup) -> list[str]:
    paths = []
    paths.extend(_ensure_list(getattr(sg, "card_images", [])))
    paths.extend([p.path for p in _ensure_list(getattr(sg, "portraits", []))])
    return _dedup_keep_order(paths)


# -------------------------
# Stage 2 helpers
# -------------------------

def build_path_index(class_groups: dict[str, ClassPhotoGroup]) -> dict:
    """
    Build a per-image index from class_groups.
    """
    index = {}

    for class_id, cg in class_groups.items():
        # separator
        if getattr(cg, "separator_image", None):
            p = cg.separator_image
            index[p] = {
                "path": p,
                "class_id": class_id,
                "group_key": None,
                "group_idx": None,
                "role": "separator",
                "is_tag_shot": True,
                "student_number": None,
                "teacher_name": None,
            }

        # teacher
        if getattr(cg, "teacher", None):
            tg = cg.teacher
            teacher_name = ""
            if getattr(tg, "teacher_info", None):
                teacher_name = tg.teacher_info.name or ""

            for p in _ensure_list(getattr(tg, "card_images", [])):
                index[p] = {
                    "path": p,
                    "class_id": class_id,
                    "group_key": f"{class_id}::teacher",
                    "group_idx": None,
                    "role": "teacher_card",
                    "is_tag_shot": True,
                    "student_number": None,
                    "teacher_name": teacher_name,
                }

            for photo in _ensure_list(getattr(tg, "portraits", [])):
                p = photo.path
                index[p] = {
                    "path": p,
                    "class_id": class_id,
                    "group_key": f"{class_id}::teacher",
                    "group_idx": None,
                    "role": "teacher_portrait",
                    "is_tag_shot": False,
                    "student_number": None,
                    "teacher_name": teacher_name,
                }

        # students
        for group_idx, sg in enumerate(getattr(cg, "students", [])):
            gkey = _group_key(class_id, group_idx)
            num = getattr(sg, "attendance_number", None)

            for p in _ensure_list(getattr(sg, "card_images", [])):
                index[p] = {
                    "path": p,
                    "class_id": class_id,
                    "group_key": gkey,
                    "group_idx": group_idx,
                    "role": "student_card",
                    "is_tag_shot": True,
                    "student_number": num,
                    "teacher_name": None,
                }

            for photo in _ensure_list(getattr(sg, "portraits", [])):
                p = photo.path
                index[p] = {
                    "path": p,
                    "class_id": class_id,
                    "group_key": gkey,
                    "group_idx": group_idx,
                    "role": "student_portrait",
                    "is_tag_shot": False,
                    "student_number": num,
                    "teacher_name": None,
                }

    return index


def _get_all_separator_paths(class_groups: dict[str, ClassPhotoGroup]) -> set[str]:
    """
    Collect all class separator image paths.
    Used to avoid sending separator-related student groups into duplicate_number.
    """
    separator_paths = set()
    for _, cg in class_groups.items():
        p = getattr(cg, "separator_image", None)
        if p:
            separator_paths.add(p)
    return separator_paths


def detect_duplicate_numbers(class_groups: dict[str, ClassPhotoGroup]) -> list[dict]:
    """
    group(class)-level hard error:
    same attendance number assigned to multiple student groups in the same class

    Special rule:
    if a duplicated group contains a class separator image, do NOT send that group
    to duplicate_number. Keep that group in the class export first.
    """
    errors = []
    separator_paths = _get_all_separator_paths(class_groups)

    for class_id, cg in class_groups.items():
        num_to_groups = defaultdict(list)

        for group_idx, sg in enumerate(getattr(cg, "students", [])):
            num = getattr(sg, "attendance_number", None)
            if num is None:
                continue
            num_to_groups[num].append((group_idx, sg))

        for num, items in num_to_groups.items():
            if len(items) <= 1:
                continue

            # Exempt groups that contain a separator image
            filtered_items = []

            for group_idx, sg in items:
                group_paths = _group_all_paths(sg)
                has_separator = any(p in separator_paths for p in group_paths)
                if not has_separator:
                    filtered_items.append((group_idx, sg))

            # If after exempting separator-related groups there is <= 1 real duplicate left,
            # do not create duplicate_number error.
            if len(filtered_items) <= 1:
                continue

            related_group_keys = []
            related_paths = []

            for group_idx, sg in filtered_items:
                related_group_keys.append(_group_key(class_id, group_idx))
                related_paths.extend(_group_all_paths(sg))

            errors.append({
                "error_type": "duplicate_number",
                "severity": "hard",
                "detection_unit": "class",
                "class_id": class_id,
                "group_key": None,
                "group_keys": related_group_keys,
                "group_idx": None,
                "student_number": num,
                "image_path": None,
                "related_paths": _dedup_keep_order(related_paths),
                "message": (
                    f"Student number {num} assigned to multiple groups in class {class_id}. "
                    f"Groups containing separator images were excluded from duplicate handling."
                ),
            })

    return errors


def detect_missing_tag_shots(class_groups: dict[str, ClassPhotoGroup]) -> list[dict]:
    """
    group-level HARD error:
    unknown student group (attendance_number is None)

    Policy:
    - exclude the whole unknown group from normal export
    - still copy related images into _errors/missing_tag_shot
    - DO NOT change any files in the testing input folder
    """
    errors = []

    for class_id, cg in class_groups.items():
        for group_idx, sg in enumerate(getattr(cg, "students", [])):
            num = getattr(sg, "attendance_number", None)
            if num is not None:
                continue

            gkey = _group_key(class_id, group_idx)
            related_paths = _group_all_paths(sg)

            if not related_paths:
                continue

            errors.append({
                "error_type": "missing_tag_shot",
                "severity": "hard",
                "detection_unit": "group",
                "class_id": class_id,
                "group_key": gkey,
                "group_keys": [gkey],
                "group_idx": group_idx,
                "student_number": None,
                "image_path": related_paths[0],
                "related_paths": related_paths,
                "message": f"Student group has no valid student tag shot in class {class_id}.",
            })

    return errors


def detect_pipeline_errors(
    class_groups: dict[str, ClassPhotoGroup],
    roster: dict | None,
    photos_path: str | Path,
) -> list[dict]:
    """
    Run agreed error detectors.
    Current version:
    - duplicate_number
    - missing_tag_shot
    - person_changed_without_new_tag removed
    """
    errors = []
    errors.extend(detect_duplicate_numbers(class_groups))
    errors.extend(detect_missing_tag_shots(class_groups))

    def _sort_key(e):
        return (
            e.get("class_id") or "",
            e.get("error_type") or "",
            e.get("image_path") or "",
            str(e.get("student_number") or ""),
        )

    return sorted(errors, key=_sort_key)


def save_error_queue(error_queue: list[dict], path: str | Path):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(error_queue, f, ensure_ascii=False, indent=2)


def attach_error_tags_to_groups(class_groups: dict[str, ClassPhotoGroup], error_queue: list[dict]):
    """
    Minimal non-invasive attachment:
    dynamically attach .error_tags to groups / photos
    """
    group_tags = defaultdict(list)
    image_tags = defaultdict(list)

    for err in error_queue:
        et = err["error_type"]

        for gk in _ensure_list(err.get("group_keys", [])):
            if gk:
                group_tags[gk].append(et)

        if err.get("group_key"):
            group_tags[err["group_key"]].append(et)

        for p in _ensure_list(err.get("related_paths", [])):
            if p:
                image_tags[p].append(et)

        if err.get("image_path"):
            image_tags[err["image_path"]].append(et)

    for class_id, cg in class_groups.items():
        for group_idx, sg in enumerate(getattr(cg, "students", [])):
            gk = _group_key(class_id, group_idx)
            setattr(sg, "error_tags", sorted(set(group_tags.get(gk, []))))

            for photo in _ensure_list(getattr(sg, "portraits", [])):
                setattr(photo, "error_tags", sorted(set(image_tags.get(photo.path, []))))

        if getattr(cg, "teacher", None):
            tg = cg.teacher
            setattr(tg, "error_tags", [])
            for photo in _ensure_list(getattr(tg, "portraits", [])):
                setattr(photo, "error_tags", sorted(set(image_tags.get(photo.path, []))))


# -------------------------
# Stage 3 helpers
# -------------------------

def _pick_best_student_patch(group: StudentPhotoGroup):
    valid = [p for p in _ensure_list(getattr(group, "portraits", [])) if not getattr(p, "is_ng", False)]
    if valid:
        group.best_shot = max(valid, key=lambda p: p.score)
    elif getattr(group, "portraits", None):
        group.best_shot = max(group.portraits, key=lambda p: p.score)
        if group.best_shot:
            group.best_shot.comment = (group.best_shot.comment or "") + " [all-NG fallback]"
    else:
        group.best_shot = None


def _pick_best_teacher_patch(group: TeacherPhotoGroup):
    valid = [p for p in _ensure_list(getattr(group, "portraits", [])) if not getattr(p, "is_ng", False)]
    if valid:
        group.best_shot = max(valid, key=lambda p: p.score)
    elif getattr(group, "portraits", None):
        group.best_shot = max(group.portraits, key=lambda p: p.score)
        if group.best_shot:
            group.best_shot.comment = (group.best_shot.comment or "") + " [all-NG fallback]"
    else:
        group.best_shot = None


def build_exportable_class_groups(
    class_groups: dict[str, ClassPhotoGroup],
    error_queue: list[dict],
) -> dict[str, ClassPhotoGroup]:
    """
    Exclude only hard-error items from normal export.

    Current policy:
    - duplicate_number -> exclude all related student groups
    - missing_tag_shot -> exclude all related unknown student groups
    - NEVER modify the testing/input folder itself
    """
    exportable = copy.deepcopy(class_groups)

    hard_group_keys = set()

    for err in error_queue:
        if err.get("severity") != "hard":
            continue

        if err.get("error_type") in {"duplicate_number", "missing_tag_shot"}:
            for gk in _ensure_list(err.get("group_keys", [])):
                if gk:
                    hard_group_keys.add(gk)

    for class_id, cg in exportable.items():
        new_students = []

        for group_idx, sg in enumerate(getattr(cg, "students", [])):
            gk = _group_key(class_id, group_idx)

            # hard-remove duplicate groups and missing-tag groups from NORMAL export
            if gk in hard_group_keys:
                continue

            _pick_best_student_patch(sg)
            new_students.append(sg)

        cg.students = new_students

        if getattr(cg, "teacher", None):
            tg = cg.teacher
            _pick_best_teacher_patch(tg)

    return exportable


def export_error_items(error_queue: list[dict], output_dir: str | Path):
    """
    Copy affected source images into:
      OUTPUT/_errors/<error_type>/

    IMPORTANT:
    - copy only
    - do NOT move
    - do NOT change/delete anything in the testing input folder
    """
    output_dir = Path(output_dir)
    error_root = output_dir / "_errors"
    error_root.mkdir(parents=True, exist_ok=True)

    copied = 0

    for err in error_queue:
        err_type = err["error_type"]
        dst_dir = error_root / err_type
        dst_dir.mkdir(parents=True, exist_ok=True)

        class_id = err.get("class_id") or "unknown_class"

        for src in _dedup_keep_order(_ensure_list(err.get("related_paths", []))):
            if not _photo_exists(src):
                continue

            base = Path(src).name
            dst_name = f"{class_id}__{base}"
            dst_path = dst_dir / dst_name

            # avoid duplicate overwrite collisions
            if dst_path.exists():
                stem = Path(base).stem
                suf = Path(base).suffix
                i = 2
                while True:
                    alt = dst_dir / f"{class_id}__{stem}__{i}{suf}"
                    if not alt.exists():
                        dst_path = alt
                        break
                    i += 1

            shutil.copy2(src, dst_path)
            copied += 1

    return copied


def write_error_log_json(error_queue: list[dict], path: str | Path):
    path = Path(path)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(error_queue, f, ensure_ascii=False, indent=2)


def write_error_log_csv(error_queue: list[dict], path: str | Path):
    path = Path(path)

    fieldnames = [
        "error_type",
        "severity",
        "detection_unit",
        "class_id",
        "group_key",
        "group_idx",
        "student_number",
        "image_path",
        "group_keys",
        "related_paths",
        "message",
    ]

    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for err in error_queue:
            row = dict(err)
            row["group_keys"] = " | ".join(_ensure_list(row.get("group_keys", [])))
            row["related_paths"] = " | ".join(_ensure_list(row.get("related_paths", [])))
            writer.writerow({k: row.get(k, "") for k in fieldnames})


def summarize_error_queue(error_queue: list[dict]) -> dict[str, int]:
    return dict(Counter(err["error_type"] for err in error_queue))

# =========================
# Manual resolution helpers
# Add this block into HelperFunctions.py
# =========================

import os
import json
from pathlib import Path


# -------------------------
# Small internal helpers
# -------------------------

def _normalise_error_type(error_type: str) -> str:
    """
    Accept both UI-facing names and pipeline/internal names.
    """
    if not error_type:
        return ""
    et = str(error_type).strip().lower()
    mapping = {
        "no_card_detected": "no_card_detected",
        "card_not_detected": "no_card_detected",
        "札未検出": "no_card_detected",
        "missing_tag_shot": "no_card_detected",

        "multiple_card_detected": "multiple_card_detected",
        "multiple_card_holders": "multiple_card_detected",
        "複数札検出": "multiple_card_detected",
        "duplicate_number": "multiple_card_detected",
    }
    return mapping.get(et, et)


def _parse_class_id_simple(class_id: str) -> tuple[str, str]:
    """
    Examples:
      '3-1'   -> ('3', '1')
      '3年A組' -> ('3', 'A')
      'A'     -> ('', 'A')
      '1'     -> ('', '1')
    """
    if not class_id:
        return "", ""

    s = str(class_id).strip()

    # e.g. 3-1
    if "-" in s:
        parts = s.split("-", 1)
        return parts[0].strip(), parts[1].strip()

    # e.g. 3年A組 / 3年1組
    import re
    m = re.search(r"(\d+)\s*年\s*([A-Za-z0-9]+)\s*組", s)
    if m:
        return m.group(1), m.group(2)

    # fallback
    return "", s


def _safe_lookup_student(
    roster: dict | None,
    class_id: str,
    student_no: int,
    class_mapping: dict | None = None,
) -> dict | None:
    """
    Try to find the student in roster using class_id.
    Compatible with both numeric-class rosters and mapped letter rosters.
    """
    if not roster:
        return None

    grade_str, class_str = _parse_class_id_simple(class_id)
    lookup_keys = []

    # 1) direct class_id
    lookup_keys.append(str(class_id))

    # 2) plain parsed class_str
    if class_str:
        lookup_keys.append(str(class_str))

    # 3) mapped roster key (if provided)
    if class_mapping and grade_str and class_str:
        key = f"{grade_str}-{class_str}"
        mapped = class_mapping.get(key)
        if mapped:
            lookup_keys.append(str(mapped))

    # 4) unique order
    seen = set()
    lookup_keys = [k for k in lookup_keys if not (k in seen or seen.add(k))]

    for k in lookup_keys:
        cls = roster.get("classes", {}).get(k)
        if not cls:
            continue
        for s in cls.get("students", []):
            try:
                if int(s.get("number")) == int(student_no):
                    return s
            except Exception:
                continue

    return None


def _sanitise_filename_local(name: str) -> str:
    """
    Remove Windows-illegal characters.
    """
    return "".join(c for c in str(name) if c not in r'<>:"/\|?*').strip()


def _build_resolved_filename_local(
    *,
    year: str,
    school_name: str,
    image_path: str,
    class_id: str,
    student_no: int,
    tag: str = "札01",
) -> str:
    """
    Build a preview filename similar to package_exporter naming.

    Example:
      26_上水高_IMG_0042_3107_札01.jpg
    """
    year = str(year).strip() if year is not None else ""
    school_name = _sanitise_filename_local(school_name or "")
    orig_stem = _sanitise_filename_local(Path(image_path).stem)
    ext = Path(image_path).suffix or ".jpg"

    grade_str, class_str = _parse_class_id_simple(class_id)
    num_part = str(int(student_no)).zfill(2) if student_no is not None else "00"

    # If grade/class can be parsed -> [grade][class][no]
    if grade_str or class_str:
        id_str = f"{grade_str}{class_str}{num_part}"
    else:
        id_str = f"{num_part}"

    parts = [p for p in [year, school_name, orig_stem, id_str, tag] if p != ""]
    return "_".join(parts) + ext


def _base_resolution_payload(error_item: dict) -> dict:
    """
    Common output structure for one resolved item.
    """
    image_path = error_item.get("image_path") or ""
    return {
        "error_id": error_item.get("error_id"),
        "error_type": _normalise_error_type(error_item.get("error_type", "")),
        "image_path": image_path,
        "image_name": Path(image_path).name if image_path else error_item.get("image_name", ""),
        "class_id": error_item.get("class_id"),
        "group_key": error_item.get("group_key"),
        "group_keys": error_item.get("group_keys", []),
        "status": "resolved",
    }


# -------------------------
# Manual resolution: no card detected
# -------------------------

def resolve_no_card_detected(
    error_item: dict,
    student_no: int,
    roster: dict | None = None,
    *,
    year: str = "",
    school_name: str = "",
    class_mapping: dict | None = None,
    tag: str = "札01",
) -> dict:
    """
    Resolve a 'no card detected' style error by manually inputting student number.

    Supported incoming error_type aliases:
      - no_card_detected
      - card_not_detected
      - missing_tag_shot
      - 札未検出

    Returns one resolution dict.
    """
    et = _normalise_error_type(error_item.get("error_type", ""))
    if et != "no_card_detected":
        raise ValueError(
            f"resolve_no_card_detected() received incompatible error_type: {error_item.get('error_type')}"
        )

    if student_no is None:
        raise ValueError("student_no cannot be None")
    student_no = int(student_no)

    class_id = error_item.get("class_id") or ""
    student = _safe_lookup_student(
        roster=roster,
        class_id=class_id,
        student_no=student_no,
        class_mapping=class_mapping,
    )

    payload = _base_resolution_payload(error_item)
    payload.update({
        "resolution_type": "manual_student_number",
        "student_no": student_no,
        "student_name": student.get("name", "") if student else "",
        "furigana": student.get("furigana", "") if student else "",
        "gender": student.get("gender", "") if student else "",
        "suggested_filename": _build_resolved_filename_local(
            year=year,
            school_name=school_name,
            image_path=error_item.get("image_path") or error_item.get("image_name") or "unknown.jpg",
            class_id=class_id,
            student_no=student_no,
            tag=tag,
        ),
        "notes": f"Resolved by manual student number input: {student_no}",
    })
    return payload


# -------------------------
# Manual resolution: multiple card detected
# -------------------------

def resolve_multiple_card_detected(
    error_item: dict,
    selected_student_no: int,
    roster: dict | None = None,
    *,
    year: str = "",
    school_name: str = "",
    class_mapping: dict | None = None,
    tag: str = "札01",
) -> dict:
    """
    Resolve a 'multiple card detected / duplicate candidate' style error
    by manually selecting the correct student number.

    Supported incoming error_type aliases:
      - multiple_card_detected
      - multiple_card_holders
      - duplicate_number
      - 複数札検出

    Returns one resolution dict.
    """
    et = _normalise_error_type(error_item.get("error_type", ""))
    if et != "multiple_card_detected":
        raise ValueError(
            f"resolve_multiple_card_detected() received incompatible error_type: {error_item.get('error_type')}"
        )

    if selected_student_no is None:
        raise ValueError("selected_student_no cannot be None")
    selected_student_no = int(selected_student_no)

    class_id = error_item.get("class_id") or ""
    student = _safe_lookup_student(
        roster=roster,
        class_id=class_id,
        student_no=selected_student_no,
        class_mapping=class_mapping,
    )

    payload = _base_resolution_payload(error_item)
    payload.update({
        "resolution_type": "manual_candidate_selection",
        "student_no": selected_student_no,
        "student_name": student.get("name", "") if student else "",
        "furigana": student.get("furigana", "") if student else "",
        "gender": student.get("gender", "") if student else "",
        "suggested_filename": _build_resolved_filename_local(
            year=year,
            school_name=school_name,
            image_path=error_item.get("image_path") or error_item.get("image_name") or "unknown.jpg",
            class_id=class_id,
            student_no=selected_student_no,
            tag=tag,
        ),
        "notes": f"Resolved by manual candidate selection: {selected_student_no}",
    })
    return payload


# -------------------------
# Resolution persistence helpers
# -------------------------

def upsert_error_resolution(
    resolutions: list[dict],
    resolution: dict,
) -> list[dict]:
    """
    Insert or replace a resolution by error_id.
    """
    if "error_id" not in resolution:
        raise ValueError("resolution must contain 'error_id'")

    error_id = resolution["error_id"]
    replaced = False
    out = []

    for r in resolutions:
        if r.get("error_id") == error_id:
            out.append(resolution)
            replaced = True
        else:
            out.append(r)

    if not replaced:
        out.append(resolution)

    return out


def save_error_resolutions(
    resolutions: list[dict],
    path: str | Path,
):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(resolutions, f, ensure_ascii=False, indent=2)


def load_error_resolutions(
    path: str | Path,
) -> list[dict]:
    path = Path(path)
    if not path.exists():
        return []
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise ValueError(f"error resolutions file must contain a list: {path}")
    return data


def mark_error_deleted(
    error_item: dict,
    *,
    reason: str = "deleted_by_user",
) -> dict:
    """
    Optional helper: when user chooses 'delete' in UI.
    """
    payload = _base_resolution_payload(error_item)
    payload.update({
        "status": "deleted",
        "resolution_type": "manual_delete",
        "student_no": None,
        "student_name": "",
        "suggested_filename": None,
        "notes": reason,
    })
    return payload

# =========================
# PATCH CELL 1
# Manual resolution + apply-to-export helpers
# Put this in a NEW CELL after the existing "Helper Functions" cell
# and before Stage 2 / Stage 3 execution.
# =========================

import copy
import json
import os
import re
from pathlib import Path


# -------------------------
# Manual resolution helpers
# -------------------------

def _normalise_error_type(error_type: str) -> str:
    """
    Accept both UI-facing names and pipeline/internal names.
    """
    if not error_type:
        return ""
    et = str(error_type).strip().lower()
    mapping = {
        "no_card_detected": "no_card_detected",
        "card_not_detected": "no_card_detected",
        "札未検出": "no_card_detected",
        "missing_tag_shot": "no_card_detected",

        "multiple_card_detected": "multiple_card_detected",
        "multiple_card_holders": "multiple_card_detected",
        "複数札検出": "multiple_card_detected",
        "duplicate_number": "multiple_card_detected",
    }
    return mapping.get(et, et)


def _parse_class_id_simple(class_id: str) -> tuple[str, str]:
    """
    Examples:
      '3-1'    -> ('3', '1')
      '3年A組' -> ('3', 'A')
      'A'      -> ('', 'A')
      '1'      -> ('', '1')
    """
    if not class_id:
        return "", ""

    s = str(class_id).strip()

    if "-" in s:
        parts = s.split("-", 1)
        return parts[0].strip(), parts[1].strip()

    m = re.search(r"(\d+)\s*年\s*([A-Za-z0-9]+)\s*組", s)
    if m:
        return m.group(1), m.group(2)

    return "", s


def _safe_lookup_student(
    roster: dict | None,
    class_id: str,
    student_no: int,
    class_mapping: dict | None = None,
) -> dict | None:
    """
    Try to find the student in roster using class_id.
    Compatible with both numeric-class rosters and mapped letter rosters.
    """
    if not roster:
        return None

    grade_str, class_str = _parse_class_id_simple(class_id)
    lookup_keys = []

    # direct class_id
    lookup_keys.append(str(class_id))

    # parsed class part
    if class_str:
        lookup_keys.append(str(class_str))

    # mapped roster key
    if class_mapping and grade_str and class_str:
        key = f"{grade_str}-{class_str}"
        mapped = class_mapping.get(key)
        if mapped:
            lookup_keys.append(str(mapped))

    seen = set()
    lookup_keys = [k for k in lookup_keys if not (k in seen or seen.add(k))]

    for k in lookup_keys:
        cls = roster.get("classes", {}).get(k)
        if not cls:
            continue
        for s in cls.get("students", []):
            try:
                if int(s.get("number")) == int(student_no):
                    return s
            except Exception:
                continue

    return None


def _sanitise_filename_local(name: str) -> str:
    return "".join(c for c in str(name) if c not in r'<>:"/\|?*').strip()


def _build_resolved_filename_local(
    *,
    year: str,
    school_name: str,
    image_path: str,
    class_id: str,
    student_no: int,
    tag: str = "札01",
) -> str:
    """
    Build a preview filename similar to package_exporter naming.
    """
    year = str(year).strip() if year is not None else ""
    school_name = _sanitise_filename_local(school_name or "")
    orig_stem = _sanitise_filename_local(Path(image_path).stem)
    ext = Path(image_path).suffix or ".jpg"

    grade_str, class_str = _parse_class_id_simple(class_id)
    num_part = str(int(student_no)).zfill(2) if student_no is not None else "00"

    if grade_str or class_str:
        id_str = f"{grade_str}{class_str}{num_part}"
    else:
        id_str = f"{num_part}"

    parts = [p for p in [year, school_name, orig_stem, id_str, tag] if p != ""]
    return "_".join(parts) + ext


def _base_resolution_payload(error_item: dict) -> dict:
    image_path = error_item.get("image_path") or ""
    return {
        "error_id": error_item.get("error_id"),
        "error_type": _normalise_error_type(error_item.get("error_type", "")),
        "image_path": image_path,
        "image_name": Path(image_path).name if image_path else error_item.get("image_name", ""),
        "class_id": error_item.get("class_id"),
        "group_key": error_item.get("group_key"),
        "group_keys": error_item.get("group_keys", []),
        "status": "resolved",
    }


def resolve_no_card_detected(
    error_item: dict,
    student_no: int,
    roster: dict | None = None,
    *,
    year: str = "",
    school_name: str = "",
    class_mapping: dict | None = None,
    tag: str = "札01",
) -> dict:
    """
    Resolve a no-card error by manually inputting student number.
    Compatible with missing_tag_shot.
    """
    et = _normalise_error_type(error_item.get("error_type", ""))
    if et != "no_card_detected":
        raise ValueError(
            f"resolve_no_card_detected() received incompatible error_type: {error_item.get('error_type')}"
        )

    if student_no is None:
        raise ValueError("student_no cannot be None")
    student_no = int(student_no)

    class_id = error_item.get("class_id") or ""
    student = _safe_lookup_student(
        roster=roster,
        class_id=class_id,
        student_no=student_no,
        class_mapping=class_mapping,
    )

    payload = _base_resolution_payload(error_item)
    payload.update({
        "resolution_type": "manual_student_number",
        "student_no": student_no,
        "student_name": student.get("name", "") if student else "",
        "furigana": student.get("furigana", "") if student else "",
        "gender": student.get("gender", "") if student else "",
        "suggested_filename": _build_resolved_filename_local(
            year=year,
            school_name=school_name,
            image_path=error_item.get("image_path") or error_item.get("image_name") or "unknown.jpg",
            class_id=class_id,
            student_no=student_no,
            tag=tag,
        ),
        "notes": f"Resolved by manual student number input: {student_no}",
    })
    return payload


def resolve_multiple_card_detected(
    error_item: dict,
    selected_student_no: int,
    roster: dict | None = None,
    *,
    year: str = "",
    school_name: str = "",
    class_mapping: dict | None = None,
    tag: str = "札01",
    selected_group_key: str | None = None,
) -> dict:
    """
    Resolve a multiple-card / duplicate candidate error
    by manually selecting the correct student number.
    Compatible with duplicate_number.
    """
    et = _normalise_error_type(error_item.get("error_type", ""))
    if et != "multiple_card_detected":
        raise ValueError(
            f"resolve_multiple_card_detected() received incompatible error_type: {error_item.get('error_type')}"
        )

    if selected_student_no is None:
        raise ValueError("selected_student_no cannot be None")
    selected_student_no = int(selected_student_no)

    class_id = error_item.get("class_id") or ""
    student = _safe_lookup_student(
        roster=roster,
        class_id=class_id,
        student_no=selected_student_no,
        class_mapping=class_mapping,
    )

    payload = _base_resolution_payload(error_item)
    payload.update({
        "resolution_type": "manual_candidate_selection",
        "student_no": selected_student_no,
        "student_name": student.get("name", "") if student else "",
        "furigana": student.get("furigana", "") if student else "",
        "gender": student.get("gender", "") if student else "",
        "group_key": selected_group_key or error_item.get("group_key"),
        "group_keys": error_item.get("group_keys", []),
        "suggested_filename": _build_resolved_filename_local(
            year=year,
            school_name=school_name,
            image_path=error_item.get("image_path") or error_item.get("image_name") or "unknown.jpg",
            class_id=class_id,
            student_no=selected_student_no,
            tag=tag,
        ),
        "notes": f"Resolved by manual candidate selection: {selected_student_no}",
    })
    return payload


def mark_error_deleted(
    error_item: dict,
    *,
    reason: str = "deleted_by_user",
) -> dict:
    payload = _base_resolution_payload(error_item)
    payload.update({
        "status": "deleted",
        "resolution_type": "manual_delete",
        "student_no": None,
        "student_name": "",
        "suggested_filename": None,
        "notes": reason,
    })
    return payload


def upsert_error_resolution(
    resolutions: list[dict],
    resolution: dict,
) -> list[dict]:
    """
    Insert or replace a resolution.
    If error_id exists, use it as primary key.
    Otherwise fallback to (error_type, image_path, group_key, class_id).
    """
    out = []
    replaced = False

    rid = resolution.get("error_id")
    rkey = (
        _normalise_error_type(resolution.get("error_type", "")),
        resolution.get("image_path"),
        resolution.get("group_key"),
        resolution.get("class_id"),
    )

    for r in resolutions:
        same = False

        if rid is not None and r.get("error_id") == rid:
            same = True
        else:
            old_key = (
                _normalise_error_type(r.get("error_type", "")),
                r.get("image_path"),
                r.get("group_key"),
                r.get("class_id"),
            )
            if old_key == rkey:
                same = True

        if same:
            out.append(resolution)
            replaced = True
        else:
            out.append(r)

    if not replaced:
        out.append(resolution)

    return out


def save_error_resolutions(
    resolutions: list[dict],
    path: str | Path,
):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(resolutions, f, ensure_ascii=False, indent=2)


def load_error_resolutions(
    path: str | Path,
) -> list[dict]:
    path = Path(path)
    if not path.exists():
        return []
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise ValueError(f"error resolutions file must contain a list: {path}")
    return data


# -------------------------
# Apply resolutions back into export flow
# -------------------------

def _find_student_group_by_group_key(class_groups: dict[str, ClassPhotoGroup], group_key: str):
    """
    Find (class_id, class_group, student_group) by group_key like '3-1::2'
    """
    if not group_key or "::" not in group_key:
        return None, None, None

    class_id, idx_str = group_key.split("::", 1)
    if idx_str == "teacher":
        return class_id, class_groups.get(class_id), None

    try:
        idx = int(idx_str)
    except Exception:
        return None, None, None

    cg = class_groups.get(class_id)
    if cg is None:
        return None, None, None

    students = getattr(cg, "students", [])
    if idx < 0 or idx >= len(students):
        return class_id, cg, None

    return class_id, cg, students[idx]


def _find_student_group_by_image_path(class_groups: dict[str, ClassPhotoGroup], image_path: str):
    """
    Find the student group containing this image path.
    Returns (class_id, class_group, student_group, group_key)
    """
    if not image_path:
        return None, None, None, None

    for class_id, cg in class_groups.items():
        for group_idx, sg in enumerate(getattr(cg, "students", [])):
            card_images = _ensure_list(getattr(sg, "card_images", []))
            portrait_paths = [p.path for p in _ensure_list(getattr(sg, "portraits", []))]
            all_paths = set(card_images + portrait_paths)

            if image_path in all_paths:
                return class_id, cg, sg, _group_key(class_id, group_idx)

    return None, None, None, None


def _move_image_from_portraits_to_cards(student_group: StudentPhotoGroup, image_path: str):
    """
    If image_path is currently inside portraits, move it to card_images.
    Useful for manually resolved no-card cases.
    """
    if not image_path:
        return

    new_portraits = []
    moved = False

    for photo in _ensure_list(getattr(student_group, "portraits", [])):
        if photo.path == image_path and not moved:
            moved = True
            continue
        new_portraits.append(photo)

    if moved:
        current_cards = _ensure_list(getattr(student_group, "card_images", []))
        if image_path not in current_cards:
            current_cards.append(image_path)
        student_group.card_images = current_cards
        student_group.portraits = new_portraits


def _remove_image_from_student_group(student_group: StudentPhotoGroup, image_path: str):
    if not image_path:
        return

    student_group.card_images = [
        p for p in _ensure_list(getattr(student_group, "card_images", []))
        if p != image_path
    ]
    student_group.portraits = [
        photo for photo in _ensure_list(getattr(student_group, "portraits", []))
        if photo.path != image_path
    ]


def _is_resolution_effective(resolution: dict) -> bool:
    status = str(resolution.get("status", "")).strip().lower()
    return status in {"resolved", "deleted"}


def filter_unresolved_error_queue(
    error_queue: list[dict],
    resolutions: list[dict],
) -> list[dict]:
    """
    Remove error items that already have an effective resolution.
    Match rule:
      1. error_id if available
      2. fallback to (error_type, image_path, group_key, class_id)
    """
    resolved_ids = set()
    resolved_keys = set()

    for r in resolutions:
        if not _is_resolution_effective(r):
            continue

        if r.get("error_id") is not None:
            resolved_ids.add(r.get("error_id"))

        resolved_keys.add((
            _normalise_error_type(r.get("error_type", "")),
            r.get("image_path"),
            r.get("group_key"),
            r.get("class_id"),
        ))

    remaining = []
    for err in error_queue:
        matched = False

        if err.get("error_id") is not None and err.get("error_id") in resolved_ids:
            matched = True
        else:
            err_key = (
                _normalise_error_type(err.get("error_type", "")),
                err.get("image_path"),
                err.get("group_key"),
                err.get("class_id"),
            )
            if err_key in resolved_keys:
                matched = True

        if not matched:
            remaining.append(err)

    return remaining


def apply_error_resolutions_to_class_groups(
    class_groups: dict[str, ClassPhotoGroup],
    resolutions: list[dict],
) -> dict[str, ClassPhotoGroup]:
    """
    Apply manual resolutions to a deep-copied class_groups object.

    Supported resolution_type:
      - manual_student_number
      - manual_candidate_selection
      - manual_delete
    """
    updated = copy.deepcopy(class_groups)

    for res in resolutions:
        if not _is_resolution_effective(res):
            continue

        resolution_type = str(res.get("resolution_type", "")).strip().lower()
        image_path = res.get("image_path")
        group_key = res.get("group_key")
        group_keys = _ensure_list(res.get("group_keys", []))
        student_no = res.get("student_no")

        # manual delete
        if resolution_type == "manual_delete":
            _, _, sg, _ = _find_student_group_by_image_path(updated, image_path)
            if sg is not None:
                _remove_image_from_student_group(sg, image_path)
                _pick_best_student_patch(sg)
            continue

        # manual student number
        if resolution_type == "manual_student_number":
            target_sg = None

            if group_key:
                _, _, target_sg = _find_student_group_by_group_key(updated, group_key)
            if target_sg is None and image_path:
                _, _, target_sg, _ = _find_student_group_by_image_path(updated, image_path)

            if target_sg is not None:
                target_sg.attendance_number = int(student_no)
                if image_path:
                    _move_image_from_portraits_to_cards(target_sg, image_path)
                _pick_best_student_patch(target_sg)

            continue

        # manual candidate selection
        if resolution_type == "manual_candidate_selection":
            chosen_group_key = group_key or (group_keys[0] if group_keys else None)

            if chosen_group_key:
                _, _, chosen_sg = _find_student_group_by_group_key(updated, chosen_group_key)
                if chosen_sg is not None:
                    chosen_sg.attendance_number = int(student_no)
                    if image_path:
                        _move_image_from_portraits_to_cards(chosen_sg, image_path)
                    _pick_best_student_patch(chosen_sg)

            # neutralize other duplicate groups
            for gk in group_keys:
                if not gk or gk == chosen_group_key:
                    continue
                _, _, other_sg = _find_student_group_by_group_key(updated, gk)
                if other_sg is not None:
                    other_sg.attendance_number = None
                    _pick_best_student_patch(other_sg)

            continue

    return updated

