from pathlib import Path
from datetime import datetime

import pytest
from openpyxl import load_workbook

from backend.excel_labels import TEACHER_SHEET_LABELS, excel_label, translate_display_value
import numpy as np
from PIL import Image
import zipfile


def test_snap_excel_labels_japanese(tmp_path: Path):
    pytest.importorskip("torch")
    from backend.snap_pipeline import SnapPipeline, PipelineResult, ImageRecord
    pipeline = SnapPipeline.__new__(SnapPipeline)

    img_path = tmp_path / "a.jpg"
    Image.new("RGB", (16, 16), color=(120, 120, 120)).save(img_path)
    rec = ImageRecord(
        path=img_path,
        capture_time=datetime(2026, 1, 2, 3, 4, 5),
        sequence_tail=1,
        embedding=np.zeros(3, dtype=np.float32),
        focus_score=1.0,
        brightness_score=1.0,
    )
    summary = PipelineResult(1, 1, 1, 0.0, 1, 1, 0, 0)

    out_dir = tmp_path / "out"
    pipeline._write_outputs(out_dir, [[rec]], [rec], [rec], [], [], summary)

    assert not (out_dir / "dedup_candidates").exists()
    assert (out_dir / "similarity_clusters" / "cluster_001" / "REP_a.jpg").is_file()
    assert (out_dir / "final_selected" / "a.jpg").is_file()
    assert (out_dir / "ng_photos").is_dir()
    assert (out_dir / "other_passing").is_dir()
    assert (out_dir / "snap_pipeline_report.xlsx").is_file()

    wb = load_workbook(out_dir / "snap_pipeline_report.xlsx")
    assert wb.sheetnames == ["ベストショット一覧", "NG写真一覧", "未選定写真一覧", "統計サマリー"]
    assert [c.value for c in wb["ベストショット一覧"][1]] == [
        "ファイル名",
        "区分",
        "総合スコア",
        "技術スコア",
        "表情/明るさスコア",
        "構図スコア",
        "希少性スコア",
        "類似グループID",
        "撮影日時",
        "コメント",
        "選定理由",
    ]
    for sheet_name in ["ベストショット一覧", "NG写真一覧", "未選定写真一覧"]:
        headers = [c.value for c in wb[sheet_name][1]]
        assert "コメント" in headers
    summary_rows = {row[0].value: row[1].value for row in wb["統計サマリー"].iter_rows(min_row=2)}
    assert summary_rows["代表候補数"] == 1
    assert summary_rows["重複削減率"] == 0
    assert wb["ベストショット一覧"][2][1].value == "ベストショット"
    assert wb["ベストショット一覧"][2][9].value == "総合評価が高く、ベストショット候補として選定されました。"


def test_club_excel_labels_japanese(tmp_path: Path):
    pytest.importorskip("cv2")
    from backend.club_pipeline import run_club_pipeline

    root = tmp_path / "in"
    club = root / "soccer"
    club.mkdir(parents=True)
    Image.new("RGB", (32, 32), color=(180, 180, 180)).save(club / "x.jpg")

    z = tmp_path / "club.zip"
    with zipfile.ZipFile(z, "w") as zf:
        zf.write(club / "x.jpg", arcname="soccer/x.jpg")

    out = run_club_pipeline(str(z), str(tmp_path / "out"))
    wb = load_workbook(out["excel_path"])

    assert wb.sheetnames == ["サマリー", "目つぶり確認サマリー", "顔検出詳細", "ベストショット順位", "リネーム結果", "NG写真・要確認"]
    assert [c.value for c in wb["リネーム結果"][1]] == ["部活動名", "元ファイル名", "リネーム後ファイル名", "撮影日", "順位"]


def test_individual_error_log_csv_localized(tmp_path: Path):
    pytest.importorskip("mediapipe")
    from backend.individual.real_pipeline_source import write_error_log_csv
    out_csv = tmp_path / "error_log.csv"
    write_error_log_csv(
        [
            {
                "error_type": "no_card_detected",
                "severity": "warning",
                "detection_unit": "layer1",
                "class_id": "3A",
                "group_key": "g1",
                "group_idx": 1,
                "student_number": 5,
                "image_path": "a.jpg",
                "group_keys": ["g1", "g2"],
                "related_paths": ["a.jpg", "b.jpg"],
                "message": "no clear card visible",
            }
        ],
        out_csv,
    )
    text = out_csv.read_text(encoding="utf-8-sig")
    assert "エラー種別" in text
    assert "重要度" in text
    assert "検出レイヤー" in text
    assert "札検出なし" in text
    assert "警告" in text
    assert "明確な札が見えません" in text


def test_teacher_label_mapping_localized():
    assert TEACHER_SHEET_LABELS["Match Results"] == "照合結果"
    assert TEACHER_SHEET_LABELS["Best Shot Scores"] == "ベストショットスコア"
    assert excel_label("Original Filename") == "元ファイル名"
    assert excel_label("Card Name (OCR)") == "札氏名（OCR）"
    assert translate_display_value("matched") == "照合済み"


def test_club_ranking_sheet_localized_values(tmp_path: Path):
    pytest.importorskip("cv2")
    from backend.club_pipeline import run_club_pipeline

    root = tmp_path / "in2"
    club = root / "basket"
    club.mkdir(parents=True)
    Image.new("RGB", (32, 32), color=(180, 180, 180)).save(club / "x.jpg")

    z = tmp_path / "club2.zip"
    with zipfile.ZipFile(z, "w") as zf:
        zf.write(club / "x.jpg", arcname="basket/x.jpg")

    out = run_club_pipeline(str(z), str(tmp_path / "out2"))
    wb = load_workbook(out["excel_path"])
    ws = wb["ベストショット順位"]
    headers = [c.value for c in ws[1]]
    assert "画質・見栄え" in headers
    assert "表情スコア" in headers
    assert "雰囲気スコア" in headers
    assert "人数スコア" in headers
    assert "ポーズ減点" in headers
    assert "コメント" in headers
    assert "NG判定" in headers
    assert "NG理由" in headers

    ng_col = headers.index("NG判定") + 1
    ng_value = ws.cell(row=2, column=ng_col).value
    assert ng_value in {"はい", "いいえ"}
